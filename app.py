from datetime import datetime, date, time, timedelta
from calendar import monthrange
from html import escape
from io import BytesIO
from pathlib import Path
from threading import Thread
import base64
import json
import os
import re
import secrets
import smtplib
import ssl
import traceback
import urllib.error
import urllib.request
import certifi
from email.utils import parseaddr
from email.message import EmailMessage

from flask import Flask, render_template_string, redirect, url_for, request, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont
try:
    from apscheduler.schedulers.background import BackgroundScheduler
except ImportError:
    BackgroundScheduler = None

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
UPLOAD_DIR = STATIC_DIR / "uploads"
CARD_DIR = STATIC_DIR / "cards"
LOGO_PATH = STATIC_DIR / "logo.png"
MEMBER_CARD_TEMPLATE_PATH = STATIC_DIR / "member_card_template.png"
for folder in [STATIC_DIR, UPLOAD_DIR, CARD_DIR]:
    folder.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-only")
database_url = os.getenv("DATABASE_URL", "sqlite:///fitness.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024

# Email : remplis ces variables dans ton terminal pour envoyer de vrais emails.
# export SMTP_HOST="brevo.com"
# export SMTP_PORT="465"
# export SMTP_USER="tonadresse@smtp-brevo.com"
# export SMTP_PASSWORD="SMTP key"
# export MAIL_FROM="Section Fitness <tonadresse@gmail.com>"
SMTP_HOST = os.getenv("SMTP_HOST")
try:
    SMTP_PORT = int(str(os.getenv("SMTP_PORT", "465")).strip().strip('"').strip("'"))
except ValueError:
    SMTP_PORT = 465
try:
    SMTP_TIMEOUT = int(str(os.getenv("SMTP_TIMEOUT", "10")).strip().strip('"').strip("'"))
except ValueError:
    SMTP_TIMEOUT = 10
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
MAIL_FROM = os.getenv("MAIL_FROM", SMTP_USER or "section-fitness@local")
BREVO_API_KEY = os.getenv("BREVO_API_KEY")
LAST_DAILY_TASK_FILE = BASE_DIR / ".last_daily_fitness_tasks"
SCHEMA_READY = False


db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="adherent")
    status = db.Column(db.String(20), nullable=False, default="autre")
    blocked_until = db.Column(db.Date, nullable=True)

    # Nouveaux champs adhérent
    full_name = db.Column(db.String(150), nullable=True)
    first_name = db.Column(db.String(80), nullable=True)
    last_name = db.Column(db.String(80), nullable=True)
    profile_photo = db.Column(db.String(255), nullable=True)
    profile_photo_data = db.Column(db.Text, nullable=True)
    profile_photo_mime = db.Column(db.String(80), nullable=True)
    subscription_type = db.Column(db.String(50), nullable=True)
    subscription_year = db.Column(db.Integer, nullable=True)
    member_profile = db.Column(db.String(30), nullable=True)
    rights_holder_name = db.Column(db.String(150), nullable=True)
    member_number = db.Column(db.String(30), nullable=True)
    member_card = db.Column(db.String(255), nullable=True)
    blocked_at = db.Column(db.Date, nullable=True)
    blocked_reason = db.Column(db.String(255), nullable=True)
    preferred_course = db.Column(db.String(100), nullable=True)
    preferred_coach = db.Column(db.String(150), nullable=True)
    preferred_slot = db.Column(db.String(80), nullable=True)

    # Gestion avancée comptes / bureau / abonnements
    admin_role = db.Column(db.String(50), nullable=True)
    account_status = db.Column(db.String(30), nullable=False, default="active")
    activation_token = db.Column(db.String(255), nullable=True)
    activation_expires_at = db.Column(db.DateTime, nullable=True)
    subscription_end_date = db.Column(db.Date, nullable=True)
    archived_at = db.Column(db.Date, nullable=True)
    archived_reason = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    coach_type = db.Column(db.String(30), nullable=False, default="titulaire")

    def set_password(self, password):
        self.password_hash = generate_password_hash(password, method="pbkdf2:sha256")

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def is_blocked(self):
        return self.blocked_until and self.blocked_until >= date.today()

    def display_name(self):
        joined = f"{self.first_name or ''} {self.last_name or ''}".strip()
        return joined or self.full_name or self.email


class CourseSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_date = db.Column(db.Date, nullable=False)
    start_time = db.Column(db.Time, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    course_name = db.Column(db.String(100), nullable=False)
    capacity = db.Column(db.Integer, nullable=False)
    booking_open_date = db.Column(db.Date, nullable=False)
    priority_until = db.Column(db.Date, nullable=True)
    coach_name = db.Column(db.String(150), nullable=True)
    is_reservable = db.Column(db.Boolean, nullable=False, default=True)
    waitlist_capacity = db.Column(db.Integer, nullable=False, default=5)


class Booking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    session_id = db.Column(db.Integer, db.ForeignKey("course_session.id"), nullable=False)
    status = db.Column(db.String(30), nullable=False, default="booked")
    attendance_status = db.Column(db.String(30), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    archived = db.Column(db.Boolean, nullable=False, default=False)
    user = db.relationship("User", backref="bookings")
    session = db.relationship("CourseSession", backref="bookings")


class MembershipPeriod(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    subscription_type = db.Column(db.String(50), nullable=False)
    subscription_year = db.Column(db.Integer, nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    annual_fee_applies = db.Column(db.Boolean, nullable=False, default=False)
    subscription_price_snapshot = db.Column(db.Float, nullable=True)
    annual_fee_snapshot = db.Column(db.Float, nullable=True)
    total_snapshot = db.Column(db.Float, nullable=True)
    tariff_snapshot_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.String(150), nullable=True)
    notes = db.Column(db.String(500), nullable=True)
    user = db.relationship("User", backref="membership_periods")


class CourseTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    weekday = db.Column(db.Integer, nullable=False)
    week_parity = db.Column(db.String(10), nullable=False, default="all")
    course_name = db.Column(db.String(100), nullable=False)
    start_time = db.Column(db.Time, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    capacity = db.Column(db.Integer, nullable=False, default=35)
    waitlist_capacity = db.Column(db.Integer, nullable=False, default=5)
    coach_name = db.Column(db.String(150), nullable=True)
    is_reservable = db.Column(db.Boolean, nullable=False, default=True)
    active = db.Column(db.Boolean, nullable=False, default=True)


class CoachAbsence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    coach_name = db.Column(db.String(150), nullable=False)
    absence_date = db.Column(db.Date, nullable=False)
    session_id = db.Column(db.Integer, db.ForeignKey("course_session.id"), nullable=True)
    status = db.Column(db.String(30), nullable=False, default="absent")
    replacement_name = db.Column(db.String(150), nullable=True)
    notes = db.Column(db.String(500), nullable=True)
    followup_status = db.Column(db.String(30), nullable=False, default="a_traiter")
    admin_notes = db.Column(db.String(500), nullable=True)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    reviewed_by = db.Column(db.String(150), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    session = db.relationship("CourseSession")


class BudgetEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    entry_date = db.Column(db.Date, nullable=False, default=date.today)
    entry_type = db.Column(db.String(20), nullable=False)
    category = db.Column(db.String(80), nullable=False)
    label = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class InventoryItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    category = db.Column(db.String(80), nullable=True)
    quantity = db.Column(db.Integer, nullable=False, default=0)
    alert_threshold = db.Column(db.Integer, nullable=False, default=1)
    unit_cost = db.Column(db.Float, nullable=True)
    acquisition_year = db.Column(db.Integer, nullable=True)
    invoice_file = db.Column(db.String(255), nullable=True)
    purchase_request_file = db.Column(db.String(255), nullable=True)
    notes = db.Column(db.String(500), nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)


class UsefulDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    category = db.Column(db.String(80), nullable=True)
    file_path = db.Column(db.String(255), nullable=False)
    notes = db.Column(db.String(500), nullable=True)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    uploaded_by = db.Column(db.String(150), nullable=True)


class AppSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(120), unique=True, nullable=False)
    value = db.Column(db.String(500), nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def is_admin():
    return current_user.is_authenticated and current_user.role == "admin"


def is_coach_or_admin():
    return current_user.is_authenticated and current_user.role in ["admin", "coach"]


def next_url(default_endpoint="index"):
    target = request.args.get("next") or request.form.get("next")
    if target and target.startswith("/"):
        return target
    return url_for(default_endpoint)


def coach_display_names():
    coaches = [c.display_name() for c in User.query.filter_by(role="coach", coach_type="titulaire").order_by(User.full_name, User.email).all()]
    defaults = ["Hayate", "Malika", "Maud", "Mathieu", "Mélanie"]
    return sorted({name for name in coaches + defaults if name})


def configured_coach_names():
    user_coaches = [c.display_name() for c in User.query.filter_by(role="coach").order_by(User.full_name, User.email).all()]
    template_coaches = [row[0] for row in db.session.query(CourseTemplate.coach_name).filter(CourseTemplate.coach_name.isnot(None)).distinct().all()]
    replacement_coaches = get_replacement_coaches()
    return sorted({name for name in user_coaches + template_coaches + replacement_coaches if name})


def coach_type_for_name(name):
    user = User.query.filter(User.role == "coach").filter((User.full_name == name) | (User.email == name)).first()
    if user:
        return user.coach_type or "titulaire"
    return "remplacant" if name in get_replacement_coaches() else "titulaire"


def configured_coach_rows():
    rows = []
    for name in configured_coach_names():
        user = User.query.filter(User.role == "coach").filter((User.full_name == name) | (User.email == name)).first()
        rows.append({"name": name, "email": user.email if user else "", "coach_type": coach_type_for_name(name), "user_id": user.id if user else None})
    return rows


def coach_identity_names(user):
    names = {user.display_name(), user.email}
    if user.full_name:
        names.add(user.full_name)
    return {name for name in names if name}


def titular_coach_names():
    replacements = set(get_replacement_coaches())
    return sorted({name for name in configured_coach_names() if coach_type_for_name(name) == "titulaire" and name not in replacements})


def all_coach_users():
    return User.query.filter_by(role="coach").order_by(User.full_name, User.email).all()


def coach_user_for_name(name):
    if not name:
        return None
    for user in all_coach_users():
        if name in {user.email, user.full_name, user.display_name()}:
            return user
    return None


def valid_email(value):
    value = (value or "").strip().lower()
    return value if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value) else ""


SUBSCRIPTION_PRICES = {
    "Annuel": 140.0, "Semestre 1": 60.0, "Semestre 2": 60.0,
    "Trimestre 1": 35.0, "T2": 35.0, "T3": 35.0, "T4": 35.0,
}
DEFAULT_2026_PROFILE_PRICES = {
    "Annuel": {"ouvrant_droit": 140.0, "ayant_droit": 140.0, "exterieur": 280.0, "retraite": 280.0},
    "Semestre 1": {"ouvrant_droit": 60.0, "ayant_droit": 60.0, "exterieur": 120.0, "retraite": 120.0},
    "Semestre 2": {"ouvrant_droit": 45.0, "ayant_droit": 45.0, "exterieur": 90.0, "retraite": 90.0},
    "Trimestre 1": {"ouvrant_droit": 35.0, "ayant_droit": 35.0, "exterieur": 70.0, "retraite": 70.0},
    "T2": {"ouvrant_droit": 35.0, "ayant_droit": 35.0, "exterieur": 70.0, "retraite": 70.0},
    "T3": {"ouvrant_droit": 25.0, "ayant_droit": 25.0, "exterieur": 50.0, "retraite": 50.0},
    "T4": {"ouvrant_droit": 30.0, "ayant_droit": 30.0, "exterieur": 60.0, "retraite": 60.0},
}
SUBSCRIPTION_ALIASES = {
    "Trimestre 2": "T2",
    "Trimestre 3": "T3",
    "Trimestre 4": "T4",
}
TEMPORARY_BOOKING_GRACE_START = date(2026, 7, 1)
TEMPORARY_BOOKING_GRACE_END = date(2026, 7, 15)
DEMO_ADHERENT_EMAIL = "adherent@fitness.local"
DEMO_COACH_EMAIL = "coach@fitness.local"
DEFAULT_ANNUAL_MEMBERSHIP_FEE = 10.0
MEMBER_PROFILE_LABELS = {
    "ouvrant_droit": "Ouvrant droit",
    "ayant_droit": "Ayant droit",
    "exterieur": "Extérieur",
    "retraite": "Retraité",
}
MEMBER_PROFILE_RATES = {
    "ouvrant_droit": 0.5,
    "ayant_droit": 0.5,
    "exterieur": 1.0,
    "retraite": 1.0,
}


def parse_amount(value, default=0.0):
    try:
        return float(str(value or default).replace(",", "."))
    except ValueError:
        return default


def setting_value(key, default=""):
    setting = AppSetting.query.filter_by(key=key).first()
    return setting.value if setting and setting.value is not None else str(default)


def set_setting_value(key, value):
    setting = AppSetting.query.filter_by(key=key).first()
    if not setting:
        setting = AppSetting(key=key)
        db.session.add(setting)
    setting.value = str(value)
    setting.updated_at = datetime.utcnow()


def subscription_price_key(subscription_type):
    return f"subscription_price_{subscription_type.lower().replace(' ', '_')}"


def subscription_profile_price_key(subscription_type, member_profile):
    safe_subscription = normalize_subscription_type(subscription_type).lower().replace(" ", "_")
    return f"subscription_price_{safe_subscription}_{member_profile}"


def normalize_subscription_type(subscription_type):
    subscription_type = (subscription_type or "Annuel").strip()
    return SUBSCRIPTION_ALIASES.get(subscription_type, subscription_type)


def parse_iso_date(value, default):
    try:
        return datetime.strptime(value or "", "%Y-%m-%d").date()
    except ValueError:
        return default


def coach_planning_period(args):
    today = date.today()
    view_mode = (args.get("view_mode") or "").strip()
    explicit_start = (args.get("start_date") or "").strip()
    explicit_end = (args.get("end_date") or "").strip()
    if not view_mode:
        if explicit_start or explicit_end:
            view_mode = "range"
        elif args.get("year") or args.get("month"):
            view_mode = "month"
        else:
            view_mode = "rolling"
    try:
        year = int(args.get("year", today.year) or today.year)
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(args.get("month", today.month) or today.month)
    except (TypeError, ValueError):
        month = today.month
    if month < 1 or month > 12:
        month = today.month
    if view_mode == "month":
        start = date(year, month, 1)
        end = date(year, month, monthrange(year, month)[1])
    elif view_mode == "range":
        start = parse_iso_date(explicit_start, today)
        end = parse_iso_date(explicit_end, start + timedelta(days=30))
        if end < start:
            end = start
        year, month = start.year, start.month
    else:
        view_mode = "rolling"
        start = today
        end = today + timedelta(days=30)
        year, month = today.year, today.month
    return view_mode, start, end, year, month


def get_subscription_prices():
    return {
        name: parse_amount(setting_value(subscription_price_key(name), price), price)
        for name, price in SUBSCRIPTION_PRICES.items()
    }


def get_subscription_price_matrix():
    base_prices = get_subscription_prices()
    matrix = {}
    for subscription_type in SUBSCRIPTION_PRICES:
        matrix[subscription_type] = {}
        base_price = base_prices.get(subscription_type, SUBSCRIPTION_PRICES[subscription_type])
        for member_profile in MEMBER_PROFILE_LABELS:
            default_price = DEFAULT_2026_PROFILE_PRICES.get(subscription_type, {}).get(member_profile, base_price * member_profile_rate(member_profile))
            key = subscription_profile_price_key(subscription_type, member_profile)
            matrix[subscription_type][member_profile] = parse_amount(setting_value(key, default_price), default_price)
    return matrix


def get_annual_membership_fee():
    return parse_amount(setting_value("annual_membership_fee", DEFAULT_ANNUAL_MEMBERSHIP_FEE), DEFAULT_ANNUAL_MEMBERSHIP_FEE)


def seed_default_2026_tariffs_once():
    if setting_value("default_2026_tariffs_seeded", "") == "yes":
        return
    set_setting_value("annual_membership_fee", DEFAULT_ANNUAL_MEMBERSHIP_FEE)
    for subscription_type, profile_prices in DEFAULT_2026_PROFILE_PRICES.items():
        for member_profile, amount in profile_prices.items():
            set_setting_value(subscription_profile_price_key(subscription_type, member_profile), amount)
    set_setting_value("default_2026_tariffs_seeded", "yes")
    db.session.commit()


def membership_tariff_snapshot(user, subscription_type, annual_fee_applies=False):
    subscription_type = normalize_subscription_type(subscription_type)
    member_profile = user.member_profile or "ouvrant_droit"
    price_matrix = get_subscription_price_matrix()
    subscription_price = price_matrix.get(subscription_type, {}).get(member_profile, 0.0)
    annual_fee = get_annual_membership_fee() if annual_fee_applies else 0.0
    return subscription_price, annual_fee, subscription_price + annual_fee


def get_replacement_coaches():
    default_names = {"Mathieu", "Mélanie"}
    raw = setting_value("replacement_coaches", "")
    setting_names = {name.strip() for name in raw.split("\n") if name.strip()}
    user_names = {u.display_name() for u in User.query.filter_by(role="coach", coach_type="remplacant").order_by(User.full_name, User.email).all()}
    return sorted(default_names | setting_names | user_names)


def coach_replacement_options():
    return sorted(set(configured_coach_names()) | set(coach_display_names()) | set(get_replacement_coaches()))


def save_replacement_coaches(names):
    set_setting_value("replacement_coaches", "\n".join(sorted({name.strip() for name in names if name.strip()})))


def get_coach_planning_weekdays():
    raw = setting_value("coach_planning_weekdays", "")
    if raw:
        selected = {int(value) for value in raw.split(",") if value.strip().isdigit()}
        return sorted(day for day in selected if 0 <= day <= 6)
    template_days = {row[0] for row in db.session.query(CourseTemplate.weekday).filter_by(active=True).distinct().all()}
    return sorted(template_days) if template_days else [0, 1, 2, 3]


def set_coach_planning_weekdays(days):
    valid_days = sorted({int(day) for day in days if str(day).isdigit() and 0 <= int(day) <= 6})
    set_setting_value("coach_planning_weekdays", ",".join(str(day) for day in valid_days))


def member_profile_label(profile):
    return MEMBER_PROFILE_LABELS.get(profile or "ouvrant_droit", "Ouvrant droit")


def member_profile_rate(profile):
    return MEMBER_PROFILE_RATES.get(profile or "ouvrant_droit", 0.5)


def split_name(full_name):
    parts = (full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def form_full_name():
    first_name = request.form.get("first_name", "").strip()
    last_name = request.form.get("last_name", "").strip()
    if first_name or last_name:
        return first_name, last_name, f"{first_name} {last_name}".strip()
    full_name = request.form.get("full_name", "").strip()
    first_name, last_name = split_name(full_name)
    return first_name, last_name, full_name


def normalize_member_status(member_profile, status):
    if member_profile in ["ayant_droit", "exterieur", "retraite"]:
        return "autre"
    return status if status in ["mensuel", "cadre", "autre"] else "autre"


def first_registration_fee_applies(user, year):
    if user.created_at:
        return user.created_at.year == year
    return user.subscription_year == year


def period_budget_line(period, seen_user_years=None):
    seen_user_years = seen_user_years if seen_user_years is not None else set()
    user = period.user
    subscription_type = normalize_subscription_type(period.subscription_type)
    user_year_key = (period.user_id, period.subscription_year)
    first_period = MembershipPeriod.query.filter_by(
        user_id=period.user_id,
        subscription_year=period.subscription_year,
    ).order_by(MembershipPeriod.start_date, MembershipPeriod.id).first()
    is_first_period_for_year = first_period.id == period.id if first_period else user_year_key not in seen_user_years
    seen_user_years.add(user_year_key)
    note_text = (period.notes or "").lower()
    is_renewal = "renouvellement" in note_text and "période précédente" not in note_text
    annual_fee_should_apply = bool(period.annual_fee_applies and is_first_period_for_year and not is_renewal)
    subscription_price = period.subscription_price_snapshot
    if subscription_price is None:
        subscription_price, fallback_fee, _ = membership_tariff_snapshot(user, subscription_type, annual_fee_should_apply)
        annual_fee = fallback_fee if annual_fee_should_apply else 0.0
    else:
        annual_fee = (period.annual_fee_snapshot or 0.0) if annual_fee_should_apply else 0.0
    total = (subscription_price or 0.0) + (annual_fee or 0.0)
    return {
        "period": period,
        "user": user,
        "subscription_type": subscription_type,
        "subscription_year": period.subscription_year,
        "member_profile_label": member_profile_label(user.member_profile or "ouvrant_droit"),
        "profile_rate": 0,
        "base_subscription_price": subscription_price or 0.0,
        "subscription_price": subscription_price or 0.0,
        "annual_fee": annual_fee or 0.0,
        "total": total,
        "annual_fee_applies": annual_fee_should_apply,
    }


def membership_period_rows(periods):
    seen_user_years = set()
    return [period_budget_line(period, seen_user_years) for period in periods]


def safe_display(value, default="-"):
    try:
        text = str(value or "").strip()
        return text or default
    except Exception:
        return default


def budget_entry_rows_plain(entries):
    rows = []
    for entry in entries:
        try:
            rows.append({
                "entry_date": entry.entry_date.strftime("%d/%m/%Y") if entry.entry_date else "-",
                "entry_type": "income" if entry.entry_type == "income" else "expense",
                "entry_type_label": "Recette" if entry.entry_type == "income" else "Dépense",
                "category": safe_display(entry.category),
                "label": safe_display(entry.label),
                "amount": float(entry.amount or 0),
                "notes": safe_display(entry.notes, ""),
            })
        except Exception:
            print("\n--- LIGNE BUDGET IGNORÉE ---")
            traceback.print_exc()
            print("----------------------------\n")
    return rows


def budget_due_rows_plain(year=None):
    rows = []
    for row in expected_dues_rows(year):
        try:
            user = row.get("user")
            display_name = user.display_name() if user else "-"
            email = user.email if user else "-"
            rows.append({
                "name": safe_display(display_name),
                "email": safe_display(email),
                "member_profile_label": safe_display(row.get("member_profile_label")),
                "subscription_type": safe_display(row.get("subscription_type")),
                "subscription_year": safe_display(row.get("subscription_year"), ""),
                "subscription_price": float(row.get("subscription_price") or 0),
                "annual_fee": float(row.get("annual_fee") or 0),
                "total": float(row.get("total") or 0),
                "note": "Première inscription annuelle" if row.get("annual_fee") else "Pas de nouvelle cotisation",
            })
        except Exception:
            print("\n--- LIGNE COTISATION IGNORÉE ---")
            traceback.print_exc()
            print("--------------------------------\n")
    return rows


def expected_dues_rows(year=None):
    year = int(year or date.today().year)
    rows = []
    periods = MembershipPeriod.query.join(User).filter(
        User.role.in_(["adherent", "admin"]),
        User.account_status != "archived",
        MembershipPeriod.subscription_year == year,
    ).order_by(User.full_name, User.email, MembershipPeriod.start_date, MembershipPeriod.id).all()
    period_user_ids = {p.user_id for p in periods}
    rows.extend(membership_period_rows(periods))
    users = active_member_query().filter(User.subscription_year == year, ~User.id.in_(period_user_ids or {0})).order_by(User.full_name, User.email).all()
    for user in users:
        member_profile = user.member_profile or "ouvrant_droit"
        subscription_type = normalize_subscription_type(user.subscription_type)
        first_fee_applies = first_registration_fee_applies(user, year)
        subscription_price, first_fee, total = membership_tariff_snapshot(user, subscription_type, first_fee_applies)
        rows.append({"user": user, "subscription_type": subscription_type, "subscription_year": user.subscription_year, "member_profile_label": member_profile_label(member_profile), "profile_rate": 0, "base_subscription_price": subscription_price, "subscription_price": subscription_price, "annual_fee": first_fee, "total": total})
    return rows


def subscription_end(subscription_type, year):
    year = int(year or date.today().year)
    subscription_type = normalize_subscription_type(subscription_type)
    mapping = {
        "Annuel": date(year, 12, 31), "Semestre 1": date(year, 6, 30), "Semestre 2": date(year, 12, 31),
        "Trimestre 1": date(year, 3, 31), "T2": date(year, 6, 30), "T3": date(year, 9, 30), "T4": date(year, 12, 31),
    }
    return mapping.get(subscription_type, date(year, 12, 31))


def subscription_start(subscription_type, year):
    year = int(year or date.today().year)
    subscription_type = normalize_subscription_type(subscription_type)
    mapping = {
        "Annuel": date(year, 1, 1), "Semestre 1": date(year, 1, 1), "Semestre 2": date(year, 7, 1),
        "Trimestre 1": date(year, 1, 1), "T2": date(year, 4, 1), "T3": date(year, 7, 1), "T4": date(year, 10, 1),
    }
    return mapping.get(subscription_type, date(year, 1, 1))


def subscription_range(subscription_type, year):
    return subscription_start(subscription_type, year), subscription_end(subscription_type, year)


def user_can_book_session(user, session):
    if user.role not in ["adherent", "admin"]:
        return False, "Seuls les adhérents peuvent réserver."
    if TEMPORARY_BOOKING_GRACE_START <= session.course_date <= TEMPORARY_BOOKING_GRACE_END:
        return True, ""
    if not user.subscription_type or not user.subscription_year:
        return False, "Votre abonnement n'est pas renseigné. Contactez la Section Fitness."
    start, end = subscription_range(user.subscription_type, user.subscription_year)
    if not (start <= session.course_date <= end):
        return False, f"Votre abonnement {user.subscription_type} {user.subscription_year} permet de réserver uniquement du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}."
    return True, ""


def create_membership_period(user, subscription_type, subscription_year, annual_fee_applies=False, created_by=None, notes=None):
    subscription_type = normalize_subscription_type(subscription_type)
    start, end = subscription_range(subscription_type, subscription_year)
    existing = MembershipPeriod.query.filter_by(
        user_id=user.id,
        subscription_type=subscription_type,
        subscription_year=subscription_year,
        start_date=start,
        end_date=end,
    ).first()
    if existing:
        old_annual_fee_applies = existing.annual_fee_applies
        existing.annual_fee_applies = existing.annual_fee_applies or annual_fee_applies
        existing.created_by = existing.created_by or created_by
        existing.notes = existing.notes or notes
        if existing.subscription_price_snapshot is None or existing.annual_fee_snapshot is None or existing.total_snapshot is None or (not old_annual_fee_applies and existing.annual_fee_applies):
            subscription_price, annual_fee, total = membership_tariff_snapshot(user, subscription_type, existing.annual_fee_applies)
            existing.subscription_price_snapshot = subscription_price
            existing.annual_fee_snapshot = annual_fee
            existing.total_snapshot = total
            existing.tariff_snapshot_at = existing.tariff_snapshot_at or datetime.utcnow()
        return existing
    subscription_price, annual_fee, total = membership_tariff_snapshot(user, subscription_type, annual_fee_applies)
    period = MembershipPeriod(
        user_id=user.id,
        subscription_type=subscription_type,
        subscription_year=subscription_year,
        start_date=start,
        end_date=end,
        annual_fee_applies=annual_fee_applies,
        subscription_price_snapshot=subscription_price,
        annual_fee_snapshot=annual_fee,
        total_snapshot=total,
        tariff_snapshot_at=datetime.utcnow(),
        created_by=created_by,
        notes=notes,
    )
    db.session.add(period)
    return period


def ensure_current_membership_period_before_change(user, new_subscription_type, new_subscription_year, created_by=None):
    old_subscription_type = normalize_subscription_type(user.subscription_type)
    old_subscription_year = int(user.subscription_year or new_subscription_year or date.today().year)
    new_subscription_type = normalize_subscription_type(new_subscription_type)
    new_subscription_year = int(new_subscription_year or old_subscription_year)
    if not old_subscription_type or old_subscription_type not in SUBSCRIPTION_PRICES:
        return None
    if old_subscription_type == new_subscription_type and old_subscription_year == new_subscription_year:
        return None
    existing = MembershipPeriod.query.filter_by(
        user_id=user.id,
        subscription_type=old_subscription_type,
        subscription_year=old_subscription_year,
    ).first()
    if existing:
        return existing
    annual_fee_applies = not MembershipPeriod.query.filter_by(user_id=user.id, subscription_year=old_subscription_year).first()
    return create_membership_period(
        user,
        old_subscription_type,
        old_subscription_year,
        annual_fee_applies=annual_fee_applies,
        created_by=created_by,
        notes="Période précédente conservée avant nouvelle adhésion",
    )


def inferred_previous_subscription(subscription_type):
    subscription_type = normalize_subscription_type(subscription_type)
    previous = {
        "Semestre 2": "Semestre 1",
        "T2": "Trimestre 1",
        "T3": "T2",
        "T4": "T3",
    }
    return previous.get(subscription_type)


def repair_missing_prior_membership_periods(user):
    repaired = False
    periods = MembershipPeriod.query.filter_by(user_id=user.id).order_by(MembershipPeriod.subscription_year, MembershipPeriod.start_date, MembershipPeriod.id).all()
    for period in periods:
        if "renouvellement" not in (period.notes or "").lower():
            continue
        prior_type = inferred_previous_subscription(period.subscription_type)
        if not prior_type:
            continue
        prior_start, prior_end = subscription_range(prior_type, period.subscription_year)
        if prior_end >= period.start_date:
            continue
        existing_prior = MembershipPeriod.query.filter_by(
            user_id=user.id,
            subscription_type=prior_type,
            subscription_year=period.subscription_year,
            start_date=prior_start,
            end_date=prior_end,
        ).first()
        has_any_prior = MembershipPeriod.query.filter(
            MembershipPeriod.user_id == user.id,
            MembershipPeriod.subscription_year == period.subscription_year,
            MembershipPeriod.end_date < period.start_date,
        ).first()
        if existing_prior or has_any_prior:
            continue
        create_membership_period(
            user,
            prior_type,
            period.subscription_year,
            annual_fee_applies=True,
            created_by=period.created_by,
            notes="Période précédente reconstituée avant nouvelle adhésion",
        )
        repaired = True
    return repaired


def recent_membership_actions(limit=40):
    return membership_actions_query().limit(limit).all()


def membership_actions_query(start_date=None, end_date=None, year=None):
    query = MembershipPeriod.query.join(User).filter(
        User.role.in_(["adherent", "admin"])
    )
    if year:
        query = query.filter(MembershipPeriod.subscription_year == int(year))
    if start_date:
        query = query.filter(MembershipPeriod.created_at >= datetime.combine(start_date, time.min))
    if end_date:
        query = query.filter(MembershipPeriod.created_at <= datetime.combine(end_date, time.max))
    return query.order_by(MembershipPeriod.created_at.desc(), MembershipPeriod.id.desc())


def delete_static_file(relative_path):
    if not relative_path:
        return
    try:
        path = (STATIC_DIR / relative_path).resolve()
        if STATIC_DIR.resolve() in path.parents and path.exists():
            path.unlink()
    except OSError:
        pass


def delete_member_completely(user):
    """Supprime un adhérent et toutes les données directement rattachées."""
    profile_photo = user.profile_photo
    member_card = user.member_card
    Booking.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    MembershipPeriod.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    db.session.delete(user)
    db.session.commit()
    delete_static_file(profile_photo)
    delete_static_file(member_card)


def make_activation_token(user):
    user.activation_token = secrets.token_urlsafe(32)
    user.activation_expires_at = datetime.utcnow() + timedelta(days=14)
    return user.activation_token


def send_activation_email(user):
    token = make_activation_token(user)
    db.session.commit()
    link = url_for("activate_account", token=token, _external=True)
    role_label = "coach" if user.role == "coach" else "adhérent" if user.role == "adherent" else "admin"
    return send_email(user.email, "Activation de votre compte Section Fitness", f"Bonjour {user.display_name()},\n\nVotre compte {role_label} Section Fitness a été pré-enregistré. Merci de créer votre mot de passe avec ce lien :\n{link}\n\nCe lien est valable 14 jours.\n\nSection Fitness")


def create_activation_link(user):
    token = make_activation_token(user)
    db.session.commit()
    return url_for("activate_account", token=token, _external=True)


def send_password_reset_email(user):
    token = make_activation_token(user)
    db.session.commit()
    link = url_for("reset_password", token=token, _external=True)
    return send_email(user.email, "Réinitialisation de votre mot de passe Section Fitness", f"Bonjour {user.display_name()},\n\nUne demande de réinitialisation de mot de passe a été faite pour votre compte Section Fitness.\n\nCréez un nouveau mot de passe avec ce lien :\n{link}\n\nCe lien est valable 14 jours.\n\nSi vous n'êtes pas à l'origine de cette demande, ignorez cet email.\n\nSection Fitness")


def archive_expired_memberships():
    today = date.today()
    if today <= TEMPORARY_BOOKING_GRACE_END:
        return 0
    users = User.query.filter(User.role == "adherent", User.email != DEMO_ADHERENT_EMAIL, User.account_status != "archived", User.subscription_end_date.isnot(None), User.subscription_end_date < today).all()
    for user in users:
        has_future_booking = Booking.query.join(CourseSession).filter(
            Booking.user_id == user.id,
            Booking.status.in_(["booked", "waiting_list"]),
            CourseSession.course_date >= today,
        ).first()
        if has_future_booking:
            continue
        user.account_status = "archived"
        user.archived_at = today
        user.archived_reason = f"Abonnement expiré le {user.subscription_end_date.strftime('%d/%m/%Y')}"
    if users:
        db.session.commit()
    return len(users)


def active_adherent_query():
    return User.query.filter(User.role == "adherent", User.email != DEMO_ADHERENT_EMAIL, User.account_status != "archived")


def active_member_query():
    return User.query.filter(
        User.account_status != "archived",
        User.email != DEMO_ADHERENT_EMAIL,
        db.or_(
            User.role == "adherent",
            db.and_(User.role == "admin", User.subscription_type.isnot(None), User.subscription_year.isnot(None)),
        ),
    )


def is_member_account(user):
    return bool(user and user.account_status != "archived" and user.email != DEMO_ADHERENT_EMAIL and (
        user.role == "adherent" or (
            user.role == "admin" and user.subscription_type is not None and user.subscription_year is not None
        )
    ))


def send_member_campaign_async(user_ids, subject, signed_body, signed_html, inline_images):
    def worker():
        with app.app_context():
            sent_count = 0
            failed_count = 0
            users = User.query.filter(User.id.in_(user_ids), User.account_status != "archived").order_by(User.role, User.full_name, User.email).all()
            for user in users:
                if send_email(user.email, subject, signed_body, html_body=signed_html, inline_images=inline_images):
                    sent_count += 1
                else:
                    failed_count += 1
            print(f"\n--- CAMPAGNE EMAIL TERMINÉE ---\nEnvoyés: {sent_count}\nÉchecs: {failed_count}\n-------------------------------\n")

    Thread(target=worker, daemon=True).start()


def notify_admins_of_coach_absence(coach_name, start_date, end_date, status, replacement_name="", notes=""):
    admins = User.query.filter_by(role="admin").all()
    if not admins:
        return 0
    status_labels = {"absent": "absence", "conge": "congé", "replaced": "remplacement", "present": "présence"}
    period = start_date.strftime("%d/%m/%Y")
    if end_date != start_date:
        period = f"du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
    details = [
        f"Coach : {coach_name}",
        f"Type : {status_labels.get(status, status)}",
        f"Période : {period}",
    ]
    if replacement_name:
        details.append(f"Remplaçant : {replacement_name}")
    if notes:
        details.append(f"Notes : {notes}")
    planning_link = url_for("admin_coach_planning", year=start_date.year, month=start_date.month, _external=True)
    body = "Bonjour,\n\nUne absence ou un congé vient d'être déclaré depuis le profil coach.\n\n"
    body += "\n".join(details)
    body += f"\n\nVoir le planning coachs :\n{planning_link}\n\nSection Fitness"
    sent = 0
    for admin in admins:
        if send_email(admin.email, "Absence / congé coach déclaré", body):
            sent += 1
    return sent


def notify_members_of_coach_absence(coach_name, start_date, end_date, status, replacement_name="", notes=""):
    if status not in ["absent", "conge"]:
        return 0
    sessions = CourseSession.query.filter(
        CourseSession.coach_name == coach_name,
        CourseSession.course_date >= start_date,
        CourseSession.course_date <= end_date,
    ).order_by(CourseSession.course_date, CourseSession.start_time).all()
    sent = 0
    for session in sessions:
        bookings = Booking.query.filter(
            Booking.session_id == session.id,
            Booking.status.in_(["booked", "waiting_list"]),
        ).all()
        for booking in bookings:
            body = (
                f"Bonjour {booking.user.display_name()},\n\n"
                f"Le coach {coach_name} a déclaré une absence pour le cours {session.course_name} "
                f"du {session.course_date.strftime('%d/%m/%Y')} à {session.start_time.strftime('%H:%M')}.\n"
            )
            if replacement_name:
                body += f"\nRemplaçant prévu : {replacement_name}.\n"
            if notes:
                body += f"\nInformation complémentaire : {notes}\n"
            body += "\nSection Fitness"
            if send_email(booking.user.email, "Information cours - absence coach", body):
                sent += 1
    return sent


def notify_coach_absence_validated(absence):
    user = coach_user_for_name(absence.coach_name)
    if not user:
        return False
    body = (
        f"Bonjour {user.display_name()},\n\n"
        f"Votre demande d'absence/congé du {absence.absence_date.strftime('%d/%m/%Y')} "
        f"({absence_session_label(absence)}) a été validée par le Bureau Fitness.\n"
    )
    if absence.replacement_name:
        body += f"\nRemplaçant identifié : {absence.replacement_name}.\n"
    body += "\nSection Fitness"
    return send_email(user.email, "Demande d'absence validée", body)


def notify_replacement_assigned(absence):
    user = coach_user_for_name(absence.replacement_name)
    if not user:
        return False
    body = (
        f"Bonjour {user.display_name()},\n\n"
        f"Vous êtes indiqué(e) comme remplaçant(e) pour {absence.coach_name} "
        f"le {absence.absence_date.strftime('%d/%m/%Y')} ({absence_session_label(absence)}).\n\n"
        "Ce cours apparaît désormais dans votre planning coach.\n\nSection Fitness"
    )
    return send_email(user.email, "Remplacement confirmé - Section Fitness", body)


def notify_coaches_replacement_needed(absence):
    subject = "Remplacement à trouver - Section Fitness"
    body = (
        "Bonjour,\n\n"
        f"Un remplacement est recherché pour {absence.coach_name} "
        f"le {absence.absence_date.strftime('%d/%m/%Y')} ({absence_session_label(absence)}).\n\n"
        "Merci de répondre via vos canaux habituels (adresse professionnelle / groupe WhatsApp) si vous pouvez assurer ce créneau.\n\n"
        "Section Fitness"
    )
    sent = 0
    for user in all_coach_users():
        if user.display_name() == absence.coach_name:
            continue
        if send_email(user.email, subject, body):
            sent += 1
    return sent


def admin_email_signature_body(body):
    return body.rstrip() + "\n\nSportivement,\nBureau Fitness,"


def admin_email_signature_html(body):
    escaped_body = escape(body).replace("\n", "<br>")
    logo_html = '<br><br><img src="cid:fitness_logo" alt="Bureau Fitness" style="width:120px;height:auto;">' if LOGO_PATH.exists() else ""
    return f"""<!doctype html><html><body style="font-family:Arial,sans-serif;font-size:15px;line-height:1.45;color:#111827">{escaped_body}<br><br>Sportivement,<br>Bureau Fitness,{logo_html}</body></html>"""


def mail_sender_payload():
    name, email = parseaddr(MAIL_FROM)
    if not email:
        email = MAIL_FROM
        name = "Section Fitness"
    payload = {"email": email}
    if name:
        payload["name"] = name
    return payload


def mail_sender_email():
    return mail_sender_payload()["email"]


def send_email_brevo_api(to, subject, body, attachments=None, html_body=None):
    attachments = attachments or []
    payload = {
        "sender": mail_sender_payload(),
        "bcc": [{"email": to}],
        "subject": subject,
        "textContent": body,
    }
    if html_body:
        payload["htmlContent"] = html_body

    api_attachments = []
    for attachment_path in attachments:
        path = Path(attachment_path)
        if path.exists():
            api_attachments.append({
                "name": path.name,
                "content": base64.b64encode(path.read_bytes()).decode("ascii"),
            })
    if api_attachments:
        payload["attachment"] = api_attachments

    request_data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email",
        data=request_data,
        headers={
            "accept": "application/json",
            "api-key": BREVO_API_KEY,
            "content-type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=SMTP_TIMEOUT, context=ssl.create_default_context(cafile=certifi.where())) as response:
        return 200 <= response.status < 300


def send_email(to, subject, body, attachments=None, html_body=None, inline_images=None):
    try:
        attachments = attachments or []
        inline_images = inline_images or {}
        if BREVO_API_KEY:
            return send_email_brevo_api(to, subject, body, attachments=attachments, html_body=html_body)

        if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD):
            print("\n--- EMAIL NON ENVOYÉ : SMTP NON CONFIGURÉ ---")
            print("Cci:", to)
            print("Subject:", subject)
            print(body)
            if html_body:
                print("HTML:", html_body)
            for cid, image_path in inline_images.items():
                print("Inline image:", cid, image_path)
            for attachment_path in attachments:
                print("Attachment:", attachment_path)
            print("--------------------------------------------\n")
            return False

        msg = EmailMessage()
        msg["From"] = MAIL_FROM
        msg["Bcc"] = to
        msg["Subject"] = subject
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype="html")
            html_part = msg.get_payload()[-1]
            for cid, image_path in inline_images.items():
                path = Path(image_path)
                if path.exists():
                    html_part.add_related(path.read_bytes(), maintype="image", subtype=path.suffix.lstrip(".").lower() or "png", cid=f"<{cid}>")

        for attachment_path in attachments:
            path = Path(attachment_path)
            if path.exists():
                msg.add_attachment(path.read_bytes(), maintype="image", subtype="png", filename=path.name)

        context = ssl.create_default_context(cafile=certifi.where())

        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT, context=context) as server:
                server.login(SMTP_USER, SMTP_PASSWORD)
                server.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT) as server:
                server.starttls(context=context)
                server.login(SMTP_USER, SMTP_PASSWORD)
                server.send_message(msg)

        return True

    except Exception as e:
        print("\n--- ERREUR ENVOI EMAIL ---")
        print(e)
        if isinstance(e, urllib.error.HTTPError):
            print(e.read().decode("utf-8", errors="replace"))
        print("--------------------------\n")
        return False


def allowed_image(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"png", "jpg", "jpeg"}


def save_profile_photo(file, user_id):
    if not file or not file.filename:
        return None
    if not allowed_image(file.filename):
        raise ValueError("Format photo non accepté. Utilisez JPG ou PNG.")
    ext = secure_filename(file.filename).rsplit(".", 1)[1].lower()
    filename = f"profile_{user_id}.{ext}"
    path = UPLOAD_DIR / filename
    file.save(path)
    return f"uploads/{filename}"


def persist_profile_photo(user, file):
    if not file or not file.filename:
        return None
    if not allowed_image(file.filename):
        raise ValueError("Format photo non accepté. Utilisez JPG ou PNG.")
    ext = secure_filename(file.filename).rsplit(".", 1)[1].lower()
    content = file.read()
    filename = f"profile_{user.id}.{ext}"
    path = UPLOAD_DIR / filename
    path.write_bytes(content)
    user.profile_photo = f"uploads/{filename}"
    user.profile_photo_data = base64.b64encode(content).decode("ascii")
    user.profile_photo_mime = "image/png" if ext == "png" else "image/jpeg"
    return user.profile_photo


def user_profile_photo_bytes(user):
    if user.profile_photo_data:
        try:
            return base64.b64decode(user.profile_photo_data), user.profile_photo_mime or "image/jpeg"
        except Exception:
            pass
    if user.profile_photo:
        path = STATIC_DIR / user.profile_photo
        if path.exists():
            mime = "image/png" if path.suffix.lower() == ".png" else "image/jpeg"
            return path.read_bytes(), mime
    return None, None


def user_profile_photo_image(user):
    data, _ = user_profile_photo_bytes(user)
    if not data:
        return None
    try:
        return Image.open(BytesIO(data)).convert("RGB")
    except Exception:
        return None


def allowed_document(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xls", "xlsx"}


def save_inventory_document(file, prefix):
    if not file or not file.filename:
        return None
    if not allowed_document(file.filename):
        raise ValueError("Format non accepté. Utilisez PDF, image, Word ou Excel.")
    original = secure_filename(file.filename)
    filename = f"{prefix}_{secrets.token_hex(8)}_{original}"
    path = UPLOAD_DIR / filename
    file.save(path)
    return f"uploads/{filename}"


def save_useful_document(file):
    if not file or not file.filename:
        raise ValueError("Merci de choisir un document à téléverser.")
    if not allowed_document(file.filename):
        raise ValueError("Format non accepté. Utilisez PDF, image, Word ou Excel.")
    original = secure_filename(file.filename)
    filename = f"infos_utiles_{secrets.token_hex(8)}_{original}"
    path = UPLOAD_DIR / filename
    file.save(path)
    return f"uploads/{filename}"


def get_font(size, bold=False):
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold else "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial Bold.ttf" if bold else "/Library/Fonts/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def next_member_number(year):
    count = User.query.filter(User.member_number.isnot(None)).count() + 1
    return f"FS-{year}-{count:04d}"


def fit_image_cover(image, size):
    target_w, target_h = size
    img_w, img_h = image.size
    scale = max(target_w / img_w, target_h / img_h)
    new_size = (int(img_w * scale), int(img_h * scale))
    image = image.resize(new_size)
    left = (image.width - target_w) // 2
    top = (image.height - target_h) // 2
    return image.crop((left, top, left + target_w, top + target_h))


def draw_text_fit(draw, position, text, font_size, max_width, fill, bold=True, min_size=24):
    text = str(text or "-").upper()
    size = font_size
    font = get_font(size, bold)
    while size > min_size and draw.textbbox((0, 0), text, font=font)[2] > max_width:
        size -= 2
        font = get_font(size, bold)
    draw.text(position, text, font=font, fill=fill)
    return font


def add_left_card_gradient(card):
    width, height = card.size
    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    pixels = overlay.load()
    for x in range(width):
        if x < 780:
            alpha = 255
        elif x < 1080:
            alpha = int(255 * (1 - ((x - 780) / 300)))
        else:
            alpha = 0
        for y in range(48, height - 48):
            pixels[x, y] = (0, 0, 0, max(0, min(255, alpha)))
    return Image.alpha_composite(card.convert("RGBA"), overlay)


def generate_member_card(user):
    width, height = 1528, 998
    green = (88, 126, 55)
    white = (255, 255, 255)
    panel_dark = (6, 13, 12)

    if MEMBER_CARD_TEMPLATE_PATH.exists():
        card = Image.open(MEMBER_CARD_TEMPLATE_PATH).convert("RGBA").resize((width, height))
    else:
        card = Image.new("RGBA", (width, height), (10, 16, 15, 255))
        fallback_draw = ImageDraw.Draw(card)
        for i in range(-250, width, 28):
            fallback_draw.arc((i, 120, i + 1200, height + 260), 205, 345, fill=(34, 62, 36), width=2)

    draw = ImageDraw.Draw(card)

    photo_x, photo_y, photo_w, photo_h = 83, 93, 310, 435
    draw.rounded_rectangle((photo_x - 3, photo_y - 3, photo_x + photo_w + 3, photo_y + photo_h + 3), radius=31, fill=green)
    draw.rounded_rectangle((photo_x, photo_y, photo_x + photo_w, photo_y + photo_h), radius=28, fill=(232, 232, 232))
    photo = user_profile_photo_image(user)
    if photo:
        photo = fit_image_cover(photo, (photo_w, photo_h))
        mask = Image.new("L", (photo_w, photo_h), 0)
        ImageDraw.Draw(mask).rounded_rectangle((0, 0, photo_w, photo_h), radius=28, fill=255)
        card.paste(photo, (photo_x, photo_y), mask)
    else:
        placeholder_font = get_font(28, True)
        draw.text((photo_x + 68, photo_y + 185), "PHOTO", font=placeholder_font, fill=(120, 120, 120))

    draw.rectangle((78, 630, 760, 744), fill=panel_dark)
    draw_text_fit(draw, (82, 635), user.display_name(), 78, 660, white, bold=True, min_size=38)

    draw.rectangle((272, 860, 548, 930), fill=panel_dark)
    draw_text_fit(draw, (276, 868), user.subscription_type or "-", 48, 250, white, bold=True, min_size=26)

    draw.rectangle((775, 860, 970, 930), fill=panel_dark)
    draw_text_fit(draw, (778, 868), str(user.subscription_year or date.today().year), 48, 180, white, bold=True, min_size=30)

    filename = f"carte_adherent_{user.id}.png"
    path = CARD_DIR / filename
    card.convert("RGB").save(path)
    user.member_card = f"cards/{filename}"
    db.session.commit()
    return path


def monday_midday_priority_applies(session):
    return (
        session.course_date.weekday() == 0
        and session.start_time in [time(11, 45), time(12, 45)]
        and session.priority_until
        and date.today() <= session.priority_until
    )


def priority_until_label(session):
    return session.priority_until.strftime("%d/%m/%Y") if session and session.priority_until else ""


def booked_count(session):
    return Booking.query.filter_by(session_id=session.id, status="booked").count()


def waiting_list_count(session):
    return Booking.query.filter_by(session_id=session.id, status="waiting_list").count()


def waitlist_rank(booking):
    if booking.status != "waiting_list":
        return None
    waiting = Booking.query.filter_by(session_id=booking.session_id, status="waiting_list").order_by(Booking.created_at, Booking.id).all()
    for index, item in enumerate(waiting, start=1):
        if item.id == booking.id:
            return index
    return None


def waitlist_capacity(session):
    return session.waitlist_capacity if session.waitlist_capacity is not None else 5


def attendance_label(booking):
    if booking.attendance_status == "present":
        return "Présent"
    if booking.attendance_status == "late":
        return "Retard"
    if booking.attendance_status == "skipped":
        return "À revoir"
    if booking.attendance_status == "absent" or booking.status == "absent_unexcused":
        return "Absent"
    if booking.status == "waiting_list":
        return "Liste d'attente"
    return "Non pointé"


def attendance_badge_class(booking):
    if booking.attendance_status == "present":
        return ""
    if booking.attendance_status == "late":
        return "wait"
    if booking.attendance_status == "skipped":
        return "wait"
    if booking.attendance_status == "absent" or booking.status == "absent_unexcused":
        return "full"
    if booking.status == "waiting_list":
        return "wait"
    return "wait"


def absence_display_label(absence):
    if not absence:
        return ""
    if absence.followup_status == "annule" or absence.status == "cancelled":
        return "cours annulé"
    if absence.followup_status == "remplacement_a_trouver":
        return "remplacement à trouver"
    if absence.followup_status == "remplacement_trouve" or absence.status == "replaced":
        return "remplacé"
    if absence.followup_status == "valide":
        return "validé"
    labels = {"absent": "absent", "conge": "congé", "present": "présent"}
    return labels.get(absence.status, absence.status)


def absence_badge_class(absence):
    if not absence:
        return ""
    if absence.followup_status == "annule" or absence.status == "cancelled":
        return "full"
    if absence.followup_status == "remplacement_trouve" or absence.status in ["present", "replaced"]:
        return ""
    if absence.followup_status in ["remplacement_a_trouver", "en_cours", "valide"]:
        return "wait"
    if absence.status in ["absent", "conge"]:
        return "full"
    return "wait"


def absence_blocks_booking(absence):
    if not absence:
        return False
    if absence.followup_status == "remplacement_trouve" or absence.status in ["present", "replaced"]:
        return False
    return absence.status in ["absent", "conge", "cancelled"] or absence.followup_status in ["annule", "remplacement_a_trouver", "en_cours", "valide"]


def absence_for_session(abs_by_key, session):
    if not session:
        return None
    return abs_by_key.get((session.coach_name, session.course_date, session.id)) or abs_by_key.get((session.coach_name, session.course_date, None))


def replacement_is_confirmed(absence):
    return bool(absence and absence.replacement_name and absence.followup_status in ["remplacement_trouve", "valide"])


def effective_coach_for_session(session, abs_by_key):
    absence = absence_for_session(abs_by_key, session)
    if replacement_is_confirmed(absence):
        return absence.replacement_name
    return session.coach_name or "-"


def coach_monthly_invoice_rows(start, end, coach_filter=None):
    sessions = CourseSession.query.filter(CourseSession.course_date >= start, CourseSession.course_date <= end).order_by(CourseSession.course_date, CourseSession.start_time).all()
    absences = CoachAbsence.query.filter(CoachAbsence.absence_date >= start, CoachAbsence.absence_date <= end).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in absences}
    rows = {}
    for session in sessions:
        absence = absence_for_session(abs_by_key, session)
        effective = effective_coach_for_session(session, abs_by_key)
        if coach_filter and effective != coach_filter and session.coach_name != coach_filter:
            continue
        row = rows.setdefault(effective, {"coach": effective, "cours": 0, "remplacements": 0, "absences": 0, "annules": 0, "details": []})
        detail_status = "effectué"
        if absence:
            label = absence_display_label(absence)
            if replacement_is_confirmed(absence):
                row["remplacements"] += 1
                detail_status = f"remplacement de {absence.coach_name}"
            elif absence.followup_status in ["annule"] or absence.status == "cancelled":
                row["annules"] += 1
                detail_status = "annulé"
            elif absence.status in ["absent", "conge"]:
                rows.setdefault(absence.coach_name, {"coach": absence.coach_name, "cours": 0, "remplacements": 0, "absences": 0, "annules": 0, "details": []})["absences"] += 1
                detail_status = label
        if detail_status not in ["annulé", "absent", "congé", "remplacement à trouver"]:
            row["cours"] += 1
        row["details"].append({
            "date": session.course_date,
            "horaire": f"{session.start_time.strftime('%H:%M')} - {session.end_time.strftime('%H:%M')}",
            "cours": session.course_name,
            "statut": detail_status,
            "coach_initial": session.coach_name or "-",
        })
    return sorted(rows.values(), key=lambda item: item["coach"])


def coach_invoice_detail_rows(start, end, coach_filter=None):
    sessions = CourseSession.query.filter(
        CourseSession.course_date >= start,
        CourseSession.course_date <= end,
    ).order_by(CourseSession.course_date, CourseSession.start_time, CourseSession.coach_name).all()
    absences = CoachAbsence.query.filter(
        CoachAbsence.absence_date >= start,
        CoachAbsence.absence_date <= end,
    ).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in absences}
    rows = []

    def add_row(coach, session, statut, absence=None, coach_initial=None, replacement_name=None):
        if not coach or coach == "-":
            return
        if coach_filter and coach != coach_filter:
            return
        duration_hours = max(
            0,
            (
                datetime.combine(session.course_date, session.end_time)
                - datetime.combine(session.course_date, session.start_time)
            ).total_seconds() / 3600,
        )
        billed_hours = 1.5 if 0 < duration_hours <= 1 else duration_hours
        rows.append({
            "date": session.course_date,
            "jour": WEEKDAY_LABELS[session.course_date.weekday()],
            "coach": coach,
            "statut": statut,
            "coach_initial": coach_initial or session.coach_name or "",
            "remplacant": replacement_name or "",
            "cours": session.course_name,
            "horaire": f"{session.start_time.strftime('%H:%M')} - {session.end_time.strftime('%H:%M')}",
            "start_time": session.start_time,
            "end_time": session.end_time,
            "duration_hours": duration_hours,
            "duration_label": format_duration_hours(duration_hours),
            "billed_hours": billed_hours,
            "billed_label": format_duration_hours(billed_hours),
            "suivi_admin": absence.followup_status if absence else "",
            "notes_admin": absence.admin_notes if absence and absence.admin_notes else "",
            "notes_coach": absence.notes if absence and absence.notes else "",
        })

    for session in sessions:
        absence = absence_for_session(abs_by_key, session)
        original_coach = session.coach_name or "-"
        if absence and replacement_is_confirmed(absence):
            add_row(absence.coach_name, session, "Absence remplacée", absence=absence, coach_initial=original_coach, replacement_name=absence.replacement_name)
            add_row(absence.replacement_name, session, "Remplacement effectué", absence=absence, coach_initial=original_coach, replacement_name=absence.replacement_name)
        elif absence and (absence.followup_status == "annule" or absence.status == "cancelled"):
            add_row(original_coach, session, "Cours annulé", absence=absence, coach_initial=original_coach, replacement_name=absence.replacement_name)
        elif absence and absence.status in ["absent", "conge"]:
            add_row(original_coach, session, absence_display_label(absence), absence=absence, coach_initial=original_coach, replacement_name=absence.replacement_name)
        else:
            add_row(original_coach, session, "Cours effectué", absence=absence, coach_initial=original_coach)
    return sorted(rows, key=lambda item: (item["coach"], item["date"], item["horaire"], item["cours"]))


def format_time_compact(value):
    return value.strftime("%Hh%M").replace("h00", "h")


def format_duration_hours(hours):
    total_minutes = int(round((hours or 0) * 60))
    whole_hours, minutes = divmod(total_minutes, 60)
    if minutes:
        return f"{whole_hours}h{minutes:02d}"
    return f"{whole_hours}h"


def coach_invoice_summary_rows(start, end, coach_filter=None):
    grouped = {}
    for row in coach_invoice_detail_rows(start, end, coach_filter=coach_filter):
        coach = row["coach"]
        item = grouped.setdefault(coach, {"coach": coach, "items": [], "total_hours": 0.0})
        label = (
            f"{row['date'].strftime('%d/%m')} de "
            f"{format_time_compact(row['start_time'])} à {format_time_compact(row['end_time'])} "
            f"({format_duration_hours(row['duration_hours'])}) - {row['cours']} - {row['statut']}"
        )
        if row.get("coach_initial") and row["coach_initial"] != coach:
            label += f" (coach initial : {row['coach_initial']})"
        if row.get("remplacant") and row["remplacant"] != coach:
            label += f" (remplaçant : {row['remplacant']})"
        notes = []
        if row.get("notes_admin"):
            notes.append(f"note admin : {row['notes_admin']}")
        if row.get("notes_coach"):
            notes.append(f"note coach : {row['notes_coach']}")
        if notes:
            label += f" [{'; '.join(notes)}]"
        item["items"].append(label)
        item["total_hours"] += row["duration_hours"] or 0
    rows = []
    for item in grouped.values():
        rows.append({
            "coach": item["coach"],
            "details": " ; ".join(item["items"]),
            "total_hours": item["total_hours"],
            "total_label": format_duration_hours(item["total_hours"]),
        })
    return sorted(rows, key=lambda item: item["coach"])


def absence_session_label(absence):
    if absence and absence.session:
        session = absence.session
        return f"{session.start_time.strftime('%H:%M')} - {session.end_time.strftime('%H:%M')} · {session.course_name}"
    return "Toute la journée"


def absence_session_options(absence):
    if not absence:
        return []
    return absence_target_sessions(absence.coach_name, absence.absence_date)


def absence_target_sessions(coach_name, day):
    return CourseSession.query.filter_by(
        coach_name=coach_name,
        course_date=day,
    ).order_by(CourseSession.start_time).all()


def upsert_coach_absence(coach_name, day, status, replacement, notes, session=None, reset_followup=False):
    existing = CoachAbsence.query.filter_by(
        coach_name=coach_name,
        absence_date=day,
        session_id=session.id if session else None,
    ).first()
    if not existing:
        existing = CoachAbsence(coach_name=coach_name, absence_date=day, session_id=session.id if session else None)
        db.session.add(existing)
    existing.status = status
    existing.replacement_name = replacement
    existing.notes = notes
    if reset_followup:
        existing.followup_status = "a_traiter"
        existing.admin_notes = None
        existing.reviewed_at = None
        existing.reviewed_by = None
    return existing


@app.context_processor
def template_helpers():
    return {
        "absence_badge_class": absence_badge_class,
        "absence_blocks_booking": absence_blocks_booking,
        "absence_display_label": absence_display_label,
        "absence_for_session": absence_for_session,
        "absence_session_label": absence_session_label,
        "absence_session_options": absence_session_options,
        "attendance_badge_class": attendance_badge_class,
        "attendance_label": attendance_label,
        "booked_count": booked_count,
        "effective_coach_for_session": effective_coach_for_session,
        "replacement_is_confirmed": replacement_is_confirmed,
        "user_can_book_session": user_can_book_session,
        "waitlist_rank": waitlist_rank,
        "split_name": split_name,
    }


def user_has_active_booking(user_id, session_id):
    return Booking.query.filter(
        Booking.user_id == user_id,
        Booking.session_id == session_id,
        Booking.status.in_(["booked", "waiting_list"])
    ).first()


def create_booking_for_user(user, session, by_admin=False):
    # Test manuel liste d'attente :
    # 1. Créer un cours avec jauge 1 et liste d'attente 5.
    # 2. Réserver avec un premier adhérent : statut booked.
    # 3. Réserver avec un deuxième adhérent : statut waiting_list, rang 1.
    # 4. Annuler la réservation du premier adhérent.
    # 5. Vérifier que le deuxième passe en booked et que les emails/flash sont clairs.
    if user_has_active_booking(user.id, session.id):
        return None, "duplicate"
    if booked_count(session) < session.capacity:
        booking = Booking(user_id=user.id, session_id=session.id, status="booked")
        db.session.add(booking)
        db.session.commit()
        subject = "Réservation confirmée" + (" par la Section Fitness" if by_admin else "")
        admin_note = " par l'administration" if by_admin else ""
        send_email(user.email, subject, f"Bonjour {user.display_name()},\n\nVotre réservation au cours {session.course_name} du {session.course_date.strftime('%d/%m/%Y')} à {session.start_time.strftime('%H:%M')} est confirmée{admin_note}.\n\nSection Fitness")
        return booking, "booked"
    if waiting_list_count(session) >= waitlist_capacity(session):
        return None, "waitlist_full"
    booking = Booking(user_id=user.id, session_id=session.id, status="waiting_list")
    db.session.add(booking)
    db.session.commit()
    rank = waitlist_rank(booking)
    subject = "Inscription en liste d’attente" + (" par la Section Fitness" if by_admin else "")
    admin_note = " par l'administration" if by_admin else ""
    send_email(user.email, subject, f"Bonjour {user.display_name()},\n\nLe cours {session.course_name} du {session.course_date.strftime('%d/%m/%Y')} à {session.start_time.strftime('%H:%M')} est complet : vous êtes en liste d'attente{admin_note}, mais vous n'avez pas encore de place confirmée.\n\nVotre rang actuel : {rank}.\n\nVous recevrez un email si une place se libère et que votre réservation est confirmée.\n\nSection Fitness")
    return booking, "waiting_list"


def cancel_booking_and_promote(booking, cancelled_by_admin=False):
    session = booking.session
    booking.status = "cancelled"
    db.session.commit()
    if cancelled_by_admin:
        send_email(booking.user.email, "Réservation annulée par la Section Fitness", f"Bonjour {booking.user.display_name()},\n\nVotre réservation au cours {session.course_name} du {session.course_date.strftime('%d/%m/%Y')} a été annulée par l'administration.\n\nSection Fitness")
    else:
        send_email(booking.user.email, "Annulation confirmée", f"Bonjour {booking.user.display_name()},\n\nVotre réservation au cours {session.course_name} du {session.course_date.strftime('%d/%m/%Y')} est annulée.\n\nSection Fitness")

    next_waiting = Booking.query.filter_by(session_id=session.id, status="waiting_list").order_by(Booking.created_at, Booking.id).first()
    if next_waiting and booked_count(session) < session.capacity:
        next_waiting.status = "booked"
        db.session.commit()
        send_email(next_waiting.user.email, "Réservation confirmée - place libérée", f"Bonjour {next_waiting.user.display_name()},\n\nUne place s'est libérée pour {session.course_name} du {session.course_date.strftime('%d/%m/%Y')} à {session.start_time.strftime('%H:%M')}.\n\nVotre réservation est maintenant confirmée.\n\nSection Fitness")
        return next_waiting
    return None


def absence_count(user):
    three_months_ago = date.today() - timedelta(days=90)
    return Booking.query.join(CourseSession).filter(
        Booking.user_id == user.id,
        Booking.status == "absent_unexcused",
        CourseSession.course_date >= three_months_ago
    ).count()


def session_slot_label(session):
    day = WEEKDAY_LABELS[session.course_date.weekday()]
    hour = session.start_time.hour
    if hour < 12:
        moment = "matin"
    elif hour < 14:
        moment = "midi"
    else:
        moment = "soir"
    return f"{day} {moment}"


def preference_options():
    sessions = CourseSession.query.order_by(CourseSession.course_name, CourseSession.coach_name, CourseSession.course_date).all()
    courses = sorted({s.course_name for s in sessions if s.course_name} | {"Pilates"})
    coaches = sorted({s.coach_name for s in sessions if s.coach_name} | {"Hayate"})
    slots = sorted({session_slot_label(s) for s in sessions} | {"Lundi midi"})
    return {"courses": courses, "coaches": coaches, "slots": slots}


def preference_stats():
    users = active_member_query().all()

    def rows(attr):
        counts = {}
        for user in users:
            value = getattr(user, attr, None)
            if value:
                counts[value] = counts.get(value, 0) + 1
        return sorted(counts.items(), key=lambda item: (-item[1], item[0]))

    return {
        "course": rows("preferred_course"),
        "coach": rows("preferred_coach"),
        "slot": rows("preferred_slot"),
    }


def section_admin_stats():
    users = active_member_query().all()
    annual = {}
    subscriptions = {}
    profiles = {}
    status_counts = {}
    for user in users:
        year = user.subscription_year or (user.created_at.year if user.created_at else date.today().year)
        annual[year] = annual.get(year, 0) + 1
        if user.subscription_type:
            subscriptions[user.subscription_type] = subscriptions.get(user.subscription_type, 0) + 1
        profile = member_profile_label(user.member_profile)
        profiles[profile] = profiles.get(profile, 0) + 1
        status_counts[user.status or "autre"] = status_counts.get(user.status or "autre", 0) + 1

    annual_rows = []
    previous = None
    for year in sorted(annual):
        count = annual[year]
        evolution = None if previous in (None, 0) else ((count - previous) / previous) * 100
        annual_rows.append({"year": year, "count": count, "evolution": evolution})
        previous = count

    return {
        "annual": annual_rows,
        "subscriptions": sorted(subscriptions.items(), key=lambda item: (-item[1], item[0])),
        "profiles": sorted(profiles.items(), key=lambda item: (-item[1], item[0])),
        "statuses": sorted(status_counts.items(), key=lambda item: (-item[1], item[0])),
    }


def course_booking_stats(start, end, course_filter=""):
    query = db.session.query(CourseSession).filter(
        CourseSession.course_date >= start,
        CourseSession.course_date <= end,
    )
    if course_filter:
        query = query.filter(CourseSession.course_name == course_filter)
    sessions = query.order_by(CourseSession.course_date, CourseSession.start_time).all()
    rows = []
    monthly = {}
    for session in sessions:
        bookings = Booking.query.join(User).filter(
            Booking.session_id == session.id,
            Booking.status.in_(["booked", "waiting_list", "absent_unexcused"]),
        ).all()
        booked = [b for b in bookings if b.status == "booked"]
        waiting = [b for b in bookings if b.status == "waiting_list"]
        absent = [b for b in bookings if b.status == "absent_unexcused" or b.attendance_status == "absent"]
        mensuels = sum(1 for b in booked if b.user.status == "mensuel")
        cadres_autres = len(booked) - mensuels
        ratio_mensuel = (mensuels / len(booked) * 100) if booked else 0
        fill_rate = (len(booked) / session.capacity * 100) if session.capacity else 0
        rows.append({
            "date": session.course_date,
            "month": session.course_date.strftime("%Y-%m"),
            "course": session.course_name,
            "coach": session.coach_name or "-",
            "time": f"{session.start_time.strftime('%H:%M')} - {session.end_time.strftime('%H:%M')}",
            "capacity": session.capacity,
            "booked": len(booked),
            "waiting": len(waiting),
            "absent": len(absent),
            "mensuels": mensuels,
            "cadres_autres": cadres_autres,
            "ratio_mensuel": ratio_mensuel,
            "fill_rate": fill_rate,
        })
        month = monthly.setdefault(session.course_date.strftime("%Y-%m"), {"month": session.course_date.strftime("%Y-%m"), "sessions": 0, "booked": 0, "mensuels": 0, "waiting": 0})
        month["sessions"] += 1
        month["booked"] += len(booked)
        month["mensuels"] += mensuels
        month["waiting"] += len(waiting)
    monthly_rows = []
    for item in sorted(monthly.values(), key=lambda row: row["month"]):
        item["ratio_mensuel"] = (item["mensuels"] / item["booked"] * 100) if item["booked"] else 0
        monthly_rows.append(item)
    return rows, monthly_rows


def course_name_options():
    return [row[0] for row in db.session.query(CourseSession.course_name).distinct().order_by(CourseSession.course_name).all()]


def apply_absence_sanction(user):
    absences = absence_count(user)
    if absences >= 2:
        user.blocked_until = date.today() + timedelta(days=30)
        user.blocked_at = date.today()
        user.blocked_reason = f"Blocage automatique : {absences} absences non excusées sur 90 jours."
        db.session.commit()
        send_email(
            user.email,
            "Blocage temporaire de votre compte Fitness",
            f"Bonjour {user.display_name()},\n\nVotre compte est temporairement bloqué jusqu'au {user.blocked_until} en raison de {absences} absences non excusées sur les 90 derniers jours.\n\nSection Fitness"
        )


def refresh_absence_block_status(user):
    absences = absence_count(user)
    if absences < 2 and user.blocked_reason and user.blocked_reason.startswith("Blocage automatique"):
        user.blocked_until = None
        user.blocked_at = None
        user.blocked_reason = None


def planned_sessions_for_day(day):
    """Retourne les cours théoriques à créer pour une date donnée, à partir des paramètres admin."""
    iso_week = day.isocalendar().week
    parity = "even" if iso_week % 2 == 0 else "odd"
    templates = CourseTemplate.query.filter_by(weekday=day.weekday(), active=True).all()
    sessions_to_create = []
    for tpl in templates:
        if tpl.week_parity in ("all", parity):
            sessions_to_create.append((tpl.course_name, tpl.start_time, tpl.end_time, tpl.capacity, tpl.coach_name, tpl.is_reservable, tpl.waitlist_capacity))
    return sessions_to_create


def create_session_if_missing(day, course_name, start, end, capacity, booking_open_date, coach_name=None, is_reservable=True, waitlist_capacity_value=5):
    existing = CourseSession.query.filter_by(
        course_date=day,
        start_time=start,
        end_time=end,
        course_name=course_name
    ).first()

    if existing:
        return False

    db.session.add(CourseSession(
        course_date=day,
        start_time=start,
        end_time=end,
        course_name=course_name,
        capacity=capacity,
        booking_open_date=booking_open_date,
        coach_name=coach_name,
        is_reservable=is_reservable,
        waitlist_capacity=waitlist_capacity_value,
        # La priorité mensuel est propre à chaque créneau nouvellement créé.
        priority_until=booking_open_date + timedelta(days=6)
    ))
    return True


def generate_month_sessions(year, month):
    """Génération manuelle par l'admin, conservée en secours.
    Chaque créneau créé reçoit une priorité de 7 jours à partir du jour de génération.
    """
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    current = first_day
    created = 0
    booking_open_date = date.today()

    while current <= last_day:
        for item in planned_sessions_for_day(current):
            course_name, start, end, capacity, coach_name, is_reservable, waitlist_capacity_value = item
            if create_session_if_missing(current, course_name, start, end, capacity, booking_open_date, coach_name, is_reservable, waitlist_capacity_value):
                created += 1
        current += timedelta(days=1)

    db.session.commit()
    return created


def generate_rolling_sessions(days_ahead=28):
    """Génère automatiquement les créneaux manquants entre aujourd'hui et J+28.
    Exemple : chaque jour, le site ajoute uniquement les nouveaux créneaux nécessaires pour
    maintenir un planning ouvert sur 28 jours glissants.
    """
    today = date.today()
    end_date = today + timedelta(days=days_ahead)
    current = today
    created = 0

    while current <= end_date:
        for item in planned_sessions_for_day(current):
            course_name, start, end, capacity, coach_name, is_reservable, waitlist_capacity_value = item
            if create_session_if_missing(current, course_name, start, end, capacity, today, coach_name, is_reservable, waitlist_capacity_value):
                created += 1
        current += timedelta(days=1)

    db.session.commit()
    return created

def archive_past_bookings():
    """Archive les réservations liées à des séances déjà passées.
    L’archivage ne supprime rien : les données restent disponibles dans les exports et les fiches de cours.
    """
    today = date.today()
    past_bookings = Booking.query.join(CourseSession).filter(
        CourseSession.course_date < today,
        Booking.archived.is_(False)
    ).all()
    for booking in past_bookings:
        booking.archived = True
    if past_bookings:
        db.session.commit()
    return len(past_bookings)


def run_daily_automation(force=False):
    """Tâches automatiques lancées une fois par jour au premier accès au site.
    - archive les réservations des séances passées ;
    - maintient automatiquement un planning ouvert sur 1 mois glissant.
    """
    today = date.today()
    try:
        last_run = LAST_DAILY_TASK_FILE.read_text().strip() if LAST_DAILY_TASK_FILE.exists() else ""
    except OSError:
        last_run = ""
    if not force and last_run == today.isoformat():
        return

    archive_past_bookings()
    archive_expired_memberships()
    generate_rolling_sessions(days_ahead=28)

    # Nettoyage prudent : supprime uniquement les créneaux trop lointains sans réservation.
    end_date = today + timedelta(days=28)
    far_empty_sessions = CourseSession.query.filter(CourseSession.course_date > end_date).all()
    for session in far_empty_sessions:
        if not session.bookings:
            db.session.delete(session)
    db.session.commit()

    try:
        LAST_DAILY_TASK_FILE.write_text(today.isoformat())
    except OSError:
        pass


@app.before_request
def before_each_request():
    if request.endpoint not in {"static"}:
        global SCHEMA_READY
        if not SCHEMA_READY:
            ensure_schema()
            create_default_admin()
            seed_default_course_templates()
            SCHEMA_READY = True
        run_daily_automation()


@app.route("/")
@login_required
def index():
    today = date.today()
    end_date = today + timedelta(days=28)
    query = CourseSession.query.filter(
        CourseSession.course_date >= today,
        CourseSession.course_date <= end_date
    )
    selected_course = request.args.get("course_filter", "")
    selected_coach = request.args.get("coach_filter", "")
    selected_slot = request.args.get("slot_filter", "")
    if current_user.role in ["adherent", "admin"]:
        if selected_course:
            query = query.filter(CourseSession.course_name == selected_course)
        if selected_coach:
            query = query.filter(CourseSession.coach_name == selected_coach)
    sessions = query.order_by(CourseSession.course_date, CourseSession.start_time).all()
    if current_user.role in ["adherent", "admin"] and selected_slot:
        sessions = [session for session in sessions if session_slot_label(session) == selected_slot]
    absences = CoachAbsence.query.filter(
        CoachAbsence.absence_date >= today,
        CoachAbsence.absence_date <= end_date,
    ).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in absences}
    current_bookings = []
    active_booking_by_session = {}
    if current_user.role in ["adherent", "admin"]:
        current_bookings = Booking.query.join(CourseSession).filter(
            Booking.user_id == current_user.id,
            Booking.status.in_(["booked", "waiting_list"]),
            CourseSession.course_date >= today,
        ).order_by(CourseSession.course_date, CourseSession.start_time, Booking.created_at).all()
        active_booking_by_session = {booking.session_id: booking for booking in current_bookings}
    stats = {
        "today_sessions": CourseSession.query.filter_by(course_date=today).count(),
        "bookings": Booking.query.filter_by(status="booked", archived=False).count(),
        "members": active_member_query().count(),
        "blocked": User.query.filter(User.blocked_until >= today).count(),
    }
    latest_bookings = Booking.query.join(CourseSession).join(Booking.user).filter(
        User.role.in_(["adherent", "admin"])
    ).order_by(Booking.created_at.desc(), Booking.id.desc()).limit(12).all() if is_admin() else []
    return render_template_string(TEMPLATE_INDEX, sessions=sessions, booked_count=booked_count, waitlist_rank=waitlist_rank, stats=stats, latest_bookings=latest_bookings, preference_options=preference_options(), preference_stats=preference_stats(), section_stats=section_admin_stats(), selected_course=selected_course, selected_coach=selected_coach, selected_slot=selected_slot, abs_by_key=abs_by_key, current_bookings=current_bookings, active_booking_by_session=active_booking_by_session, temporary_booking_grace_start=TEMPORARY_BOOKING_GRACE_START, temporary_booking_grace_end=TEMPORARY_BOOKING_GRACE_END)


@app.route("/admin/statistics")
@login_required
def admin_statistics():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    today = date.today()
    default_start = date(today.year, today.month, 1)
    default_end = date(today.year, today.month, monthrange(today.year, today.month)[1])
    stats_start = parse_iso_date(request.args.get("start_date", ""), default_start)
    stats_end = parse_iso_date(request.args.get("end_date", ""), default_end)
    if stats_end < stats_start:
        stats_end = stats_start
    course_filter = request.args.get("course_filter", "").strip()
    course_rows, course_monthly_rows = course_booking_stats(stats_start, stats_end, course_filter)
    stats = {
        "today_sessions": CourseSession.query.filter_by(course_date=date.today()).count(),
        "bookings": Booking.query.filter(Booking.status.in_(["booked", "waiting_list"])).count(),
        "members": active_member_query().count(),
        "blocked": User.query.filter(User.blocked_until >= date.today()).count()
    }
    filter_values = {"start_date": stats_start.isoformat(), "end_date": stats_end.isoformat(), "course_filter": course_filter}
    return render_template_string(TEMPLATE_STATISTICS, stats=stats, preference_stats=preference_stats(), section_stats=section_admin_stats(), course_rows=course_rows, course_monthly_rows=course_monthly_rows, course_options=course_name_options(), filter_values=filter_values)


@app.route("/register", methods=["GET", "POST"])
def register():
    flash("La création de compte est réservée à l'administration. Utilisez le lien d'activation reçu par email.")
    return redirect(url_for("login"))
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        member_profile = request.form.get("member_profile", "ouvrant_droit")
        status = normalize_member_status(member_profile, request.form.get("status", "autre"))
        rights_holder_name = request.form.get("rights_holder_name", "").strip()
        first_name, last_name, full_name = form_full_name()
        if not first_name or not last_name:
            flash("Merci d'indiquer votre prénom et votre nom.")
            return redirect(url_for("register"))
        subscription_type = normalize_subscription_type(request.form["subscription_type"])
        subscription_year = int(request.form["subscription_year"])
        photo = request.files.get("profile_photo")

        if User.query.filter_by(email=email).first():
            flash("Compte déjà existant.")
            return redirect(url_for("register"))
        if not photo or not photo.filename:
            flash("Merci d'ajouter une photo de profil.")
            return redirect(url_for("register"))
        if member_profile == "ayant_droit" and not rights_holder_name:
            flash("Merci d'indiquer le nom et prénom de l'ouvrant droit.")
            return redirect(url_for("register"))

        user = User(email=email, role="adherent", status=status, full_name=full_name, first_name=first_name, last_name=last_name, member_profile=member_profile, rights_holder_name=rights_holder_name, subscription_type=subscription_type, subscription_year=subscription_year, subscription_end_date=subscription_end(subscription_type, subscription_year), account_status="active", member_number=next_member_number(subscription_year))
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        create_membership_period(user, subscription_type, subscription_year, annual_fee_applies=True, created_by=user.display_name(), notes="Création compte adhérent")
        db.session.commit()

        try:
            persist_profile_photo(user, photo)
            db.session.commit()
            card_path = generate_member_card(user)
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("register"))

        sent = send_email(
            user.email,
            "Bienvenue dans la Section Fitness - votre carte adhérent",
            f"Bonjour {user.display_name()},\n\nVotre compte adhérent a été créé avec succès. Votre carte adhérent est jointe à cet email.\n\nType d'abonnement : {user.subscription_type}\nAnnée : {user.subscription_year}\nID adhérent : {user.member_number}\n\nSection Fitness",
            attachments=[card_path]
        )
        flash("Compte créé. Email envoyé." if sent else "Compte créé. SMTP non configuré : l'email apparaît dans la console.")
        return redirect(url_for("login"))
    return render_template_string(TEMPLATE_REGISTER, current_year=date.today().year)


@app.route("/profile", methods=["GET", "POST"])
@login_required
def member_profile():
    if current_user.role not in ["adherent", "admin"]:
        flash("Accès réservé aux adhérents.")
        return redirect(url_for("index"))
    if request.method == "POST":
        profile = request.form.get("member_profile", current_user.member_profile or "ouvrant_droit")
        rights_holder_name = request.form.get("rights_holder_name", "").strip()
        if profile not in MEMBER_PROFILE_LABELS:
            profile = "ouvrant_droit"
        if profile == "ayant_droit" and not rights_holder_name:
            flash("Merci d'indiquer le nom et prénom de l'ouvrant droit.")
            return redirect(url_for("member_profile"))
        first_name, last_name, full_name = form_full_name()
        current_user.first_name = first_name
        current_user.last_name = last_name
        current_user.full_name = full_name
        current_user.member_profile = profile
        if current_user.role == "admin":
            current_user.status = normalize_member_status(profile, request.form.get("status", current_user.status or "autre"))
            current_user.subscription_type = normalize_subscription_type(request.form.get("subscription_type", current_user.subscription_type or "Annuel"))
            current_user.subscription_year = int(request.form.get("subscription_year") or current_user.subscription_year or date.today().year)
            current_user.subscription_end_date = subscription_end(current_user.subscription_type, current_user.subscription_year)
            if not current_user.member_number:
                current_user.member_number = next_member_number(current_user.subscription_year)
            create_membership_period(
                current_user,
                current_user.subscription_type,
                current_user.subscription_year,
                annual_fee_applies=not MembershipPeriod.query.filter_by(user_id=current_user.id, subscription_year=current_user.subscription_year).first(),
                created_by=current_user.display_name(),
                notes="Profil adhérent admin",
            )
        else:
            current_user.status = normalize_member_status(profile, current_user.status)
        current_user.rights_holder_name = rights_holder_name if profile == "ayant_droit" else None
        current_user.preferred_course = request.form.get("preferred_course", "").strip() or None
        current_user.preferred_coach = request.form.get("preferred_coach", "").strip() or None
        current_user.preferred_slot = request.form.get("preferred_slot", "").strip() or None
        photo = request.files.get("profile_photo")
        if photo and photo.filename:
            try:
                persist_profile_photo(current_user, photo)
            except ValueError as exc:
                flash(str(exc))
                return redirect(url_for("member_profile"))
            generate_member_card(current_user)
        db.session.commit()
        flash("Profil mis à jour.")
        return redirect(url_for("member_profile"))
    return render_template_string(TEMPLATE_MEMBER_PROFILE, preference_options=preference_options(), current_year=date.today().year)


@app.route("/planning-coachs")
@login_required
def member_coach_planning():
    if current_user.role not in ["adherent", "admin"]:
        flash("Accès réservé aux adhérents.")
        return redirect(url_for("index"))
    ensure_coach_absence_schema()
    view_mode, start, end, year, month = coach_planning_period(request.args)
    sessions = CourseSession.query.filter(
        CourseSession.course_date >= start,
        CourseSession.course_date <= end,
    ).order_by(CourseSession.course_date, CourseSession.start_time).all()
    absences = CoachAbsence.query.filter(
        CoachAbsence.absence_date >= start,
        CoachAbsence.absence_date <= end,
    ).order_by(CoachAbsence.absence_date, CoachAbsence.coach_name).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in absences}
    effective_names = {effective_coach_for_session(s, abs_by_key) for s in sessions if effective_coach_for_session(s, abs_by_key) != "-"}
    coach_names = sorted({name for name in effective_names if coach_type_for_name(name) == "titulaire"} | set(titular_coach_names()) | {name for name in effective_names if name in get_replacement_coaches()})
    planning_weekdays = set(get_coach_planning_weekdays())
    month_days = [start + timedelta(days=i) for i in range((end - start).days + 1) if (start + timedelta(days=i)).weekday() in planning_weekdays]
    coach_agenda = {}
    for session in sessions:
        display_coach = effective_coach_for_session(session, abs_by_key)
        if display_coach in coach_names:
            coach_agenda.setdefault((display_coach, session.course_date), []).append(session)
    active_bookings = Booking.query.join(CourseSession).filter(
        Booking.user_id == current_user.id,
        Booking.status.in_(["booked", "waiting_list"]),
        CourseSession.course_date >= start,
        CourseSession.course_date <= end,
    ).all()
    active_booking_by_session = {booking.session_id: booking for booking in active_bookings}
    return render_template_string(
        TEMPLATE_MEMBER_COACH_PLANNING,
        abs_by_key=abs_by_key,
        active_booking_by_session=active_booking_by_session,
        booked_count=booked_count,
        coach_agenda=coach_agenda,
        coach_names=coach_names,
        month_days=month_days,
        waitlist_rank=waitlist_rank,
        weekday_labels=WEEKDAY_LABELS,
        year=year,
        month=month,
        view_mode=view_mode,
        start=start,
        end=end,
        range_label=f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}",
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            if user.account_status == "pending":
                flash("Compte non activé : utilisez le lien reçu par email pour créer votre mot de passe.")
                return redirect(url_for("login"))
            if (
                user.role == "adherent"
                and user.account_status == "archived"
                and date.today() <= TEMPORARY_BOOKING_GRACE_END
                and user.subscription_type
                and user.subscription_year
            ):
                user.account_status = "active"
                user.archived_at = None
                user.archived_reason = None
                db.session.commit()
            if user.role == "adherent" and user.account_status == "archived":
                flash("Votre compte est archivé car votre abonnement est expiré. Contactez la Section Fitness pour renouveler votre abonnement.")
                return redirect(url_for("login"))
            login_user(user)
            return redirect(url_for("index"))
        flash("Identifiants incorrects.")
    return render_template_string(TEMPLATE_LOGIN)


@app.route("/coach", methods=["GET", "POST"])
def coach_login():
    flash("Les coachs se connectent maintenant depuis la page de connexion principale.")
    return redirect(url_for("login"))
    """Connexion réservée aux coachs préalablement créés par l'admin."""
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        user = User.query.filter_by(email=email, role="coach").first()
        if user and user.check_password(password):
            if user.account_status == "pending":
                flash("Compte coach non activé : utilisez le lien reçu par email pour créer votre mot de passe.")
                return redirect(url_for("coach_login"))
            login_user(user)
            flash("Connexion coach réussie.")
            return redirect(url_for("index"))
        flash("Identifiants coach incorrects ou compte non enregistré.")
    return render_template_string(TEMPLATE_COACH_LOGIN)


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = User.query.filter_by(email=email).first()
        if user:
            send_password_reset_email(user)
        flash("Si un compte existe avec cet email, un lien de réinitialisation vient d'être envoyé.")
        return redirect(url_for("login"))
    return render_template_string(TEMPLATE_FORGOT_PASSWORD)


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    user = User.query.filter_by(activation_token=token).first()
    if not user or not user.activation_expires_at or user.activation_expires_at < datetime.utcnow():
        flash("Lien invalide ou expiré.")
        return redirect(url_for("forgot_password"))
    if request.method == "POST":
        password = request.form.get("password", "")
        if len(password) < 8:
            flash("Le mot de passe doit contenir au moins 8 caractères.")
            return render_template_string(TEMPLATE_RESET_PASSWORD, user=user)
        user.set_password(password)
        user.activation_token = None
        user.activation_expires_at = None
        if user.account_status == "pending":
            user.account_status = "active"
        db.session.commit()
        flash("Mot de passe réinitialisé. Vous pouvez vous connecter.")
        return redirect(url_for("login"))
    return render_template_string(TEMPLATE_RESET_PASSWORD, user=user)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/book/<int:session_id>")
@login_required
def book(session_id):
    session = db.session.get(CourseSession, session_id) or CourseSession.query.get_or_404(session_id)
    redirect_target = next_url()
    if current_user.role not in ["adherent", "admin"]:
        flash("Seuls les adhérents peuvent réserver.")
        return redirect(redirect_target)
    if not session.is_reservable:
        flash("Ce cours ne nécessite pas de réservation.")
        return redirect(redirect_target)
    if current_user.is_blocked():
        flash(f"Vous êtes bloqué jusqu'au {current_user.blocked_until}.")
        return redirect(redirect_target)
    can_book, reason = user_can_book_session(current_user, session)
    if not can_book:
        flash(reason)
        return redirect(redirect_target)
    absence = absence_for_session(
        {(a.coach_name, a.absence_date, a.session_id): a for a in CoachAbsence.query.filter_by(absence_date=session.course_date).all()},
        session,
    )
    if absence_blocks_booking(absence):
        flash("Réservation indisponible : le créneau est marqué absent/congé.")
        return redirect(redirect_target)
    if monday_midday_priority_applies(session) and current_user.status != "mensuel":
        flash(f"Priorité réservée aux adhérents mensuels jusqu'au {priority_until_label(session)} inclus.")
        return redirect(redirect_target)

    booking, result = create_booking_for_user(current_user, session)
    if result == "duplicate":
        flash("Vous êtes déjà inscrit ou en liste d’attente.")
    elif result == "waitlist_full":
        flash("Liste d’attente complète.")
    elif result == "booked":
        flash("Réservation confirmée.")
    elif result == "waiting_list":
        flash(f"Cours complet. Vous êtes inscrit en liste d’attente — rang {waitlist_rank(booking)}.")
    return redirect(redirect_target)


@app.route("/cancel/<int:booking_id>")
@login_required
def cancel(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    redirect_target = next_url()
    if booking.user_id != current_user.id and not is_admin():
        flash("Action non autorisée.")
        return redirect(redirect_target)
    session_datetime = datetime.combine(booking.session.course_date, booking.session.start_time)
    if datetime.now() > session_datetime - timedelta(hours=2):
        flash("Annulation impossible à moins de 2h du cours.")
        return redirect(redirect_target)
    promoted = cancel_booking_and_promote(booking, cancelled_by_admin=is_admin() and booking.user_id != current_user.id)
    if promoted:
        flash(f"Réservation annulée. {promoted.user.display_name()} a été promu depuis la liste d’attente.")
    else:
        flash("Réservation annulée.")
    return redirect(redirect_target)


@app.route("/admin/generate", methods=["GET", "POST"])
@login_required
def admin_generate():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    if request.method == "POST":
        created = generate_month_sessions(int(request.form["year"]), int(request.form["month"]))
        flash(f"Créneaux générés : {created} nouveau(x) créneau(x).")
        return redirect(url_for("index"))
    return render_template_string(TEMPLATE_GENERATE, current_year=date.today().year, current_month=date.today().month)


@app.route("/session/<int:session_id>")
@login_required
def session_detail(session_id):
    if not is_coach_or_admin():
        flash("Accès réservé au coach ou à l’admin.")
        return redirect(url_for("index"))
    ensure_coach_absence_schema()
    session = CourseSession.query.get_or_404(session_id)
    bookings = Booking.query.filter(
        Booking.session_id == session.id,
        Booking.status.in_(["booked", "waiting_list", "absent_unexcused"]),
    ).order_by(Booking.status, Booking.created_at).all()
    bookings = sorted(bookings, key=lambda b: (
        -1 if b.status == "absent_unexcused" or b.attendance_status == "absent" else 2 if b.status == "waiting_list" else 0,
        b.created_at or datetime.utcnow(),
        b.id,
    ))
    return render_template_string(TEMPLATE_SESSION_DETAIL, session=session, bookings=bookings, waitlist_rank=waitlist_rank)


@app.route("/presence/present/<int:booking_id>")
@login_required
def mark_present(booking_id):
    if not is_coach_or_admin():
        flash("Accès réservé à la coach ou à l’admin.")
        return redirect(url_for("index"))
    booking = Booking.query.get_or_404(booking_id)
    if booking.status not in ["booked", "absent_unexcused"]:
        flash("Seules les réservations confirmées peuvent être pointées présentes.")
        return redirect(url_for("session_detail", session_id=booking.session_id))
    booking.status = "booked"
    booking.attendance_status = "present"
    db.session.commit()
    flash(f"{booking.user.display_name()} marqué présent.")
    return redirect(url_for("session_detail", session_id=booking.session_id))


@app.route("/presence/skip/<int:booking_id>")
@login_required
def mark_skipped(booking_id):
    if not is_coach_or_admin():
        flash("Accès réservé à la coach ou à l’admin.")
        return redirect(url_for("index"))
    booking = Booking.query.get_or_404(booking_id)
    if booking.status != "booked":
        flash("Seules les réservations confirmées peuvent être mises de côté.")
        return redirect(url_for("session_detail", session_id=booking.session_id))
    booking.attendance_status = "skipped"
    db.session.commit()
    flash(f"{booking.user.display_name()} mis de côté pour revenir dessus après.")
    return redirect(url_for("session_detail", session_id=booking.session_id))


@app.route("/presence/absent/<int:booking_id>")
@login_required
def mark_absent(booking_id):
    if not is_coach_or_admin():
        flash("Accès réservé à la coach ou à l’admin.")
        return redirect(url_for("index"))
    booking = Booking.query.get_or_404(booking_id)
    if booking.status != "booked":
        flash("Seules les réservations confirmées peuvent être marquées absentes.")
        return redirect(url_for("session_detail", session_id=booking.session_id))
    booking.status = "absent_unexcused"
    booking.attendance_status = "absent"
    db.session.commit()
    apply_absence_sanction(booking.user)
    send_email(
        booking.user.email,
        "Absence enregistrée - Section Fitness",
        f"Bonjour {booking.user.display_name()},\n\nVotre absence au cours {booking.session.course_name} du {booking.session.course_date.strftime('%d/%m/%Y')} a été enregistrée comme non excusée, car la réservation n'avait pas été annulée dans les délais.\n\nSection Fitness"
    )
    flash("Absence non excusée enregistrée.")
    return redirect(url_for("session_detail", session_id=booking.session_id))


@app.route("/presence/late/<int:booking_id>")
@login_required
def mark_late(booking_id):
    if not is_coach_or_admin():
        flash("Accès réservé à la coach ou à l’admin.")
        return redirect(url_for("index"))
    booking = Booking.query.get_or_404(booking_id)
    if booking.status != "absent_unexcused":
        flash("Seule une absence peut être transformée en retard.")
        return redirect(url_for("session_detail", session_id=booking.session_id))
    booking.status = "booked"
    booking.attendance_status = "late"
    refresh_absence_block_status(booking.user)
    db.session.commit()
    flash(f"{booking.user.display_name()} marqué en retard : aucune pénalité d'absence.")
    return redirect(url_for("session_detail", session_id=booking.session_id))


@app.route("/admin/members/edit/<int:user_id>", methods=["GET", "POST"])
@login_required
def admin_edit_member(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if not is_member_account(user):
        flash("Seuls les comptes ayant un profil adhérent peuvent être modifiés depuis cet écran.")
        return redirect(url_for("admin_members"))
    if request.method == "POST":
        new_email = request.form["email"].strip().lower()
        duplicate = User.query.filter(User.email == new_email, User.id != user.id).first()
        if duplicate:
            flash("Un autre compte existe déjà avec cet email.")
            return redirect(url_for("admin_edit_member", user_id=user.id))
        user.email = new_email
        user.member_profile = request.form.get("member_profile", "ouvrant_droit")
        first_name, last_name, full_name = form_full_name()
        user.first_name = first_name
        user.last_name = last_name
        user.full_name = full_name
        user.status = normalize_member_status(user.member_profile, request.form.get("status", "autre"))
        user.rights_holder_name = request.form.get("rights_holder_name", "").strip()
        if user.member_profile == "ayant_droit" and not user.rights_holder_name:
            flash("Merci d'indiquer le nom et prénom de l'ouvrant droit.")
            return redirect(url_for("admin_edit_member", user_id=user.id))
        new_subscription_type = normalize_subscription_type(request.form.get("subscription_type"))
        new_subscription_year = int(request.form.get("subscription_year") or date.today().year)
        ensure_current_membership_period_before_change(user, new_subscription_type, new_subscription_year, created_by=current_user.display_name())
        user.subscription_type = new_subscription_type
        user.subscription_year = new_subscription_year
        user.subscription_end_date = subscription_end(user.subscription_type, user.subscription_year)
        create_membership_period(user, user.subscription_type, user.subscription_year, annual_fee_applies=not MembershipPeriod.query.filter_by(user_id=user.id, subscription_year=user.subscription_year).first(), created_by=current_user.display_name(), notes="Modification admin")
        if user.account_status == "archived" and user.subscription_end_date >= date.today():
            user.account_status = "active"
            user.archived_at = None
            user.archived_reason = None
        new_password = request.form.get("password", "").strip()
        if new_password:
            user.set_password(new_password)
        photo = request.files.get("profile_photo")
        try:
            if photo and photo.filename:
                persist_profile_photo(user, photo)
            generate_member_card(user)
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("admin_edit_member", user_id=user.id))
        db.session.commit()
        flash("Informations adhérent mises à jour.")
        return redirect(url_for("admin_members"))
    if repair_missing_prior_membership_periods(user):
        db.session.commit()
    membership_periods = membership_period_rows(MembershipPeriod.query.filter_by(user_id=user.id).order_by(MembershipPeriod.subscription_year.desc(), MembershipPeriod.start_date.desc()).all())
    return render_template_string(TEMPLATE_ADMIN_EDIT_MEMBER, user=user, current_year=date.today().year, membership_periods=membership_periods, subscription_options=SUBSCRIPTION_PRICES.keys())


@app.route("/admin/members/<int:user_id>/renew", methods=["POST"])
@login_required
def admin_renew_member(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if not is_member_account(user):
        flash("Compte sans profil adhérent.")
        return redirect(url_for("admin_members"))
    subscription_type = normalize_subscription_type(request.form.get("subscription_type"))
    if subscription_type not in SUBSCRIPTION_PRICES:
        flash("Type d'abonnement invalide.")
        return redirect(url_for("admin_edit_member", user_id=user.id))
    subscription_year = int(request.form.get("subscription_year") or date.today().year)
    ensure_current_membership_period_before_change(user, subscription_type, subscription_year, created_by=current_user.display_name())
    annual_fee_applies = not MembershipPeriod.query.filter_by(
        user_id=user.id,
        subscription_year=subscription_year,
    ).first()
    create_membership_period(
        user,
        subscription_type,
        subscription_year,
        annual_fee_applies=annual_fee_applies,
        created_by=current_user.display_name(),
        notes="Renouvellement admin",
    )
    user.subscription_type = subscription_type
    user.subscription_year = subscription_year
    user.subscription_end_date = subscription_end(subscription_type, subscription_year)
    if user.account_status == "archived" and user.subscription_end_date >= date.today():
        user.account_status = "active"
        user.archived_at = None
        user.archived_reason = None
    db.session.commit()
    fee_label = "avec cotisation annuelle" if annual_fee_applies else "sans nouvelle cotisation annuelle"
    flash(f"Adhésion renouvelée ({fee_label}).")
    return redirect(url_for("admin_edit_member", user_id=user.id))


@app.route("/admin/members/<int:user_id>/reservations")
@login_required
def admin_member_reservations(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if not is_member_account(user):
        flash("Compte sans profil adhérent.")
        return redirect(url_for("admin_members"))
    today = date.today()
    end_date = today + timedelta(days=28)
    sessions = CourseSession.query.filter(
        CourseSession.course_date >= today,
        CourseSession.course_date <= end_date
    ).order_by(CourseSession.course_date, CourseSession.start_time).all()
    bookings = Booking.query.filter_by(user_id=user.id).join(CourseSession).order_by(CourseSession.course_date.desc(), CourseSession.start_time.desc()).all()
    active_session_ids = {b.session_id for b in bookings if b.status in ["booked", "waiting_list"]}
    return render_template_string(TEMPLATE_ADMIN_MEMBER_RESERVATIONS, user=user, sessions=sessions, bookings=bookings, active_session_ids=active_session_ids, booked_count=booked_count, waitlist_rank=waitlist_rank, monday_midday_priority_applies=monday_midday_priority_applies)


@app.route("/admin/members/<int:user_id>/book/<int:session_id>")
@login_required
def admin_book_for_member(user_id, session_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    session = CourseSession.query.get_or_404(session_id)
    if not is_member_account(user):
        flash("Compte sans profil adhérent.")
        return redirect(url_for("admin_members"))
    can_book, reason = user_can_book_session(user, session)
    if not can_book:
        flash(reason)
        return redirect(url_for("admin_member_reservations", user_id=user.id))
    booking, result = create_booking_for_user(user, session, by_admin=True)
    if result == "duplicate":
        flash("Cet adhérent est déjà inscrit ou en liste d’attente sur ce créneau.")
    elif result == "waitlist_full":
        flash("Liste d’attente complète.")
    elif result == "booked":
        flash("Réservation confirmée pour l’adhérent.")
    elif result == "waiting_list":
        flash(f"Cours complet : adhérent inscrit en liste d’attente — rang {waitlist_rank(booking)}.")
    return redirect(url_for("admin_member_reservations", user_id=user.id))


@app.route("/admin/members/<int:user_id>/cancel/<int:booking_id>")
@login_required
def admin_cancel_member_booking(user_id, booking_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    booking = Booking.query.get_or_404(booking_id)
    if booking.user_id != user.id:
        flash("Cette réservation ne correspond pas à cet adhérent.")
        return redirect(url_for("admin_member_reservations", user_id=user.id))
    if booking.status not in ["booked", "waiting_list"]:
        flash("Cette réservation n’est plus active.")
        return redirect(url_for("admin_member_reservations", user_id=user.id))
    promoted = cancel_booking_and_promote(booking, cancelled_by_admin=True)
    if promoted:
        flash(f"Réservation annulée pour l’adhérent. {promoted.user.display_name()} a été promu depuis la liste d’attente.")
    else:
        flash("Réservation annulée pour l’adhérent.")
    return redirect(url_for("admin_member_reservations", user_id=user.id))


@app.route("/admin/members/create", methods=["GET", "POST"])
@login_required
def admin_create_member():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        member_profile = request.form.get("member_profile", "ouvrant_droit")
        status = normalize_member_status(member_profile, request.form.get("status", "autre"))
        rights_holder_name = request.form.get("rights_holder_name", "").strip()
        first_name, last_name, full_name = form_full_name()
        subscription_type = normalize_subscription_type(request.form["subscription_type"])
        subscription_year = int(request.form["subscription_year"])
        photo = request.files.get("profile_photo")

        if User.query.filter_by(email=email).first():
            flash("Un compte existe déjà avec cet email.")
            return redirect(url_for("admin_create_member"))
        if member_profile == "ayant_droit" and not rights_holder_name:
            flash("Merci d'indiquer le nom et prénom de l'ouvrant droit.")
            return redirect(url_for("admin_create_member"))

        user = User(
            email=email,
            role="adherent",
            status=status,
            full_name=full_name or email,
            first_name=first_name,
            last_name=last_name,
            member_profile=member_profile,
            rights_holder_name=rights_holder_name,
            subscription_type=subscription_type,
            subscription_year=subscription_year,
            subscription_end_date=subscription_end(subscription_type, subscription_year),
            account_status="pending",
            member_number=next_member_number(subscription_year)
        )
        user.set_password(secrets.token_urlsafe(12))
        db.session.add(user)
        db.session.commit()
        create_membership_period(user, subscription_type, subscription_year, annual_fee_applies=True, created_by=current_user.display_name(), notes="Création admin")
        db.session.commit()

        try:
            if photo and photo.filename:
                persist_profile_photo(user, photo)
                db.session.commit()
            card_path = generate_member_card(user)
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("admin_create_member"))

        send_activation_email(user)
        flash("Adhérent créé. Un lien d'activation lui a été envoyé pour compléter son profil et créer son mot de passe.")
        return redirect(url_for("admin_members"))
    return render_template_string(TEMPLATE_ADMIN_CREATE_MEMBER, current_year=date.today().year)


@app.route("/admin/members/delete/<int:user_id>", methods=["GET", "POST"])
@login_required
def admin_delete_member(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if user.role != "adherent":
        flash("Seuls les comptes adhérents peuvent être supprimés depuis cet écran.")
        return redirect(url_for("admin_members"))
    deleted_email = user.email
    delete_member_completely(user)
    flash(f"Adhérent {deleted_email} supprimé avec ses réservations, adhésions, photo et carte.")
    return redirect(url_for("admin_members"))


@app.route("/admin/members")
@login_required
def admin_members():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    query = active_member_query()
    search = request.args.get("search", "").strip()
    profile = request.args.get("member_profile", "").strip()
    subscription = request.args.get("subscription_type", "").strip()
    year = request.args.get("subscription_year", "").strip()
    account_status = request.args.get("account_status", "").strip()
    followup_year = request.args.get("followup_year", "").strip()
    followup_start = request.args.get("followup_start", "").strip()
    followup_end = request.args.get("followup_end", "").strip()
    if search:
        like = f"%{search.lower()}%"
        query = query.filter(db.or_(db.func.lower(User.full_name).like(like), db.func.lower(User.first_name).like(like), db.func.lower(User.last_name).like(like), db.func.lower(User.email).like(like), db.func.lower(User.member_number).like(like)))
    if profile:
        query = query.filter(User.member_profile == profile)
    if subscription:
        query = query.filter(User.subscription_type.in_([subscription, *[old for old, new in SUBSCRIPTION_ALIASES.items() if new == subscription]]))
    if year.isdigit():
        query = query.filter(User.subscription_year == int(year))
    if account_status:
        query = query.filter(User.account_status == account_status)
    users = query.order_by(User.full_name, User.email).all()
    filter_values = {
        "search": search,
        "member_profile": profile,
        "subscription_type": subscription,
        "subscription_year": year,
        "account_status": account_status,
        "followup_year": followup_year,
        "followup_start": followup_start,
        "followup_end": followup_end,
    }
    start_date = parse_iso_date(followup_start, None) if followup_start else None
    end_date = parse_iso_date(followup_end, None) if followup_end else None
    followup_year_value = int(followup_year) if followup_year.isdigit() else None
    membership_actions = membership_period_rows(membership_actions_query(start_date=start_date, end_date=end_date, year=followup_year_value).limit(25).all())
    return render_template_string(TEMPLATE_MEMBERS, users=users, absence_count=absence_count, filter_values=filter_values, member_profile_labels=MEMBER_PROFILE_LABELS, subscription_options=SUBSCRIPTION_PRICES.keys(), membership_actions=membership_actions)


@app.route("/admin/members/membership-followup/export")
@login_required
def export_membership_followup():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    followup_year = request.args.get("followup_year", "").strip()
    followup_start = request.args.get("followup_start", "").strip()
    followup_end = request.args.get("followup_end", "").strip()
    start_date = parse_iso_date(followup_start, None) if followup_start else None
    end_date = parse_iso_date(followup_end, None) if followup_end else None
    followup_year_value = int(followup_year) if followup_year.isdigit() else None
    rows = membership_period_rows(membership_actions_query(start_date=start_date, end_date=end_date, year=followup_year_value).all())
    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi inscriptions"
    ws.append(["Date action", "Adhérent", "Email", "Abonnement", "Année", "Période", "Tarif abonnement figé", "Cotisation annuelle figée", "Total figé", "Créé par", "Note"])
    for row in rows:
        period = row["period"]
        ws.append([
            period.created_at.strftime("%d/%m/%Y %H:%M") if period.created_at else "",
            row["user"].display_name(),
            row["user"].email,
            row["subscription_type"],
            row["subscription_year"],
            f"{period.start_date.strftime('%d/%m/%Y')} - {period.end_date.strftime('%d/%m/%Y')}",
            row["subscription_price"],
            row["annual_fee"],
            row["total"],
            period.created_by or "",
            period.notes or "",
        ])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name="suivi_inscriptions_renouvellements.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/coaches", methods=["GET", "POST"])
@login_required
def admin_coaches():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        full_name = request.form.get("full_name", "").strip() or email
        existing = User.query.filter_by(email=email).first()
        if existing:
            existing.role = "coach"
            existing.full_name = full_name
            existing.coach_type = existing.coach_type or "titulaire"
            existing.account_status = "pending"
            existing.set_password(secrets.token_urlsafe(12))
            db.session.commit()
            send_activation_email(existing)
            flash("Adresse existante convertie en compte coach. Un lien d'activation a été envoyé si le SMTP est configuré.")
        else:
            coach = User(email=email, role="coach", status="autre", full_name=full_name, account_status="pending", coach_type="titulaire")
            coach.set_password(secrets.token_urlsafe(12))
            db.session.add(coach)
            db.session.commit()
            send_activation_email(coach)
            flash("Coach ajouté. Un lien unique de création de mot de passe a été envoyé si le SMTP est configuré.")
        return redirect(url_for("admin_coaches"))
    coaches = User.query.filter_by(role="coach").order_by(User.full_name, User.email).all()
    return render_template_string(TEMPLATE_COACHES, coaches=coaches)


@app.route("/admin/coaches/delete/<int:user_id>")
@login_required
def admin_delete_coach(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if user.role != "coach":
        flash("Ce compte n'est pas un compte coach.")
        return redirect(url_for("admin_coaches"))
    db.session.delete(user)
    db.session.commit()
    flash("Coach supprimé.")
    return redirect(url_for("admin_coaches"))


@app.route("/admin/coaches/send-activation/<int:user_id>")
@login_required
def admin_send_coach_activation(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if user.role != "coach":
        flash("Ce compte n'est pas un compte coach.")
        return redirect(url_for("admin_coaches"))
    user.account_status = "pending"
    user.set_password(secrets.token_urlsafe(12))
    db.session.commit()
    send_activation_email(user)
    flash("Lien d'activation coach envoyé ou affiché dans la console si SMTP non configuré.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/accounts/<int:user_id>/send-reset")
@login_required
def admin_send_password_reset(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    send_password_reset_email(user)
    flash(f"Lien de réinitialisation envoyé à {user.email} si le SMTP est configuré.")
    return redirect(request.referrer or url_for("admin_settings"))


@app.route("/coach/profile", methods=["GET", "POST"])
@login_required
def coach_profile():
    if current_user.role != "coach" and not is_admin():
        flash("Accès réservé aux coachs.")
        return redirect(url_for("index"))
    ensure_coach_absence_schema()
    coaches = coach_display_names()
    coach_name = request.values.get("coach_name", "").strip() if is_admin() else current_user.display_name()
    if not coach_name:
        coach_name = coaches[0] if coaches else current_user.display_name()
    if request.method == "POST":
        coach_name = request.form.get("coach_name", coach_name).strip() if is_admin() else current_user.display_name()
        start_date = datetime.strptime(request.form["start_date"], "%Y-%m-%d").date()
        end_date = datetime.strptime(request.form.get("end_date") or request.form["start_date"], "%Y-%m-%d").date()
        if end_date < start_date:
            flash("La date de fin doit être postérieure ou égale à la date de début.")
            return redirect(url_for("coach_profile", coach_name=coach_name))
        status = request.form.get("status", "absent")
        replacement = request.form.get("replacement_name", "").strip()
        notes = request.form.get("notes", "").strip()
        current_day = start_date
        saved = 0
        while current_day <= end_date:
            target_sessions = absence_target_sessions(coach_name, current_day)
            if target_sessions:
                for session in target_sessions:
                    upsert_coach_absence(coach_name, current_day, status, replacement, notes, session=session, reset_followup=current_user.role == "coach")
                    saved += 1
            else:
                pass
            current_day += timedelta(days=1)
        db.session.commit()
        if saved == 0:
            flash("Aucune absence créée : aucun cours n'existe pour cette coach sur la période sélectionnée.")
            return redirect(url_for("coach_profile", coach_name=coach_name))
        if current_user.role == "coach":
            sent = notify_admins_of_coach_absence(coach_name, start_date, end_date, status, replacement, notes)
            member_sent = notify_members_of_coach_absence(coach_name, start_date, end_date, status, replacement, notes)
            flash(f"Absence/congé enregistré sur {saved} jour(s). Email envoyé à {sent} admin(s) et {member_sent} adhérent(s) inscrit(s)." if sent or member_sent else f"Absence/congé enregistré sur {saved} jour(s). Aucun email envoyé.")
        else:
            member_sent = notify_members_of_coach_absence(coach_name, start_date, end_date, status, replacement, notes)
            flash(f"Absence/congé enregistré sur {saved} jour(s). Email envoyé à {member_sent} adhérent(s) inscrit(s)." if member_sent else f"Absence/congé enregistré sur {saved} jour(s).")
        return redirect(url_for("coach_profile", coach_name=coach_name))
    today = date.today()
    absences = CoachAbsence.query.filter(
        CoachAbsence.coach_name == coach_name,
        CoachAbsence.absence_date >= today - timedelta(days=30)
    ).order_by(CoachAbsence.absence_date.desc()).all()
    return render_template_string(TEMPLATE_COACH_PROFILE, coach_name=coach_name, coaches=coaches, replacement_coaches=coach_replacement_options(), absences=absences, today=today)


@app.route("/coach/schedule")
@login_required
def coach_schedule():
    if current_user.role != "coach":
        flash("Accès réservé aux coachs.")
        return redirect(url_for("index"))
    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
        month = int(request.args.get("month", today.month))
        if month < 1 or month > 12:
            raise ValueError
    except ValueError:
        year = today.year
        month = today.month
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    identities = coach_identity_names(current_user)
    all_sessions = CourseSession.query.filter(
        CourseSession.course_date >= start,
        CourseSession.course_date <= end,
    ).order_by(CourseSession.course_date, CourseSession.start_time).all()
    absences = CoachAbsence.query.filter(
        CoachAbsence.absence_date >= start,
        CoachAbsence.absence_date <= end,
        CoachAbsence.replacement_name.in_(identities),
    ).order_by(CoachAbsence.absence_date, CoachAbsence.coach_name).all()
    own_absences = CoachAbsence.query.filter(
        CoachAbsence.absence_date >= start,
        CoachAbsence.absence_date <= end,
        CoachAbsence.coach_name.in_(identities),
    ).order_by(CoachAbsence.absence_date, CoachAbsence.coach_name).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in (own_absences + absences)}
    all_absences = CoachAbsence.query.filter(CoachAbsence.absence_date >= start, CoachAbsence.absence_date <= end).all()
    all_abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in all_absences}
    sessions = [
        session for session in all_sessions
        if session.coach_name in identities or effective_coach_for_session(session, all_abs_by_key) in identities
    ]
    replacements = []
    for absence in absences:
        if absence.session_id:
            replacements.append({"absence": absence, "session": absence.session})
        else:
            replacement_sessions = absence_target_sessions(absence.coach_name, absence.absence_date)
            if replacement_sessions:
                for session in replacement_sessions:
                    replacements.append({"absence": absence, "session": session})
            else:
                replacements.append({"absence": absence, "session": None})
    invoice_rows = coach_monthly_invoice_rows(start, end, coach_filter=current_user.display_name())
    invoice_detail_rows = coach_invoice_detail_rows(start, end, coach_filter=current_user.display_name())
    invoice_summary_rows = coach_invoice_summary_rows(start, end, coach_filter=current_user.display_name())
    return render_template_string(TEMPLATE_COACH_SCHEDULE, sessions=sessions, replacements=replacements, abs_by_key=abs_by_key, invoice_rows=invoice_rows, invoice_detail_rows=invoice_detail_rows, invoice_summary_rows=invoice_summary_rows, year=year, month=month, weekday_labels=WEEKDAY_LABELS)


@app.route("/coach/profile/delete/<int:absence_id>")
@login_required
def delete_coach_absence(absence_id):
    absence = CoachAbsence.query.get_or_404(absence_id)
    if current_user.role != "admin" and (current_user.role != "coach" or absence.coach_name != current_user.display_name()):
        flash("Action non autorisée.")
        return redirect(url_for("index"))
    coach_name = absence.coach_name
    absence_date = absence.absence_date
    db.session.delete(absence)
    db.session.commit()
    flash("Absence supprimée.")
    if request.args.get("source") == "admin_planning" and is_admin():
        return redirect(url_for("admin_coach_planning", view_mode=request.args.get("view_mode", "rolling"), start_date=request.args.get("start_date", ""), end_date=request.args.get("end_date", ""), year=request.args.get("year", absence_date.year), month=request.args.get("month", absence_date.month)))
    return redirect(url_for("coach_profile", coach_name=coach_name))


@app.route("/admin/archives")
@login_required
def admin_archives():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    archive_past_bookings()
    bookings = Booking.query.join(CourseSession).filter(Booking.archived.is_(True)).order_by(CourseSession.course_date.desc(), CourseSession.start_time.desc(), Booking.created_at).all()
    return render_template_string(TEMPLATE_ARCHIVES, bookings=bookings)


@app.route("/admin/members/export")
@login_required
def export_members_excel():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Adhérents"
    ws.append(["Nom", "Email", "Statut", "Profil adhérent", "Ouvrant droit lié", "Abonnement", "Année", "ID adhérent", "Bloqué jusqu'au", "Absences 90j"])
    users = User.query.filter_by(role="adherent").order_by(User.full_name, User.email).all()
    for u in users:
        ws.append([u.display_name(), u.email, u.status, u.member_profile or "", u.rights_holder_name or "", u.subscription_type, u.subscription_year, u.member_number, str(u.blocked_until or ""), absence_count(u)])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name="adherents_fitness.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/members/email", methods=["GET", "POST"])
@login_required
def admin_email_members():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    selected_ids = request.values.getlist("user_ids")
    target_roles = request.values.getlist("target_roles") or ["adherent"]
    valid_roles = {"adherent", "admin", "coach"}
    target_roles = [role for role in target_roles if role in valid_roles] or ["adherent"]
    if selected_ids:
        selected_member_ids = [int(i) for i in selected_ids if str(i).isdigit()]
        user_filters = []
        if "adherent" in target_roles and selected_member_ids:
            user_filters.append(db.and_(User.role == "adherent", User.id.in_(selected_member_ids)))
        if "admin" in target_roles:
            user_filters.append(User.role == "admin")
        if "coach" in target_roles:
            user_filters.append(User.role == "coach")
        users = User.query.filter(User.account_status != "archived", db.or_(*user_filters)).order_by(User.role, User.full_name, User.email).all() if user_filters else []
    else:
        users = User.query.filter(User.role.in_(target_roles), User.account_status != "archived").order_by(User.role, User.full_name, User.email).all()
    if request.method == "POST":
        subject = request.form["subject"].strip()
        body = request.form["body"].strip()
        if not subject or not body:
            flash("Merci de renseigner un objet et un message.")
            return render_template_string(TEMPLATE_EMAIL_MEMBERS, users=users, target_roles=target_roles)
        signed_body = admin_email_signature_body(body)
        signed_html = admin_email_signature_html(body)
        inline_images = {"fitness_logo": LOGO_PATH} if LOGO_PATH.exists() else {}
        user_ids = [u.id for u in users]
        send_member_campaign_async(user_ids, subject, signed_body, signed_html, inline_images)
        flash(f"Campagne email lancée pour {len(user_ids)} destinataire(s). L'envoi continue en arrière-plan pour éviter une erreur serveur.")
        return redirect(url_for("admin_members"))
    return render_template_string(TEMPLATE_EMAIL_MEMBERS, users=users, target_roles=target_roles)


@app.route("/admin/blocked")
@login_required
def blocked_members():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    users = User.query.filter(User.blocked_until >= date.today()).order_by(User.blocked_until.desc()).all()
    recent_absences = Booking.query.join(CourseSession).join(Booking.user).filter(
        Booking.status == "absent_unexcused",
        CourseSession.course_date >= date.today() - timedelta(days=90),
    ).order_by(CourseSession.course_date.desc(), CourseSession.start_time.desc()).all()
    return render_template_string(TEMPLATE_BLOCKED, users=users, absence_count=absence_count, recent_absences=recent_absences)


@app.route("/admin/unblock/<int:user_id>")
@login_required
def unblock_member(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    user.blocked_until = None
    user.blocked_at = None
    user.blocked_reason = None
    db.session.commit()
    send_email(user.email, "Déblocage de votre compte Fitness", f"Bonjour {user.display_name()},\n\nVotre compte Fitness a été débloqué. Vous pouvez à nouveau réserver des cours.\n\nSection Fitness")
    flash("Adhérent débloqué.")
    return redirect(url_for("blocked_members"))


@app.route("/admin/absence/remove/<int:booking_id>")
@login_required
def remove_unexcused_absence(booking_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    booking = Booking.query.get_or_404(booking_id)
    if booking.status != "absent_unexcused":
        flash("Cette réservation n'est pas marquée absente.")
        return redirect(url_for("blocked_members"))
    user = booking.user
    booking.status = "booked"
    booking.attendance_status = None
    refresh_absence_block_status(user)
    db.session.commit()
    flash(f"Marquage absent retiré pour {user.display_name()}.")
    return redirect(url_for("blocked_members"))


@app.route("/card/<int:user_id>")
@login_required
def download_card(user_id):
    user = User.query.get_or_404(user_id)
    if current_user.id != user.id and not is_admin():
        flash("Action non autorisée.")
        return redirect(url_for("index"))
    generate_member_card(user)
    return send_file(STATIC_DIR / user.member_card, as_attachment=True, download_name=f"carte_adherent_{user.id}.png")


@app.route("/profile-photo/<int:user_id>")
@login_required
def profile_photo_file(user_id):
    user = User.query.get_or_404(user_id)
    if current_user.id != user.id and not is_coach_or_admin():
        flash("Action non autorisée.")
        return redirect(url_for("index"))
    data, mime = user_profile_photo_bytes(user)
    if not data:
        flash("Aucune photo disponible.")
        return redirect(url_for("index"))
    return send_file(BytesIO(data), mimetype=mime)


@app.route("/admin/export")
@login_required
def export_excel():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Présences"
    ws.append(["Date", "Cours", "Horaire", "Nom", "Email", "Statut adhérent", "Abonnement", "Année", "Statut réservation"])
    bookings = Booking.query.join(CourseSession).order_by(CourseSession.course_date, CourseSession.start_time).all()
    for b in bookings:
        ws.append([b.session.course_date.strftime("%d/%m/%Y"), b.session.course_name, f"{b.session.start_time.strftime('%H:%M')} - {b.session.end_time.strftime('%H:%M')}", b.user.display_name(), b.user.email, b.user.status, b.user.subscription_type, b.user.subscription_year, b.status])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name="presences_fitness.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/statistics/export")
@login_required
def export_statistics_excel():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    today = date.today()
    start = parse_iso_date(request.args.get("start_date", ""), date(today.year, today.month, 1))
    end = parse_iso_date(request.args.get("end_date", ""), date(today.year, today.month, monthrange(today.year, today.month)[1]))
    if end < start:
        end = start
    course_filter = request.args.get("course_filter", "").strip()
    rows, monthly_rows = course_booking_stats(start, end, course_filter)
    wb = Workbook()
    ws = wb.active
    ws.title = "Stats cours"
    ws.append(["Date", "Cours", "Coach", "Horaire", "Jauge", "Réservés", "Liste attente", "Absents", "Mensuels", "Cadres/autres", "Ratio mensuels %", "Remplissage %"])
    for row in rows:
        ws.append([
            row["date"].strftime("%d/%m/%Y"),
            row["course"],
            row["coach"],
            row["time"],
            row["capacity"],
            row["booked"],
            row["waiting"],
            row["absent"],
            row["mensuels"],
            row["cadres_autres"],
            round(row["ratio_mensuel"], 1),
            round(row["fill_rate"], 1),
        ])
    ws2 = wb.create_sheet("Stats mensuelles")
    ws2.append(["Mois", "Séances", "Réservations", "Mensuels", "Liste attente", "Ratio mensuels %"])
    for row in monthly_rows:
        ws2.append([row["month"], row["sessions"], row["booked"], row["mensuels"], row["waiting"], round(row["ratio_mensuel"], 1)])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name=f"statistiques_cours_{start.isoformat()}_{end.isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/email-diagnostic")
@login_required
def admin_email_diagnostic():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    print("\n--- DIAGNOSTIC SMTP ---")
    print("SMTP_HOST:", SMTP_HOST or "NON DEFINI")
    print("SMTP_PORT:", SMTP_PORT)
    print("SMTP_TIMEOUT:", SMTP_TIMEOUT)
    print("SMTP_USER défini:", bool(SMTP_USER))
    print("SMTP_PASSWORD défini:", bool(SMTP_PASSWORD))
    print("MAIL_FROM:", MAIL_FROM)
    print("BREVO_API_KEY défini:", bool(BREVO_API_KEY))
    print("-----------------------\n")
    ok = send_email(current_user.email, "Test email Section Fitness", "Bonjour,\n\nCeci est un test d'envoi email depuis la Section Fitness.\n\nSection Fitness")
    flash("Email test envoyé. Vérifiez votre boîte mail." if ok else "Email test non envoyé. Vérifiez les variables SMTP dans Render et les logs.")
    return redirect(url_for("index"))


@app.route("/infos-utiles", methods=["GET", "POST"])
@login_required
def useful_info():
    ensure_useful_documents_schema()
    if request.method == "POST":
        if not is_admin():
            flash("Accès réservé à l’admin.")
            return redirect(url_for("useful_info"))
        try:
            file_path = save_useful_document(request.files.get("document_file"))
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("useful_info"))
        title = request.form.get("title", "").strip() or Path(request.files["document_file"].filename).stem
        doc = UsefulDocument(
            title=title,
            category=request.form.get("category", "").strip(),
            file_path=file_path,
            notes=request.form.get("notes", "").strip(),
            uploaded_by=current_user.display_name(),
        )
        db.session.add(doc)
        db.session.commit()
        flash("Document ajouté aux infos utiles.")
        return redirect(url_for("useful_info"))
    documents = UsefulDocument.query.order_by(UsefulDocument.category, UsefulDocument.title).all()
    return render_template_string(TEMPLATE_USEFUL_INFO, documents=documents)


@app.route("/infos-utiles/delete/<int:document_id>")
@login_required
def delete_useful_document(document_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("useful_info"))
    ensure_useful_documents_schema()
    document = UsefulDocument.query.get_or_404(document_id)
    db.session.delete(document)
    db.session.commit()
    flash("Document retiré des infos utiles.")
    return redirect(url_for("useful_info"))


BASE_TEMPLATE_STYLE = """
<style>
:root{--green:#34a853;--green2:#8ee35f;--dark:#061417;--muted:#6b7280;--line:#e5e7eb;--bg:#f6f8fb;--danger:#ef4444;--orange:#f59e0b}
*{box-sizing:border-box} body{margin:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:var(--bg);color:#111827}.layout{display:flex;min-height:100vh}.side{width:250px;background:linear-gradient(180deg,#07161a,#031014);color:white;padding:26px 18px;position:sticky;top:0;height:100vh}.logo{width:118px;height:118px;border-radius:50%;background:white;display:block;margin:0 auto 26px;object-fit:contain}.nav a{display:block;color:white;text-decoration:none;padding:12px 14px;border-radius:10px;margin:8px 0;font-weight:600}.nav a:hover,.nav .active{background:var(--green)}.logout{color:#ff6b6b!important;margin-top:26px}.main{flex:1;padding:30px}.top{display:flex;justify-content:space-between;align-items:center;margin-bottom:24px}.btn{border:0;background:var(--green);color:white;padding:11px 16px;border-radius:10px;text-decoration:none;font-weight:700;display:inline-block;cursor:pointer}.btn.secondary{background:white;color:#111827;border:1px solid var(--line)}.btn.danger{background:var(--danger)}.grid{display:grid;grid-template-columns:repeat(4,1fr);gap:18px;margin-bottom:24px}.card{background:white;border:1px solid var(--line);border-radius:18px;padding:20px;box-shadow:0 10px 25px rgba(15,23,42,.05)}.stat{font-size:34px;font-weight:800;margin-top:8px}.muted{color:var(--muted)}.content-grid{display:grid;grid-template-columns:1.1fr .9fr;gap:22px}.session{display:flex;justify-content:space-between;gap:14px;align-items:center;border:1px solid var(--line);padding:15px;border-radius:14px;margin:12px 0;background:#fff}.badge{padding:6px 10px;border-radius:999px;font-size:13px;font-weight:700;background:#e8f8ed;color:#18793a}.badge.full{background:#fee2e2;color:#b91c1c}.badge.wait{background:#fff7ed;color:#c2410c}.table{width:100%;border-collapse:collapse;background:white;border-radius:18px;overflow:hidden}.table th,.table td{padding:14px;border-bottom:1px solid var(--line);text-align:left}.table th{background:#f9fafb}.form-wrap{max-width:760px;margin:30px auto}.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}.field label{font-weight:700;display:block;margin-bottom:7px}.field input,.field select{width:100%;padding:13px;border:1px solid #d1d5db;border-radius:10px;font-size:15px}.flash{background:#ecfdf5;border:1px solid #bbf7d0;color:#166534;padding:12px 16px;border-radius:12px;margin-bottom:18px}.login{min-height:100vh;display:grid;place-items:center;padding:22px}.login-box{width:100%;max-width:430px}.photo-preview{width:115px;height:115px;border-radius:50%;object-fit:cover;background:#e5e7eb}.admin-photo{width:54px;height:54px;border-radius:12px;object-fit:cover;background:#e5e7eb;border:1px solid var(--line)}.card-preview{width:100%;border-radius:16px;border:1px solid var(--line)}@media(max-width:900px){.layout{display:block}.side{width:auto;height:auto;position:relative}.grid,.content-grid,.form-grid{grid-template-columns:1fr}.main{padding:18px}}
</style>
"""


def shell(content, active=""):
    logo = url_for('static', filename='logo.png') if LOGO_PATH.exists() else ''
    admin_links = ""
    if current_user.is_authenticated and current_user.role == "admin":
        admin_links = (
            f'<a class="{"active" if active=="members" else ""}" href="{url_for("admin_members")}">Adhérents</a>'
            f'<a class="{"active" if active=="office" else ""}" href="{url_for("admin_office")}">Bureau / Admins</a>'
            f'<a class="{"active" if active=="coach_planning" else ""}" href="{url_for("admin_coach_planning")}">Planning coachs</a>'
            f'<a class="{"active" if active=="statistics" else ""}" href="{url_for("admin_statistics")}">Statistiques</a>'
            f'<a class="{"active" if active=="budget" else ""}" href="{url_for("admin_budget")}">Budget</a>'
            f'<a class="{"active" if active=="inventory" else ""}" href="{url_for("admin_inventory")}">Inventaire</a>'
            f'<a class="{"active" if active=="settings" else ""}" href="{url_for("admin_settings")}">Paramètres</a>'
            f'<a class="{"active" if active=="useful_info" else ""}" href="{url_for("useful_info")}">Infos utiles</a>'
            f'<a class="{"active" if active=="blocked" else ""}" href="{url_for("blocked_members")}">Adhérents bloqués</a>'
            f'<a class="{"active" if active=="archives" else ""}" href="{url_for("archived_members")}">Archives</a>'
            f'<a class="{"active" if active=="member_profile" else ""}" href="{url_for("member_profile")}">Mon profil adhérent</a>'
            f'<a class="{"active" if active=="member_coach_planning" else ""}" href="{url_for("member_coach_planning")}">Réserver mes cours</a>'
        )
    member_links = ""
    if current_user.is_authenticated and current_user.role == "adherent":
        member_links = (
            f'<a class="{"active" if active=="member_profile" else ""}" href="{url_for("member_profile")}">Mon profil</a>'
            f'<a class="{"active" if active=="member_coach_planning" else ""}" href="{url_for("member_coach_planning")}">Planning coachs</a>'
            f'<a class="{"active" if active=="useful_info" else ""}" href="{url_for("useful_info")}">Infos utiles</a>'
        )
    coach_links = ""
    if current_user.is_authenticated and current_user.role == "coach":
        coach_links = (
            f'<a class="{"active" if active=="coach_schedule" else ""}" href="{url_for("coach_schedule")}">Planning</a>'
            f'<a class="{"active" if active=="coach_profile" else ""}" href="{url_for("coach_profile")}">Absence</a>'
        )
    return f"""
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">{BASE_TEMPLATE_STYLE}<title>Section Fitness</title></head>
<body><div class="layout"><aside class="side">{f'<img class="logo" src="{logo}">' if logo else '<div class="logo"></div>'}<div class="nav">
<a class="{'active' if active=='home' else ''}" href="{url_for('index')}">Tableau de bord</a>
{admin_links}
{member_links}
{coach_links}
<a class="logout" href="{url_for('logout')}">Déconnexion</a>
</div></aside><main class="main">{content}</main></div></body></html>
"""


TEMPLATE_INDEX = """
{% set content %}
<div class="top"><div><h1>Bienvenue, {{ current_user.display_name() }} 👋</h1><p class="muted">Voici le planning de la Section Fitness.</p></div>{% if current_user.role == 'adherent' %}<a class="btn secondary" href="{{ url_for('download_card', user_id=current_user.id) }}">Ma carte adhérent</a>{% endif %}</div>
{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}
{% if current_user.is_blocked() %}<div class="flash" style="background:#fef2f2;border-color:#fecaca;color:#991b1b">Votre compte est bloqué jusqu'au {{ current_user.blocked_until }}.</div>{% endif %}
<div class="content-grid"><section class="card"><h2>Prochaines séances</h2>{% if current_user.role not in ['admin','coach'] %}<form method="get" style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:14px;padding:14px;margin:12px 0 18px"><h3 style="margin-top:0">Filtres</h3><div class="form-grid"><div class="field"><label>Cours</label><select name="course_filter"><option value="">Tous</option>{% for name in preference_options.courses %}<option value="{{ name }}" {% if selected_course == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field"><label>Coach</label><select name="coach_filter"><option value="">Tous</option>{% for name in preference_options.coaches %}<option value="{{ name }}" {% if selected_coach == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field"><label>Créneau</label><select name="slot_filter"><option value="">Tous</option>{% for name in preference_options.slots %}<option value="{{ name }}" {% if selected_slot == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div></div><br><button class="btn" type="submit">Filtrer</button> <a class="btn secondary" href="{{ url_for('index') }}">Réinitialiser</a></form>{% endif %}{% for s in sessions %}<div class="session"><div><div class="muted">{{ s.course_date.strftime('%A %d/%m/%Y') }} · {{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</div><strong>{{ s.course_name }}</strong>{% if s.is_reservable %}<div class="muted">{{ booked_count(s) }} / {{ s.capacity }} inscrits</div>{% else %}<div class="muted">Pas de réservation</div>{% endif %}</div><div>{% if not s.is_reservable %}<span class="badge wait">Sans réservation</span>{% elif booked_count(s) >= s.capacity %}<span class="badge full">Complet</span>{% else %}<span class="badge">{{ s.capacity - booked_count(s) }} places</span>{% endif %}<br><br>{% if current_user.role == 'adherent' and s.is_reservable %}<a class="btn" href="{{ url_for('book', session_id=s.id) }}">Réserver</a>{% endif %}{% if current_user.role in ['admin','coach'] %}<a class="btn secondary" href="{{ url_for('session_detail', session_id=s.id) }}">Voir liste</a>{% endif %}</div></div>{% endfor %}</section>
{% if current_user.role == 'admin' %}<section class="card"><h2>Dernières actions adhérents</h2><table class="table"><tr><th>Date action</th><th>Adhérent</th><th>Cours</th><th>Statut</th><th>Actions</th></tr>{% for b in latest_bookings %}<tr><td>{{ b.created_at.strftime('%d/%m/%Y %H:%M') if b.created_at else '-' }}</td><td>{{ b.user.display_name() }}<br><small class="muted">{{ b.user.email }}</small></td><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}<br>{{ b.session.course_name }}</td><td>{% if b.status == 'waiting_list' %}<span class="badge wait">Liste d’attente — rang {{ waitlist_rank(b) }}</span>{% elif b.status == 'booked' %}<span class="badge">Réservé</span>{% else %}<span class="badge full">{{ b.status }}</span>{% endif %}</td><td><a class="btn secondary" href="{{ url_for('session_detail', session_id=b.session_id) }}">Modifier</a>{% if b.status in ['booked','waiting_list'] %} <a class="btn danger" href="{{ url_for('cancel', booking_id=b.id) }}" onclick="return confirm('Annuler cette réservation ?')">Supprimer</a>{% endif %}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune réservation récente.</td></tr>{% endfor %}</table></section>{% else %}<section class="card"><h2>Mes réservations</h2><table class="table"><tr><th>Date</th><th>Cours</th><th>Statut</th><th></th></tr>{% for b in current_user.bookings %}<tr><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}</td><td>{{ b.session.course_name }}</td><td>{% if b.status == 'waiting_list' %}<span class="badge wait">Vous êtes en liste d’attente — rang {{ waitlist_rank(b) }}</span>{% else %}<span class="badge">{{ b.status }}</span>{% endif %}</td><td>{% if b.status in ['booked','waiting_list'] %}<a class="btn danger" href="{{ url_for('cancel', booking_id=b.id) }}">Annuler</a>{% endif %}</td></tr>{% endfor %}</table></section>{% endif %}</div>
{% endset %}{{ shell(content, 'home')|safe }}
"""
TEMPLATE_INDEX = TEMPLATE_INDEX.replace(
    """{% for s in sessions %}<div class="session"><div><div class="muted">{{ s.course_date.strftime('%A %d/%m/%Y') }} · {{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</div><strong>{{ s.course_name }}</strong>""",
    """{% for s in sessions %}{% set a = absence_for_session(abs_by_key, s) %}<div class="session"><div><div class="muted">{{ s.course_date.strftime('%A %d/%m/%Y') }} · {{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</div><strong>{{ s.course_name }}</strong>{% if a %}<br><span class="badge {{ absence_badge_class(a) }}">{{ absence_display_label(a) }}</span>{% if a.replacement_name %}<br><small>Remplaçant : {{ a.replacement_name }}</small>{% endif %}{% endif %}""",
    1,
)
TEMPLATE_INDEX = TEMPLATE_INDEX.replace(
    """{% if current_user.role == 'adherent' and s.is_reservable %}<a class="btn" href="{{ url_for('book', session_id=s.id) }}">Réserver</a>{% endif %}""",
    """{% if current_user.role == 'adherent' and s.is_reservable %}{% set can_book, reason = user_can_book_session(current_user, s) %}{% if a and absence_blocks_booking(a) %}<span class="badge full">Indisponible</span>{% elif not can_book %}<span class="badge wait">{{ reason }}</span>{% else %}<a class="btn" href="{{ url_for('book', session_id=s.id, next=request.full_path) }}">Réserver</a>{% endif %}{% endif %}""",
    1,
)
TEMPLATE_INDEX = TEMPLATE_INDEX.replace(
    """{% for b in current_user.bookings %}""",
    """{% for b in current_bookings %}""",
    1,
)

TEMPLATE_INDEX = """
{% set content %}
<div class="top"><div><h1>Bienvenue, {{ current_user.display_name() }} 👋</h1><p class="muted">Voici le planning de la Section Fitness.</p></div>{% if current_user.role == 'adherent' %}<a class="btn secondary" href="{{ url_for('download_card', user_id=current_user.id) }}">Ma carte adhérent</a>{% endif %}</div>
{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}
{% if current_user.is_blocked() %}<div class="flash" style="background:#fef2f2;border-color:#fecaca;color:#991b1b">Votre compte est bloqué jusqu'au {{ current_user.blocked_until }}.</div>{% endif %}
{% if current_user.role == 'adherent' %}<div class="flash">Période de reprise : les réservations du {{ temporary_booking_grace_start.strftime('%d/%m/%Y') }} au {{ temporary_booking_grace_end.strftime('%d/%m/%Y') }} restent ouvertes pendant la mise à jour des adhésions.</div>{% endif %}
<div class="content-grid"><section class="card"><h2>Prochaines séances</h2>{% if current_user.role not in ['admin','coach'] %}<form method="get" style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:14px;padding:14px;margin:12px 0 18px"><h3 style="margin-top:0">Filtres</h3><div class="form-grid"><div class="field"><label>Cours</label><select name="course_filter"><option value="">Tous</option>{% for name in preference_options.courses %}<option value="{{ name }}" {% if selected_course == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field"><label>Coach</label><select name="coach_filter"><option value="">Tous</option>{% for name in preference_options.coaches %}<option value="{{ name }}" {% if selected_coach == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field"><label>Créneau</label><select name="slot_filter"><option value="">Tous</option>{% for name in preference_options.slots %}<option value="{{ name }}" {% if selected_slot == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div></div><br><button class="btn" type="submit">Filtrer</button> <a class="btn secondary" href="{{ url_for('index') }}">Réinitialiser</a></form>{% endif %}{% for s in sessions %}{% set a = absence_for_session(abs_by_key, s) %}{% set booking = active_booking_by_session.get(s.id) %}<div class="session"><div><div class="muted">{{ s.course_date.strftime('%A %d/%m/%Y') }} · {{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</div><strong>{{ s.course_name }}</strong>{% if a %}<br><span class="badge {{ absence_badge_class(a) }}">{{ absence_display_label(a) }}</span>{% if a.replacement_name %}<br><small>Remplaçant : {{ a.replacement_name }}</small>{% endif %}{% endif %}{% if s.is_reservable %}<div class="muted">{{ booked_count(s) }} / {{ s.capacity }} inscrits</div>{% else %}<div class="muted">Pas de réservation</div>{% endif %}</div><div>{% if not s.is_reservable %}<span class="badge wait">Sans réservation</span>{% elif booking %}{% if booking.status == 'waiting_list' %}<span class="badge wait">Liste d’attente — rang {{ waitlist_rank(booking) }}</span>{% else %}<span class="badge">Déjà réservé</span>{% endif %}{% elif booked_count(s) >= s.capacity %}<span class="badge full">Complet</span>{% else %}<span class="badge">{{ s.capacity - booked_count(s) }} places</span>{% endif %}<br><br>{% if current_user.role == 'adherent' and s.is_reservable %}{% set can_book, reason = user_can_book_session(current_user, s) %}{% if booking %}<a class="btn danger" href="{{ url_for('cancel', booking_id=booking.id, next=request.full_path) }}">Annuler</a>{% elif a and absence_blocks_booking(a) %}<span class="badge full">Indisponible</span>{% elif not can_book %}<span class="badge wait">{{ reason }}</span>{% else %}<a class="btn" href="{{ url_for('book', session_id=s.id, next=request.full_path) }}">Réserver</a>{% endif %}{% endif %}{% if current_user.role in ['admin','coach'] %}<a class="btn secondary" href="{{ url_for('session_detail', session_id=s.id) }}">Voir liste</a>{% endif %}</div></div>{% else %}<p class="muted">Aucune séance à venir.</p>{% endfor %}</section>
{% if current_user.role == 'admin' %}<section class="card"><h2>Dernières actions adhérents</h2><table class="table"><tr><th>Date action</th><th>Adhérent</th><th>Cours</th><th>Statut</th><th>Actions</th></tr>{% for b in latest_bookings %}<tr><td>{{ b.created_at.strftime('%d/%m/%Y %H:%M') if b.created_at else '-' }}</td><td>{{ b.user.display_name() }}<br><small class="muted">{{ b.user.email }}</small></td><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}<br>{{ b.session.course_name }}</td><td>{% if b.status == 'waiting_list' %}<span class="badge wait">Liste d’attente — rang {{ waitlist_rank(b) }}</span>{% elif b.status == 'booked' %}<span class="badge">Réservé</span>{% else %}<span class="badge full">{{ b.status }}</span>{% endif %}</td><td><a class="btn secondary" href="{{ url_for('session_detail', session_id=b.session_id) }}">Modifier</a>{% if b.status in ['booked','waiting_list'] %} <a class="btn danger" href="{{ url_for('cancel', booking_id=b.id) }}" onclick="return confirm('Annuler cette réservation ?')">Supprimer</a>{% endif %}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune réservation récente.</td></tr>{% endfor %}</table></section>{% else %}<section class="card"><h2>Mes réservations à venir</h2><table class="table"><tr><th>Date</th><th>Cours</th><th>Statut</th><th></th></tr>{% for b in current_bookings %}<tr><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}<br><small>{{ b.session.start_time.strftime('%H:%M') }} - {{ b.session.end_time.strftime('%H:%M') }}</small></td><td>{{ b.session.course_name }}</td><td>{% if b.status == 'waiting_list' %}<span class="badge wait">Liste d’attente — rang {{ waitlist_rank(b) }}</span>{% else %}<span class="badge">Réservé</span>{% endif %}</td><td>{% if b.status in ['booked','waiting_list'] %}<a class="btn danger" href="{{ url_for('cancel', booking_id=b.id, next=request.full_path) }}">Annuler</a>{% endif %}</td></tr>{% else %}<tr><td colspan="4" class="muted">Aucune réservation à venir.</td></tr>{% endfor %}</table><br><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Règles de réservation</h2><p>Annulation possible jusqu'à 2h avant le cours.</p><p>Deux absences injustifiées sur 90 jours entraînent un blocage temporaire des réservations.</p><p>Si vous arrivez en retard, la coach peut corriger l'appel : le retard n'entraîne pas de pénalité.</p></div></section>{% endif %}</div>
{% endset %}{{ shell(content, 'home')|safe }}
"""
TEMPLATE_INDEX = TEMPLATE_INDEX.replace(
    """<section class="card"><h2>Dernières actions adhérents</h2><table class="table">""",
    """<section class="card"><h2>Dernières actions adhérents</h2><div class="card" style="box-shadow:none;background:#f9fafb;margin-bottom:14px"><h2>Vue adhérent démo</h2><p class="muted">Pour voir l'affichage exact d'un profil adhérent, déconnectez-vous puis connectez-vous avec <code>adherent@fitness.local</code> et le mot de passe <code>adherent123</code>.</p></div><div class="card" style="box-shadow:none;background:#f9fafb;margin-bottom:14px"><h2>Vue coach démo</h2><p class="muted">Pour voir l'affichage exact d'un profil coach, déconnectez-vous puis connectez-vous avec <code>coach@fitness.local</code> et le mot de passe <code>coach123</code>.</p></div><table class="table">""",
    1,
)
TEMPLATE_INDEX = TEMPLATE_INDEX.replace(
    """{% if current_user.role == 'adherent' %}<div class="flash">Période de reprise : les réservations du {{ temporary_booking_grace_start.strftime('%d/%m/%Y') }} au {{ temporary_booking_grace_end.strftime('%d/%m/%Y') }} restent ouvertes pendant la mise à jour des adhésions.</div>{% endif %}""",
    """{% if current_user.role == 'adherent' %}<div class="flash">Période de reprise : les réservations du {{ temporary_booking_grace_start.strftime('%d/%m/%Y') }} au {{ temporary_booking_grace_end.strftime('%d/%m/%Y') }} restent ouvertes pendant la mise à jour des adhésions.</div><section class="card"><h2>Mes cours réservés</h2><p class="muted">Vos prochaines réservations, pour vous les remémorer ou annuler rapidement.</p><table class="table"><tr><th>Date</th><th>Cours</th><th>Statut</th><th>Action</th></tr>{% for b in current_bookings %}<tr><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}<br><small>{{ b.session.start_time.strftime('%H:%M') }} - {{ b.session.end_time.strftime('%H:%M') }}</small></td><td>{{ b.session.course_name }}<br><small class="muted">{{ b.session.coach_name or '-' }}</small></td><td>{% set a = absence_for_session(abs_by_key, b.session) %}{% if a %}<span class="badge {{ absence_badge_class(a) }}">{{ absence_display_label(a) }}</span><br>{% endif %}{% if b.status == 'waiting_list' %}<span class="badge wait">Liste d'attente — rang {{ waitlist_rank(b) }}</span>{% else %}<span class="badge">Réservé</span>{% endif %}</td><td><a class="btn danger" href="{{ url_for('cancel', booking_id=b.id, next=request.full_path) }}">Annuler</a></td></tr>{% else %}<tr><td colspan="4" class="muted">Aucune réservation à venir.</td></tr>{% endfor %}</table></section><br>{% endif %}""",
    1,
)
TEMPLATE_INDEX = TEMPLATE_INDEX.replace(
    """<div class="card" style="box-shadow:none;background:#f9fafb"><h2>Règles de réservation</h2><p>Annulation possible jusqu'à 2h avant le cours.</p><p>Deux absences injustifiées sur 90 jours entraînent un blocage temporaire des réservations.</p><p>Si vous arrivez en retard, la coach peut corriger l'appel : le retard n'entraîne pas de pénalité.</p></div>""",
    """<div class="card" style="box-shadow:none;background:#f9fafb"><h2>Règles de réservation</h2><p>Les cours sont créés automatiquement 28 jours avant leur date.</p><p>Pour les créneaux réservables, les adhérents mensuels disposent d'une priorité de réservation pendant les 7 premiers jours.</p><p>Après ces 7 jours, les places restantes sont ouvertes à tous les statuts : cadres et autres peuvent alors réserver jusqu'à 21 jours avant la date du cours, selon les places disponibles.</p><p>Chaque adhérent est autonome pour réserver et annuler ses créneaux depuis son profil. Les membres du Bureau Fitness n'ont pas la main pour annuler une réservation à la place d'un adhérent.</p><p>Annulation possible jusqu'à 2h avant le cours.</p><p>Deux absences injustifiées sur 90 jours entraînent un blocage temporaire des réservations.</p><p>Si vous arrivez en retard, la coach peut corriger l'appel : le retard n'entraîne pas de pénalité.</p></div>""",
    1,
)

TEMPLATE_MEMBER_PROFILE = """
{% set content %}<div class="card form-wrap"><h1>Mon profil</h1><p class="muted">Modifier votre statut adhérent, vos préférences ou votre photo de profil.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}{% if current_user.profile_photo or current_user.profile_photo_data %}<img class="photo-preview" src="{{ url_for('profile_photo_file', user_id=current_user.id) }}" alt="Photo profil"><br><br>{% endif %}<form method="post" enctype="multipart/form-data"><div class="form-grid"><div class="field"><label>Nom complet</label><input value="{{ current_user.display_name() }}" disabled></div><div class="field"><label>Email</label><input value="{{ current_user.email }}" disabled></div><div class="field"><label>Statut prioritaire</label><select name="status"><option value="mensuel" {% if current_user.status == 'mensuel' %}selected{% endif %}>Mensuel</option><option value="cadre" {% if current_user.status == 'cadre' %}selected{% endif %}>Cadre</option><option value="autre" {% if current_user.status == 'autre' %}selected{% endif %}>Autre</option></select></div><div class="field"><label>Profil adhérent</label><select name="member_profile"><option value="ouvrant_droit" {% if current_user.member_profile == 'ouvrant_droit' or not current_user.member_profile %}selected{% endif %}>Ouvrant droit - personnel Thales, alternant, stagiaire, CDD</option><option value="ayant_droit" {% if current_user.member_profile == 'ayant_droit' %}selected{% endif %}>Ayant droit - proche d'un ouvrant droit</option><option value="exterieur" {% if current_user.member_profile == 'exterieur' %}selected{% endif %}>Extérieur - prestataire sur site Thales</option><option value="retraite" {% if current_user.member_profile == 'retraite' %}selected{% endif %}>Retraité</option></select></div><div class="field" style="grid-column:1/-1"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label><input name="rights_holder_name" value="{{ current_user.rights_holder_name or '' }}"></div><div class="field"><label>Cours préféré</label><select name="preferred_course"><option value="">-</option>{% for name in preference_options.courses %}<option value="{{ name }}" {% if current_user.preferred_course == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field"><label>Coach préféré</label><select name="preferred_coach"><option value="">-</option>{% for name in preference_options.coaches %}<option value="{{ name }}" {% if current_user.preferred_coach == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field"><label>Créneau préféré</label><select name="preferred_slot"><option value="">-</option>{% for name in preference_options.slots %}<option value="{{ name }}" {% if current_user.preferred_slot == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div><div class="field" style="grid-column:1/-1"><label>Nouvelle photo de profil JPG/PNG, facultative</label><input name="profile_photo" type="file" accept="image/png,image/jpeg"></div></div><br><button class="btn" type="submit">Enregistrer</button> <a class="btn secondary" href="{{ url_for('download_card', user_id=current_user.id) }}">Télécharger ma carte</a></form></div>{% endset %}{{ shell(content, 'member_profile')|safe }}
"""
TEMPLATE_MEMBER_PROFILE = TEMPLATE_MEMBER_PROFILE.replace(
    """<div class="field"><label>Email</label><input value="{{ current_user.email }}" disabled></div><div class="field"><label>Statut prioritaire</label>""",
    """<div class="field"><label>Email</label><input value="{{ current_user.email }}" disabled></div><div class="field"><label>Abonnement</label>{% if current_user.role == 'admin' %}<select name="subscription_type">{% for opt in ['Annuel','Semestre 1','Semestre 2','Trimestre 1','T2','T3','T4'] %}<option {% if current_user.subscription_type == opt %}selected{% endif %}>{{ opt }}</option>{% endfor %}</select>{% else %}<input value="{{ current_user.subscription_type or '-' }}" disabled>{% endif %}</div><div class="field"><label>Année d'abonnement</label>{% if current_user.role == 'admin' %}<input name="subscription_year" type="number" min="2024" max="2100" value="{{ current_user.subscription_year or current_year }}">{% else %}<input value="{{ current_user.subscription_year or '-' }}" disabled>{% endif %}</div><div class="field"><label>Statut prioritaire</label>""",
    1,
)
TEMPLATE_MEMBER_PROFILE = TEMPLATE_MEMBER_PROFILE.replace(
    """<div class="field"><label>Statut prioritaire</label><select name="status"><option value="mensuel" {% if current_user.status == 'mensuel' %}selected{% endif %}>Mensuel</option><option value="cadre" {% if current_user.status == 'cadre' %}selected{% endif %}>Cadre</option><option value="autre" {% if current_user.status == 'autre' %}selected{% endif %}>Autre</option></select></div>""",
    """<div class="field"><label>Statut prioritaire</label>{% if current_user.role == 'admin' %}<select name="status"><option value="mensuel" {% if current_user.status == 'mensuel' %}selected{% endif %}>Mensuel</option><option value="cadre" {% if current_user.status == 'cadre' %}selected{% endif %}>Cadre</option><option value="autre" {% if current_user.status == 'autre' %}selected{% endif %}>Autre</option></select>{% else %}<input value="{% if current_user.status == 'mensuel' %}Mensuel{% elif current_user.status == 'cadre' %}Cadre{% else %}Autre{% endif %}" disabled>{% endif %}</div>""",
    1,
)
TEMPLATE_MEMBER_PROFILE = TEMPLATE_MEMBER_PROFILE.replace(
    """<div class="field"><label>Nom complet</label><input value="{{ current_user.display_name() }}" disabled></div>""",
    """<div class="field"><label>Prénom</label><input name="first_name" value="{{ current_user.first_name or split_name(current_user.full_name)[0] }}" required></div><div class="field"><label>Nom</label><input name="last_name" value="{{ current_user.last_name or split_name(current_user.full_name)[1] }}" required></div>""",
    1,
)

TEMPLATE_REGISTER = """
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">__STYLE__<title>Créer un compte</title></head><body><div class="login"><div class="card form-wrap"><h1>Créer un compte adhérent</h1><p class="muted">La carte adhérent sera générée automatiquement et jointe à l'email de confirmation.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" enctype="multipart/form-data"><div class="form-grid"><div class="field"><label>Nom complet</label><input name="full_name" required></div><div class="field"><label>Email</label><input name="email" type="email" required></div><div class="field"><label>Mot de passe</label><input name="password" type="password" required></div><div class="field"><label>Statut adhérent</label><select name="status"><option value="mensuel">Mensuel</option><option value="cadre">Cadre</option><option value="autre">Autre</option></select></div><div class="field"><label>Profil adhérent</label><select name="member_profile"><option value="ouvrant_droit">Ouvrant droit - personnel Thales, alternant, stagiaire, CDD</option><option value="ayant_droit">Ayant droit - proche d'un ouvrant droit</option><option value="exterieur">Extérieur - prestataire sur site Thales</option><option value="retraite">Retraité</option></select></div><div class="field"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label><input name="rights_holder_name" placeholder="Ex. Marie Dupont"></div><div class="field"><label>Type d'abonnement</label><select name="subscription_type" required><option>Annuel</option><option>Semestre 1</option><option>Semestre 2</option><option>Trimestre 1</option><option>Trimestre 2</option><option>Trimestre 3</option><option>Trimestre 4</option></select></div><div class="field"><label>Année</label><input name="subscription_year" type="number" min="2024" max="2100" value="{{ current_year }}" required></div><div class="field" style="grid-column:1/-1"><label>Photo de profil JPG/PNG</label><input name="profile_photo" type="file" accept="image/png,image/jpeg" required></div></div><br><button class="btn" type="submit">Créer le compte</button> <a class="btn secondary" href="{{ url_for('login') }}">Déjà un compte ?</a></form></div></div></body></html>
""".replace("__STYLE__", BASE_TEMPLATE_STYLE)
TEMPLATE_REGISTER = TEMPLATE_REGISTER.replace(
    """<div class="field"><label>Nom complet</label><input name="full_name" required></div>""",
    """<div class="field"><label>Prénom</label><input name="first_name" required></div><div class="field"><label>Nom</label><input name="last_name" required></div>""",
    1,
)
TEMPLATE_REGISTER = TEMPLATE_REGISTER.replace(
    """<option>Trimestre 2</option><option>Trimestre 3</option><option>Trimestre 4</option>""",
    """<option>T2</option><option>T3</option><option>T4</option>""",
    1,
)

TEMPLATE_LOGIN = """
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">__STYLE__<title>Connexion</title></head><body><div class="login"><div class="card login-box"><h1>Connexion</h1>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post"><div class="field"><label>Email</label><input name="email" type="email" required></div><br><div class="field"><label>Mot de passe</label><input name="password" type="password" required></div><br><button class="btn" type="submit">Connexion</button></form><p><a href="{{ url_for('forgot_password') }}">Mot de passe oublié ?</a></p></div></div></body></html>
""".replace("__STYLE__", BASE_TEMPLATE_STYLE)

TEMPLATE_GENERATE = """
{% set content %}<div class="card form-wrap"><h1>Générer les créneaux</h1>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post"><div class="form-grid"><div class="field"><label>Année</label><input name="year" type="number" value="{{ current_year }}" required></div><div class="field"><label>Mois</label><input name="month" type="number" min="1" max="12" value="{{ current_month }}" required></div></div><br><button class="btn" type="submit">Générer manuellement ce mois</button> <a class="btn secondary" href="{{ url_for('index') }}">Retour</a></form></div>{% endset %}{{ shell(content, 'generate')|safe }}
"""

TEMPLATE_SESSION_DETAIL = """
{% set content %}<style>.attendance-list{display:grid;gap:14px}.attendance-card{display:grid;grid-template-columns:76px 1fr;gap:14px;align-items:center;border:1px solid #e5e7eb;border-radius:16px;padding:14px;background:#fff}.attendance-photo{width:76px;height:76px;border-radius:16px;object-fit:cover;background:#e5e7eb;border:1px solid #e5e7eb}.attendance-name{font-size:20px;font-weight:800;margin-bottom:4px}.attendance-actions{grid-column:1/-1;display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}.attendance-actions .btn{text-align:center;padding:13px 8px}.btn.warn{background:#f59e0b}@media(max-width:700px){.main{padding:14px}.card{padding:16px;border-radius:14px}.attendance-card{grid-template-columns:68px 1fr;padding:12px}.attendance-photo{width:68px;height:68px}.attendance-name{font-size:18px}.attendance-actions{grid-template-columns:1fr;gap:8px}.attendance-actions .btn{width:100%;font-size:16px}}</style><div class="card"><h1>{{ session.course_name }}</h1><p class="muted">{{ session.course_date.strftime('%d/%m/%Y') }} · {{ session.start_time.strftime('%H:%M') }} - {{ session.end_time.strftime('%H:%M') }}</p><p class="muted">Appel mobile : traiter chaque adhérent, ou utiliser “Passer” pour revenir dessus après.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<div class="attendance-list">{% for b in bookings %}<div class="attendance-card">{% if b.user.profile_photo or b.user.profile_photo_data %}<img class="attendance-photo" src="{{ url_for('profile_photo_file', user_id=b.user.id) }}" alt="Photo {{ b.user.display_name() }}">{% else %}<div class="attendance-photo"></div>{% endif %}<div><div class="attendance-name">{{ b.user.display_name() }}</div><div class="muted">{{ b.user.email }}</div><div style="margin-top:8px"><span class="badge {{ attendance_badge_class(b) }}">{{ attendance_label(b) }}</span>{% if b.status == 'waiting_list' %} <span class="badge wait">rang {{ waitlist_rank(b) }}</span>{% endif %}</div></div><div class="attendance-actions">{% if b.status in ['booked','absent_unexcused'] %}<a class="btn" href="{{ url_for('mark_present', booking_id=b.id) }}">Présent</a>{% if b.status == 'booked' %}<a class="btn warn" href="{{ url_for('mark_skipped', booking_id=b.id) }}">Passer</a><a class="btn danger" href="{{ url_for('mark_absent', booking_id=b.id) }}" onclick="return confirm('Marquer cet adhérent absent ?')">Absent</a>{% else %}<span class="btn secondary">Passer</span><span class="btn danger">Absent</span>{% endif %}{% else %}<span class="muted">Liste d'attente, pas d'appel.</span>{% endif %}</div></div>{% else %}<p class="muted">Aucune réservation confirmée.</p>{% endfor %}</div><br><a class="btn secondary" href="{{ url_for('index') }}">Retour</a></div>{% endset %}{{ shell(content, 'home')|safe }}
"""

TEMPLATE_SESSION_DETAIL = """
{% set content %}<style>.attendance-list{display:grid;gap:14px}.attendance-card{display:grid;grid-template-columns:76px 1fr;gap:14px;align-items:center;border:1px solid #e5e7eb;border-radius:16px;padding:14px;background:#fff}.attendance-card.absent{border-left:6px solid var(--danger);background:#fff7f7}.attendance-photo{width:76px;height:76px;border-radius:16px;object-fit:cover;background:#e5e7eb;border:1px solid #e5e7eb}.attendance-name{font-size:20px;font-weight:800;margin-bottom:4px}.attendance-actions{grid-column:1/-1;display:grid;grid-template-columns:1fr;gap:10px}.attendance-actions .btn{text-align:center;padding:14px 8px}.btn.warn{background:#f59e0b}@media(max-width:700px){.main{padding:14px}.card{padding:16px;border-radius:14px}.attendance-card{grid-template-columns:68px 1fr;padding:12px}.attendance-photo{width:68px;height:68px}.attendance-name{font-size:18px}.attendance-actions .btn{width:100%;font-size:16px}}</style><div class="card"><h1>{{ session.course_name }}</h1><p class="muted">{{ session.course_date.strftime('%d/%m/%Y') }} · {{ session.start_time.strftime('%H:%M') }} - {{ session.end_time.strftime('%H:%M') }}</p><p class="muted">Appel mobile : marquez uniquement les absents. Si une personne arrive après l'appel, utilisez “Retard” pour retirer la pénalité.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<div class="attendance-list">{% for b in bookings %}<div class="attendance-card {% if b.status == 'absent_unexcused' or b.attendance_status == 'absent' %}absent{% endif %}">{% if b.user.profile_photo or b.user.profile_photo_data %}<img class="attendance-photo" src="{{ url_for('profile_photo_file', user_id=b.user.id) }}" alt="Photo {{ b.user.display_name() }}">{% else %}<div class="attendance-photo"></div>{% endif %}<div><div class="attendance-name">{{ b.user.display_name() }}</div><div class="muted">{{ b.user.email }}</div><div style="margin-top:8px"><span class="badge {{ attendance_badge_class(b) }}">{{ attendance_label(b) }}</span>{% if b.status == 'waiting_list' %} <span class="badge wait">rang {{ waitlist_rank(b) }}</span>{% endif %}</div></div><div class="attendance-actions">{% if b.status == 'booked' %}<a class="btn danger" href="{{ url_for('mark_absent', booking_id=b.id) }}" onclick="return confirm('Marquer cet adhérent absent ?')">Absent</a>{% elif b.status == 'absent_unexcused' %}<a class="btn warn" href="{{ url_for('mark_late', booking_id=b.id) }}">Retard</a>{% else %}<span class="muted">Liste d'attente, pas d'appel.</span>{% endif %}</div></div>{% else %}<p class="muted">Aucune réservation confirmée.</p>{% endfor %}</div><br><a class="btn secondary" href="{{ url_for('index') }}">Retour</a></div>{% endset %}{{ shell(content, 'home')|safe }}
"""

TEMPLATE_MEMBERS = """
{% set content %}<div class="card"><div class="top"><div><h1>Adhérents</h1><p class="muted">Annuaire des adhérents pour suivi, modification, réservations et campagnes d'emailing.</p></div><div><a class="btn" href="{{ url_for('admin_create_member') }}">Créer un adhérent</a> <a class="btn secondary" href="{{ url_for('admin_import_members') }}">Import Excel</a> <a class="btn secondary" href="{{ url_for('export_members_excel') }}">Export adhérents</a> <a class="btn" href="{{ url_for('admin_email_members') }}">Campagne email</a></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="get" action="{{ url_for('admin_email_members') }}"><table class="table"><tr><th><input type="checkbox" onclick="document.querySelectorAll('.member-check').forEach(c=>c.checked=this.checked)"></th><th>Photo</th><th>Nom</th><th>Email</th><th>Statut</th><th>Profil</th><th>Abonnement</th><th>ID</th><th>Absences 90j</th><th>Compte</th><th>Blocage</th><th>Actions</th></tr>{% for u in users %}<tr><td><input class="member-check" type="checkbox" name="user_ids" value="{{ u.id }}"></td><td>{% if u.profile_photo or u.profile_photo_data %}<img class="admin-photo" src="{{ url_for('profile_photo_file', user_id=u.id) }}" alt="Photo {{ u.display_name() }}">{% else %}<span class="muted">-</span>{% endif %}</td><td>{{ u.display_name() }}</td><td><a href="mailto:{{ u.email }}">{{ u.email }}</a></td><td>{{ u.status }}</td><td>{{ u.member_profile or '-' }}{% if u.rights_holder_name %}<br><small>{{ u.rights_holder_name }}</small>{% endif %}</td><td>{{ u.subscription_type or '-' }} {{ u.subscription_year or '' }}</td><td>{{ u.member_number or '-' }}</td><td>{{ absence_count(u) }}</td><td>{% if u.account_status == 'pending' %}<span class="badge wait">activation à faire</span>{% else %}<span class="badge">{{ u.account_status }}</span>{% endif %}</td><td>{% if u.is_blocked() %}<span class="badge full">bloqué jusqu'au {{ u.blocked_until }}</span>{% else %}<span class="badge">non bloqué</span>{% endif %}</td><td><a class="btn secondary" href="{{ url_for('admin_edit_member', user_id=u.id) }}">Modifier</a> <a class="btn secondary" href="{{ url_for('admin_member_reservations', user_id=u.id) }}">Réservations</a> <a class="btn secondary" href="{{ url_for('admin_send_activation', user_id=u.id) }}">Lien activation</a> <a class="btn secondary" href="{{ url_for('admin_send_password_reset', user_id=u.id) }}">Réinitialiser MDP</a> <a class="btn secondary" href="{{ url_for('download_card', user_id=u.id) }}">Générer carte</a> {% if u.role == 'adherent' %}<a class="btn danger" href="{{ url_for('admin_delete_member', user_id=u.id) }}" onclick="return confirm('Supprimer cet adhérent et ses réservations ?')">Supprimer</a>{% else %}<span class="badge wait">Admin adhérent</span>{% endif %}</td></tr>{% else %}<tr><td colspan="12" class="muted">Aucun adhérent.</td></tr>{% endfor %}</table><br><button class="btn" type="submit">Écrire aux adhérents sélectionnés</button></form></div>{% endset %}{{ shell(content, 'members')|safe }}
"""
TEMPLATE_MEMBERS = TEMPLATE_MEMBERS.replace(
    """{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="get" action="{{ url_for('admin_email_members') }}">""",
    """{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="get" action="{{ url_for('admin_members') }}" class="card" style="box-shadow:none;background:#f9fafb"><h3>Filtres</h3><div class="form-grid"><div class="field"><label>Recherche</label><input name="search" value="{{ filter_values.search }}" placeholder="Nom, email, ID"></div><div class="field"><label>Profil</label><select name="member_profile"><option value="">Tous</option>{% for key, label in member_profile_labels.items() %}<option value="{{ key }}" {% if filter_values.member_profile == key %}selected{% endif %}>{{ label }}</option>{% endfor %}</select></div><div class="field"><label>Abonnement</label><select name="subscription_type"><option value="">Tous</option>{% for opt in subscription_options %}<option value="{{ opt }}" {% if filter_values.subscription_type == opt %}selected{% endif %}>{{ opt }}</option>{% endfor %}</select></div><div class="field"><label>Année</label><input name="subscription_year" type="number" value="{{ filter_values.subscription_year }}" placeholder="2026"></div><div class="field"><label>Compte</label><select name="account_status"><option value="">Tous</option><option value="active" {% if filter_values.account_status == 'active' %}selected{% endif %}>Actif</option><option value="pending" {% if filter_values.account_status == 'pending' %}selected{% endif %}>Activation à faire</option></select></div></div><br><button class="btn secondary" type="submit">Filtrer</button> <a class="btn secondary" href="{{ url_for('admin_members') }}">Réinitialiser</a></form><br><form method="get" action="{{ url_for('admin_email_members') }}">""",
    1,
)
TEMPLATE_MEMBERS = TEMPLATE_MEMBERS.replace(
    """</form><br><form method="get" action="{{ url_for('admin_email_members') }}">""",
    """</form><br><details class="card" style="box-shadow:none;background:#f9fafb" {% if filter_values.followup_year or filter_values.followup_start or filter_values.followup_end %}open{% endif %}><summary style="cursor:pointer;font-weight:800;font-size:18px">Suivi inscriptions / renouvellements</summary><p class="muted">Affichage limité aux 25 dernières lignes de la période choisie. Les montants sont figés à la date d'inscription.</p><form method="get" action="{{ url_for('admin_members') }}"><div class="form-grid"><div class="field"><label>Année d'adhésion</label><input name="followup_year" type="number" value="{{ filter_values.followup_year }}" placeholder="2026"></div><div class="field"><label>Date action début</label><input name="followup_start" type="date" value="{{ filter_values.followup_start }}"></div><div class="field"><label>Date action fin</label><input name="followup_end" type="date" value="{{ filter_values.followup_end }}"></div></div><br><button class="btn secondary" type="submit">Afficher le suivi</button> <a class="btn secondary" href="{{ url_for('admin_members') }}">Réinitialiser</a> <a class="btn" href="{{ url_for('export_membership_followup', followup_year=filter_values.followup_year, followup_start=filter_values.followup_start, followup_end=filter_values.followup_end) }}">Exporter ce suivi</a></form><br><table class="table"><tr><th>Date action</th><th>Adhérent</th><th>Abonnement</th><th>Période</th><th>Tarif abo</th><th>Cotisation</th><th>Total</th><th>Créé par</th><th>Note</th></tr>{% for row in membership_actions %}{% set p = row.period %}<tr><td>{{ p.created_at.strftime('%d/%m/%Y %H:%M') if p.created_at else '-' }}</td><td>{{ row.user.display_name() }}<br><small class="muted">{{ row.user.email }}</small></td><td>{{ row.subscription_type }} {{ row.subscription_year }}</td><td>{{ p.start_date.strftime('%d/%m/%Y') }} - {{ p.end_date.strftime('%d/%m/%Y') }}</td><td>{{ '%.2f'|format(row.subscription_price or 0) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee or 0) }} €{% else %}<span class="muted">Non</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total or 0) }} €</strong></td><td>{{ p.created_by or '-' }}</td><td>{{ p.notes or '' }}</td></tr>{% else %}<tr><td colspan="9" class="muted">Aucune action d'adhésion sur cette période.</td></tr>{% endfor %}</table></details><br><form method="get" action="{{ url_for('admin_email_members') }}">""",
    1,
)



TEMPLATE_ADMIN_CREATE_MEMBER = """
{% set content %}<div class="card form-wrap"><h1>Créer un adhérent</h1><p class="muted">Création manuelle depuis le profil admin. La carte adhérent est générée automatiquement ; la photo est recommandée mais non obligatoire pour une création admin.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" enctype="multipart/form-data"><div class="form-grid"><div class="field"><label>Nom complet</label><input name="full_name" required></div><div class="field"><label>Email</label><input name="email" type="email" required></div><div class="field"><label>Mot de passe provisoire</label><input name="password" type="text" value="fitness123" required></div><div class="field"><label>Statut prioritaire</label><select name="status"><option value="mensuel">Mensuel</option><option value="cadre">Cadre</option><option value="autre">Autre</option></select></div><div class="field"><label>Profil adhérent</label><select name="member_profile"><option value="ouvrant_droit">Ouvrant droit - personnel Thales, alternant, stagiaire, CDD</option><option value="ayant_droit">Ayant droit - proche d'un ouvrant droit</option><option value="exterieur">Extérieur - prestataire sur site Thales</option><option value="retraite">Retraité</option></select></div><div class="field"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label><input name="rights_holder_name" placeholder="Ex. Marie Dupont"></div><div class="field"><label>Type d'abonnement</label><select name="subscription_type" required><option>Annuel</option><option>Semestre 1</option><option>Semestre 2</option><option>Trimestre 1</option><option>Trimestre 2</option><option>Trimestre 3</option><option>Trimestre 4</option></select></div><div class="field"><label>Année</label><input name="subscription_year" type="number" min="2024" max="2100" value="{{ current_year }}" required></div><div class="field" style="grid-column:1/-1"><label>Photo de profil JPG/PNG</label><input name="profile_photo" type="file" accept="image/png,image/jpeg"></div></div><br><button class="btn" type="submit">Créer l'adhérent</button> <a class="btn secondary" href="{{ url_for('admin_members') }}">Retour</a></form></div>{% endset %}{{ shell(content, 'members')|safe }}
"""
TEMPLATE_ADMIN_CREATE_MEMBER = TEMPLATE_ADMIN_CREATE_MEMBER.replace(
    """<p class="muted">Création manuelle depuis le profil admin. La carte adhérent est générée automatiquement ; la photo est recommandée mais non obligatoire pour une création admin.</p>""",
    """<p class="muted">Création rapide : l'email et l'abonnement suffisent. L'adhérent complétera son nom, prénom, photo, catégorie et statut via le lien d'activation.</p>""",
    1,
)
TEMPLATE_ADMIN_CREATE_MEMBER = TEMPLATE_ADMIN_CREATE_MEMBER.replace(
    """<div class="field"><label>Nom complet</label><input name="full_name" required></div><div class="field"><label>Email</label>""",
    """<div class="field"><label>Prénom</label><input name="first_name"></div><div class="field"><label>Nom</label><input name="last_name"></div><div class="field"><label>Email</label>""",
    1,
)
TEMPLATE_ADMIN_CREATE_MEMBER = TEMPLATE_ADMIN_CREATE_MEMBER.replace(
    """<div class="field"><label>Mot de passe provisoire</label><input name="password" type="text" value="fitness123" required></div>""",
    "",
    1,
)
TEMPLATE_ADMIN_CREATE_MEMBER = TEMPLATE_ADMIN_CREATE_MEMBER.replace(
    """<label>Photo de profil JPG/PNG</label><input name="profile_photo" type="file" accept="image/png,image/jpeg">""",
    """<label>Photo de profil JPG/PNG, facultative</label><input name="profile_photo" type="file" accept="image/png,image/jpeg">""",
    1,
)
TEMPLATE_ADMIN_CREATE_MEMBER = TEMPLATE_ADMIN_CREATE_MEMBER.replace(
    """<option>Trimestre 2</option><option>Trimestre 3</option><option>Trimestre 4</option>""",
    """<option>T2</option><option>T3</option><option>T4</option>""",
    1,
)

TEMPLATE_ADMIN_EDIT_MEMBER = """
{% set content %}<div class="card form-wrap"><h1>Modifier l'adhérent</h1><p class="muted">Mettez à jour les informations administratives de l'adhérent. Si vous ajoutez une nouvelle photo, la carte adhérent est régénérée.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}{% if user.profile_photo or user.profile_photo_data %}<div class="card" style="box-shadow:none;background:#f9fafb"><strong>Photo rattachée à cet adhérent</strong><br><br><img class="photo-preview" src="{{ url_for('profile_photo_file', user_id=user.id) }}" alt="Photo {{ user.display_name() }}"> <a class="btn secondary" href="{{ url_for('download_card', user_id=user.id) }}">Générer la carte avec cette photo</a></div><br>{% endif %}<form method="post" enctype="multipart/form-data"><div class="form-grid"><div class="field"><label>Nom complet</label><input name="full_name" value="{{ user.full_name or '' }}" required></div><div class="field"><label>Email</label><input name="email" type="email" value="{{ user.email }}" required></div><div class="field"><label>Nouveau mot de passe, facultatif</label><input name="password" type="text" placeholder="Laisser vide pour ne pas modifier"></div><div class="field"><label>Statut prioritaire</label><select name="status"><option value="mensuel" {% if user.status == 'mensuel' %}selected{% endif %}>Mensuel</option><option value="cadre" {% if user.status == 'cadre' %}selected{% endif %}>Cadre</option><option value="autre" {% if user.status == 'autre' %}selected{% endif %}>Autre</option></select></div><div class="field"><label>Profil adhérent</label><select name="member_profile"><option value="ouvrant_droit" {% if user.member_profile == 'ouvrant_droit' or not user.member_profile %}selected{% endif %}>Ouvrant droit - personnel Thales, alternant, stagiaire, CDD</option><option value="ayant_droit" {% if user.member_profile == 'ayant_droit' %}selected{% endif %}>Ayant droit - proche d'un ouvrant droit</option><option value="exterieur" {% if user.member_profile == 'exterieur' %}selected{% endif %}>Extérieur - prestataire sur site Thales</option><option value="retraite" {% if user.member_profile == 'retraite' %}selected{% endif %}>Retraité</option></select></div><div class="field"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label><input name="rights_holder_name" value="{{ user.rights_holder_name or '' }}" placeholder="Ex. Marie Dupont"></div><div class="field"><label>Type d'abonnement</label><select name="subscription_type" required>{% for opt in ['Annuel','Semestre 1','Semestre 2','Trimestre 1','Trimestre 2','Trimestre 3','Trimestre 4'] %}<option {% if user.subscription_type == opt %}selected{% endif %}>{{ opt }}</option>{% endfor %}</select></div><div class="field"><label>Année</label><input name="subscription_year" type="number" min="2024" max="2100" value="{{ user.subscription_year or current_year }}" required></div><div class="field" style="grid-column:1/-1"><label>Nouvelle photo de profil JPG/PNG, facultative</label><input name="profile_photo" type="file" accept="image/png,image/jpeg"></div></div><br><button class="btn" type="submit">Enregistrer</button> <a class="btn secondary" href="{{ url_for('download_card', user_id=user.id) }}">Générer la carte</a> <a class="btn secondary" href="{{ url_for('admin_members') }}">Retour</a></form></div>{% endset %}{{ shell(content, 'members')|safe }}
"""
TEMPLATE_ADMIN_EDIT_MEMBER = TEMPLATE_ADMIN_EDIT_MEMBER.replace(
    """<div class="field"><label>Nom complet</label><input name="full_name" value="{{ user.full_name or '' }}" required></div>""",
    """<div class="field"><label>Prénom</label><input name="first_name" value="{{ user.first_name or split_name(user.full_name)[0] }}" required></div><div class="field"><label>Nom</label><input name="last_name" value="{{ user.last_name or split_name(user.full_name)[1] }}" required></div>""",
    1,
)
TEMPLATE_ADMIN_EDIT_MEMBER = TEMPLATE_ADMIN_EDIT_MEMBER.replace(
    """<div class="field"><label>Type d'abonnement</label><select name="subscription_type" required>{% for opt in ['Annuel','Semestre 1','Semestre 2','Trimestre 1','Trimestre 2','Trimestre 3','Trimestre 4'] %}<option {% if user.subscription_type == opt %}selected{% endif %}>{{ opt }}</option>{% endfor %}</select></div>""",
    """<div class="field"><label>Type d'abonnement actuel</label><select name="subscription_type" required>{% for opt in subscription_options %}<option {% if user.subscription_type == opt %}selected{% endif %}>{{ opt }}</option>{% endfor %}</select></div>""",
    1,
)
TEMPLATE_ADMIN_EDIT_MEMBER = TEMPLATE_ADMIN_EDIT_MEMBER.replace(
    """<a class="btn secondary" href="{{ url_for('download_card', user_id=user.id) }}">Générer la carte</a> <a class="btn secondary" href="{{ url_for('admin_members') }}">Retour</a></form></div>{% endset %}""",
    """<a class="btn secondary" href="{{ url_for('download_card', user_id=user.id) }}">Générer la carte</a> <a class="btn secondary" href="{{ url_for('admin_members') }}">Retour</a></form><br><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Renouveler l'adhésion</h2><p class="muted">Ajoute une nouvelle période d'adhésion dans l'historique. Les tarifs sont figés à la date du renouvellement.</p><form method="post" action="{{ url_for('admin_renew_member', user_id=user.id) }}"><div class="form-grid"><div class="field"><label>Nouvel abonnement</label><select name="subscription_type" required>{% for opt in subscription_options %}<option>{{ opt }}</option>{% endfor %}</select></div><div class="field"><label>Année</label><input name="subscription_year" type="number" min="2024" max="2100" value="{{ current_year }}" required></div></div><br><button class="btn" type="submit">Renouveler</button></form><br><h3>Historique adhésions</h3><table class="table"><tr><th>Abonnement</th><th>Période</th><th>Tarif abo</th><th>Cotisation</th><th>Total</th><th>Créé par</th><th>Note</th></tr>{% for row in membership_periods %}{% set p = row.period %}<tr><td>{{ row.subscription_type }} {{ row.subscription_year }}</td><td>{{ p.start_date.strftime('%d/%m/%Y') }} - {{ p.end_date.strftime('%d/%m/%Y') }}</td><td>{{ '%.2f'|format(row.subscription_price or 0) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee or 0) }} €{% else %}<span class="muted">Non</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total or 0) }} €</strong></td><td>{{ p.created_by or '-' }}</td><td>{{ p.notes or '' }}</td></tr>{% else %}<tr><td colspan="7" class="muted">Aucun historique d'adhésion.</td></tr>{% endfor %}</table></div></div>{% endset %}""",
    1,
)

TEMPLATE_ADMIN_MEMBER_RESERVATIONS = """
{% set content %}<div class="card"><div class="top"><div><h1>Réservations - {{ user.display_name() }}</h1><p class="muted">Réserver ou annuler des créneaux pour cet adhérent depuis le profil admin.</p></div><a class="btn secondary" href="{{ url_for('admin_members') }}">Retour adhérents</a></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<h2>Réservations de l'adhérent</h2><table class="table"><tr><th>Date</th><th>Horaire</th><th>Cours</th><th>Statut</th><th>Action</th></tr>{% for b in bookings %}<tr><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}</td><td>{{ b.session.start_time.strftime('%H:%M') }} - {{ b.session.end_time.strftime('%H:%M') }}</td><td>{{ b.session.course_name }}</td><td>{% if b.status == 'waiting_list' %}<span class="badge wait">Liste d’attente — rang {{ waitlist_rank(b) }}</span>{% else %}<span class="badge {% if b.status == 'absent_unexcused' %}full{% endif %}">{{ b.status }}</span>{% endif %}</td><td>{% if b.status in ['booked','waiting_list'] %}<a class="btn danger" href="{{ url_for('admin_cancel_member_booking', user_id=user.id, booking_id=b.id) }}" onclick="return confirm('Annuler cette réservation ?')">Annuler</a>{% else %}<span class="muted">-</span>{% endif %}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune réservation.</td></tr>{% endfor %}</table><br><h2>Créneaux ouverts</h2><table class="table"><tr><th>Date</th><th>Horaire</th><th>Cours</th><th>Jauge</th><th>Priorité</th><th>Action</th></tr>{% for s in sessions %}<tr><td>{{ s.course_date.strftime('%d/%m/%Y') }}</td><td>{{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</td><td>{{ s.course_name }}</td><td>{{ booked_count(s) }} / {{ s.capacity }}</td><td>{% if monday_midday_priority_applies(s) %}<span class="badge wait">priorité mensuels jusqu'au {{ s.priority_until.strftime('%d/%m/%Y') }}</span>{% else %}<span class="muted">ouverte</span>{% endif %}</td><td>{% if s.id in active_session_ids %}<span class="muted">Déjà inscrit</span>{% else %}<a class="btn" href="{{ url_for('admin_book_for_member', user_id=user.id, session_id=s.id) }}">Réserver</a>{% endif %}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucun créneau à venir.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'members')|safe }}
"""

TEMPLATE_COACH_LOGIN = """
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">__STYLE__<title>Accès coach</title></head><body><div class="login"><div class="card login-box"><h1>Accès coach</h1><p class="muted">Connectez-vous avec l'email coach et le mot de passe créé depuis le lien d'activation.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post"><div class="field"><label>Email coach</label><input name="email" type="email" required></div><br><div class="field"><label>Mot de passe</label><input name="password" type="password" required></div><br><button class="btn" type="submit">Accéder à mon profil coach</button> <a class="btn secondary" href="{{ url_for('login') }}">Retour</a></form><p><a href="{{ url_for('forgot_password') }}">Mot de passe oublié ?</a></p></div></div></body></html>
""".replace("__STYLE__", BASE_TEMPLATE_STYLE)

TEMPLATE_COACHES = """
{% set content %}<div class="card form-wrap"><h1>Coachs</h1><p class="muted">Enregistrez ici les coachs. Un lien unique leur est envoyé pour créer leur mot de passe avant connexion.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post"><div class="form-grid"><div class="field"><label>Nom coach</label><input name="full_name" placeholder="Ex. Coach Fitness"></div><div class="field"><label>Email coach</label><input name="email" type="email" required></div></div><br><button class="btn" type="submit">Ajouter et envoyer le lien</button></form><br><table class="table"><tr><th>Nom</th><th>Email</th><th>Rôle</th><th>Compte</th><th>Action</th></tr>{% for c in coaches %}<tr><td>{{ c.display_name() }}</td><td>{{ c.email }}</td><td>{{ c.coach_type or 'titulaire' }}</td><td>{% if c.account_status == 'pending' %}<span class="badge wait">activation à faire</span>{% else %}<span class="badge">{{ c.account_status }}</span>{% endif %}</td><td><a class="btn secondary" href="{{ url_for('admin_coach_planning') }}">Planning admin</a> <a class="btn secondary" href="{{ url_for('coach_profile', coach_name=c.display_name()) }}">Absence</a> <a class="btn secondary" href="{{ url_for('admin_send_coach_activation', user_id=c.id) }}">Renvoyer lien</a> <a class="btn danger" href="{{ url_for('admin_delete_coach', user_id=c.id) }}" onclick="return confirm('Supprimer cette coach ?')">Supprimer</a></td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune coach enregistrée.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'coaches')|safe }}
"""

TEMPLATE_COACH_PROFILE = """
{% set content %}<div class="card"><div class="top"><div><h1>Absence</h1><p class="muted">Déclarer les absences, congés ou remplacements.</p></div>{% if current_user.role == 'admin' %}<form method="get"><select name="coach_name" style="padding:10px;border-radius:10px;border:1px solid #ddd">{% for c in coaches %}<option {% if c == coach_name %}selected{% endif %}>{{ c }}</option>{% endfor %}</select> <button class="btn secondary" type="submit">Afficher</button></form>{% endif %}</div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>{{ coach_name }}</h3>{% if current_user.role == 'admin' %}<input type="hidden" name="coach_name" value="{{ coach_name }}">{% endif %}<div class="form-grid"><div class="field"><label>Début</label><input name="start_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Fin</label><input name="end_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="status"><option value="absent">Absence</option><option value="conge">Congé</option><option value="replaced">Remplacé</option></select></div><div class="field"><label>Remplaçant</label><select name="replacement_name"><option value="">-</option>{% for c in replacement_coaches %}<option>{{ c }}</option>{% endfor %}</select></div><div class="field" style="grid-column:1/-1"><label>Notes</label><input name="notes" placeholder="Motif, précision ou consigne interne"></div></div><br><button class="btn" type="submit">Enregistrer</button></form><br><h2>Absences et congés récents / à venir</h2><table class="table"><tr><th>Date</th><th>Type</th><th>Remplaçant</th><th>Notes</th><th>Action</th></tr>{% for a in absences %}<tr><td>{{ a.absence_date.strftime('%d/%m/%Y') }}</td><td><span class="badge {% if a.status in ['absent','conge'] %}full{% elif a.status == 'replaced' %}wait{% endif %}">{{ a.status }}</span></td><td>{{ a.replacement_name or '-' }}</td><td>{{ a.notes or '' }}</td><td><a class="btn danger" href="{{ url_for('delete_coach_absence', absence_id=a.id) }}" onclick="return confirm('Supprimer cette absence ?')">Supprimer</a></td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune absence enregistrée.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'coach_profile')|safe }}
"""

TEMPLATE_COACH_SCHEDULE = """
{% set content %}<div class="card"><div class="top"><div><h1>Mon planning</h1><p class="muted">Cours rattachés à votre compte coach.</p></div><form method="get"><input name="year" type="number" value="{{ year }}" style="width:90px;padding:10px;border-radius:10px;border:1px solid #ddd"> <input name="month" type="number" min="1" max="12" value="{{ month }}" style="width:70px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button></form></div><div class="grid"><div class="card"><span class="muted">Cours rattachés</span><div class="stat">{{ sessions|length }}</div></div><div class="card"><span class="muted">Remplacements</span><div class="stat">{{ replacements|length }}</div></div></div><h2>Cours prévus</h2><table class="table"><tr><th>Date</th><th>Horaire</th><th>Cours</th><th>Réservation</th></tr>{% for s in sessions %}<tr><td>{{ weekday_labels[s.course_date.weekday()] }} {{ s.course_date.strftime('%d/%m/%Y') }}</td><td>{{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</td><td>{{ s.course_name }}</td><td>{% if s.is_reservable %}<span class="badge">Réservable</span>{% else %}<span class="badge wait">Sans résa</span>{% endif %}</td></tr>{% else %}<tr><td colspan="4" class="muted">Aucun cours rattaché sur ce mois.</td></tr>{% endfor %}</table><br><h2>Remplacements</h2><table class="table"><tr><th>Date</th><th>Coach absent</th><th>Cours</th><th>Horaire</th><th>Suivi</th></tr>{% for item in replacements %}<tr><td>{{ weekday_labels[item.absence.absence_date.weekday()] }} {{ item.absence.absence_date.strftime('%d/%m/%Y') }}</td><td>{{ item.absence.coach_name }}</td><td>{% if item.session %}{{ item.session.course_name }}{% else %}<span class="muted">Cours non trouvé</span>{% endif %}</td><td>{% if item.session %}{{ item.session.start_time.strftime('%H:%M') }} - {{ item.session.end_time.strftime('%H:%M') }}{% else %}-{% endif %}</td><td>{{ item.absence.followup_status or '-' }}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucun remplacement sur ce mois.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'coach_schedule')|safe }}
"""

TEMPLATE_STATISTICS = """
{% set content %}<div class="card"><h1>Statistiques</h1><p class="muted">Données utiles pour piloter la section Fitness.</p><div class="grid"><div class="card"><span class="muted">Séances aujourd'hui</span><div class="stat">{{ stats.today_sessions }}</div></div><div class="card"><span class="muted">Réservations</span><div class="stat">{{ stats.bookings }}</div></div><div class="card"><span class="muted">Adhérents</span><div class="stat">{{ stats.members }}</div></div><div class="card"><span class="muted">Bloqués</span><div class="stat">{{ stats.blocked }}</div></div></div><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Préférences adhérents</h2><div class="grid"><div><h3>Cours</h3><table class="table">{% for label, count in preference_stats.course %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Coachs</h3><table class="table">{% for label, count in preference_stats.coach %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Créneaux</h3><table class="table">{% for label, count in preference_stats.slot %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div></div></div><br><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Statistiques section</h2><div class="grid"><div><h3>Adhérents par année</h3><table class="table"><tr><th>Année</th><th>Adhérents</th><th>Évolution</th></tr>{% for row in section_stats.annual %}<tr><td>{{ row.year }}</td><td><strong>{{ row.count }}</strong></td><td>{% if row.evolution is none %}<span class="muted">-</span>{% else %}{{ '%+.1f'|format(row.evolution) }} %{% endif %}</td></tr>{% else %}<tr><td class="muted" colspan="3">Aucune donnée</td></tr>{% endfor %}</table></div><div><h3>Abonnements</h3><table class="table">{% for label, count in section_stats.subscriptions %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Profils</h3><table class="table">{% for label, count in section_stats.profiles %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Statuts</h3><table class="table">{% for label, count in section_stats.statuses %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div></div></div></div>{% endset %}{{ shell(content, 'statistics')|safe }}
"""

TEMPLATE_STATISTICS = """
{% set content %}<div class="card"><h1>Statistiques</h1><p class="muted">Données utiles pour piloter la section Fitness.</p><div class="grid"><div class="card"><span class="muted">Séances aujourd'hui</span><div class="stat">{{ stats.today_sessions }}</div></div><div class="card"><span class="muted">Réservations</span><div class="stat">{{ stats.bookings }}</div></div><div class="card"><span class="muted">Adhérents</span><div class="stat">{{ stats.members }}</div></div><div class="card"><span class="muted">Bloqués</span><div class="stat">{{ stats.blocked }}</div></div></div><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Statistiques réservations par cours</h2><form method="get"><div class="form-grid"><div class="field"><label>Début</label><input type="date" name="start_date" value="{{ filter_values.start_date }}"></div><div class="field"><label>Fin</label><input type="date" name="end_date" value="{{ filter_values.end_date }}"></div><div class="field"><label>Cours</label><select name="course_filter"><option value="">Tous</option>{% for name in course_options %}<option value="{{ name }}" {% if filter_values.course_filter == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div></div><br><button class="btn secondary" type="submit">Filtrer</button> <a class="btn" href="{{ url_for('export_statistics_excel', start_date=filter_values.start_date, end_date=filter_values.end_date, course_filter=filter_values.course_filter) }}">Exporter Excel</a></form><br><h3>Résumé mensuel</h3><table class="table"><tr><th>Mois</th><th>Séances</th><th>Réservations</th><th>Mensuels</th><th>Liste attente</th><th>Ratio mensuels</th></tr>{% for row in course_monthly_rows %}<tr><td>{{ row.month }}</td><td>{{ row.sessions }}</td><td>{{ row.booked }}</td><td>{{ row.mensuels }}</td><td>{{ row.waiting }}</td><td>{{ '%.1f'|format(row.ratio_mensuel) }} %</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune donnée sur cette période.</td></tr>{% endfor %}</table><br><h3>Détail par cours</h3><table class="table"><tr><th>Date</th><th>Cours</th><th>Coach</th><th>Horaire</th><th>Jauge</th><th>Réservés</th><th>Attente</th><th>Absents</th><th>Mensuels</th><th>Cadres/autres</th><th>Ratio mensuels</th><th>Remplissage</th></tr>{% for row in course_rows %}<tr><td>{{ row.date.strftime('%d/%m/%Y') }}</td><td>{{ row.course }}</td><td>{{ row.coach }}</td><td>{{ row.time }}</td><td>{{ row.capacity }}</td><td><strong>{{ row.booked }}</strong></td><td>{{ row.waiting }}</td><td>{{ row.absent }}</td><td>{{ row.mensuels }}</td><td>{{ row.cadres_autres }}</td><td>{{ '%.1f'|format(row.ratio_mensuel) }} %</td><td>{{ '%.1f'|format(row.fill_rate) }} %</td></tr>{% else %}<tr><td colspan="12" class="muted">Aucun cours sur cette période.</td></tr>{% endfor %}</table></div><br><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Préférences adhérents</h2><div class="grid"><div><h3>Cours</h3><table class="table">{% for label, count in preference_stats.course %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Coachs</h3><table class="table">{% for label, count in preference_stats.coach %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Créneaux</h3><table class="table">{% for label, count in preference_stats.slot %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div></div></div><br><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Statistiques section</h2><div class="grid"><div><h3>Adhérents par année</h3><table class="table"><tr><th>Année</th><th>Adhérents</th><th>Évolution</th></tr>{% for row in section_stats.annual %}<tr><td>{{ row.year }}</td><td><strong>{{ row.count }}</strong></td><td>{% if row.evolution is none %}<span class="muted">-</span>{% else %}{{ '%+.1f'|format(row.evolution) }} %{% endif %}</td></tr>{% else %}<tr><td class="muted" colspan="3">Aucune donnée</td></tr>{% endfor %}</table></div><div><h3>Abonnements</h3><table class="table">{% for label, count in section_stats.subscriptions %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Profils</h3><table class="table">{% for label, count in section_stats.profiles %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div><div><h3>Statuts</h3><table class="table">{% for label, count in section_stats.statuses %}<tr><td>{{ label }}</td><td><strong>{{ count }}</strong></td></tr>{% else %}<tr><td class="muted">Aucune donnée</td><td></td></tr>{% endfor %}</table></div></div></div></div>{% endset %}{{ shell(content, 'statistics')|safe }}
"""

TEMPLATE_ARCHIVES = """
{% set content %}<div class="card"><h1>Archives des réservations</h1><p class="muted">Les réservations passées sont archivées automatiquement. Cette page permet de vérifier ultérieurement qui était inscrit à un cours.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<table class="table"><tr><th>Date</th><th>Horaire</th><th>Cours</th><th>Nom</th><th>Email</th><th>Statut réservation</th><th>Date inscription</th></tr>{% for b in bookings %}<tr><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}</td><td>{{ b.session.start_time.strftime('%H:%M') }} - {{ b.session.end_time.strftime('%H:%M') }}</td><td>{{ b.session.course_name }}</td><td>{{ b.user.display_name() }}</td><td>{{ b.user.email }}</td><td><span class="badge {% if b.status == 'absent_unexcused' %}full{% elif b.status == 'waiting_list' %}wait{% endif %}">{{ b.status }}</span></td><td>{{ b.created_at.strftime('%d/%m/%Y %H:%M') if b.created_at else '-' }}</td></tr>{% else %}<tr><td colspan="7" class="muted">Aucune réservation archivée pour l'instant.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'archives')|safe }}
"""

TEMPLATE_EMAIL_MEMBERS = """
{% set content %}<div class="card form-wrap"><h1>Campagne email</h1><p class="muted">Choisissez les groupes destinataires. La signature du Bureau Fitness et le logo sont ajoutés automatiquement.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post">{% for u in users if u.role == 'adherent' %}<input type="hidden" name="user_ids" value="{{ u.id }}">{% endfor %}<div class="card" style="box-shadow:none;background:#f9fafb"><strong>Destinataires</strong><div style="display:flex;gap:16px;flex-wrap:wrap;margin-top:12px"><label><input type="checkbox" name="target_roles" value="adherent" {% if 'adherent' in target_roles %}checked{% endif %}> Adhérents</label><label><input type="checkbox" name="target_roles" value="admin" {% if 'admin' in target_roles %}checked{% endif %}> Admins</label><label><input type="checkbox" name="target_roles" value="coach" {% if 'coach' in target_roles %}checked{% endif %}> Coachs</label></div><p class="muted">{{ users|length }} destinataire(s) actuellement listé(s). Si des adhérents ont été sélectionnés depuis l'onglet Adhérents, seuls ces adhérents sont repris.</p><p class="muted">{% for u in users %}{{ u.display_name() }} &lt;{{ u.email }}&gt;{% if not loop.last %}, {% endif %}{% else %}Aucun destinataire pour cette sélection.{% endfor %}</p></div><br><div class="field"><label>Objet</label><input name="subject" required placeholder="Ex. Informations Section Fitness"></div><br><div class="field"><label>Message</label><textarea name="body" required rows="10" style="width:100%;padding:13px;border:1px solid #d1d5db;border-radius:10px;font-size:15px"></textarea></div><br><button class="btn" type="submit">Envoyer</button> <a class="btn secondary" href="{{ url_for('admin_members') }}">Retour</a></form></div>{% endset %}{{ shell(content, 'members')|safe }}
"""

TEMPLATE_BLOCKED = """
{% set content %}<div class="card"><h1>Adhérents bloqués</h1><p class="muted">Liste des adhérents actuellement bloqués automatiquement.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<table class="table"><tr><th>Nom</th><th>Email</th><th>Jusqu'au</th><th>Motif</th><th>Absences 90j</th><th>Action</th></tr>{% for u in users %}<tr><td>{{ u.display_name() }}</td><td>{{ u.email }}</td><td>{{ u.blocked_until }}</td><td>{{ u.blocked_reason or '-' }}</td><td>{{ absence_count(u) }}</td><td><a class="btn" href="{{ url_for('unblock_member', user_id=u.id) }}">Débloquer</a></td></tr>{% else %}<tr><td colspan="6" class="muted">Aucun adhérent bloqué.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'blocked')|safe }}
"""
TEMPLATE_BLOCKED = """
{% set content %}<div class="card"><h1>Adhérents bloqués</h1><p class="muted">Suivi des blocages automatiques et des absences non excusées à corriger si besoin.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<h2>Comptes bloqués</h2><table class="table"><tr><th>Nom</th><th>Email</th><th>Jusqu'au</th><th>Motif</th><th>Absences 90j</th><th>Action</th></tr>{% for u in users %}<tr><td>{{ u.display_name() }}</td><td>{{ u.email }}</td><td>{{ u.blocked_until }}</td><td>{{ u.blocked_reason or '-' }}</td><td>{{ absence_count(u) }}</td><td><a class="btn" href="{{ url_for('unblock_member', user_id=u.id) }}">Débloquer</a></td></tr>{% else %}<tr><td colspan="6" class="muted">Aucun adhérent bloqué.</td></tr>{% endfor %}</table><br><h2>Absences à vérifier</h2><p class="muted">Cette liste inclut les absences non excusées des 90 derniers jours, même si l'adhérent n'est pas encore bloqué.</p><table class="table"><tr><th>Date</th><th>Horaire</th><th>Cours</th><th>Adhérent</th><th>Email</th><th>Absences 90j</th><th>Action</th></tr>{% for b in recent_absences %}<tr><td>{{ b.session.course_date.strftime('%d/%m/%Y') }}</td><td>{{ b.session.start_time.strftime('%H:%M') }} - {{ b.session.end_time.strftime('%H:%M') }}</td><td>{{ b.session.course_name }}</td><td>{{ b.user.display_name() }}</td><td>{{ b.user.email }}</td><td>{{ absence_count(b.user) }}</td><td><a class="btn secondary" href="{{ url_for('session_detail', session_id=b.session_id) }}">Voir cours</a> <a class="btn danger" href="{{ url_for('remove_unexcused_absence', booking_id=b.id) }}" onclick="return confirm('Retirer ce marquage absent ?')">Retirer absence</a></td></tr>{% else %}<tr><td colspan="7" class="muted">Aucune absence non excusée récente.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'blocked')|safe }}
"""

TEMPLATE_USEFUL_INFO = """
{% set content %}<div class="card"><div class="top"><div><h1>Infos utiles</h1><p class="muted">Documents partagés par le Bureau Fitness.</p></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}{% if current_user.role == 'admin' %}<form method="post" enctype="multipart/form-data" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter un document</h3><div class="form-grid"><div class="field"><label>Titre</label><input name="title" placeholder="Ex. Planning 2026"></div><div class="field"><label>Catégorie</label><input name="category" placeholder="Planning, adhésion, tarifs..."></div><div class="field" style="grid-column:1/-1"><label>Document</label><input name="document_file" type="file" accept=".pdf,.png,.jpg,.jpeg,.doc,.docx,.xls,.xlsx" required></div><div class="field" style="grid-column:1/-1"><label>Notes</label><input name="notes" placeholder="Information complémentaire facultative"></div></div><br><button class="btn" type="submit">Téléverser</button></form><br>{% endif %}<table class="table"><tr><th>Document</th><th>Catégorie</th><th>Notes</th><th>Ajouté le</th><th>Action</th></tr>{% for d in documents %}<tr><td><strong>{{ d.title }}</strong></td><td>{{ d.category or '-' }}</td><td>{{ d.notes or '' }}</td><td>{{ d.uploaded_at.strftime('%d/%m/%Y') if d.uploaded_at else '-' }}</td><td><a class="btn secondary" href="{{ url_for('static', filename=d.file_path) }}" target="_blank">Ouvrir</a>{% if current_user.role == 'admin' %} <a class="btn danger" href="{{ url_for('delete_useful_document', document_id=d.id) }}" onclick="return confirm('Retirer ce document des infos utiles ?')">Supprimer</a>{% endif %}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucun document disponible pour le moment.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'useful_info')|safe }}
"""




TEMPLATE_ACTIVATE = """
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">__STYLE__<title>Activation compte</title></head><body><div class="login"><div class="card login-box"><h1>Activer mon compte</h1><p class="muted">Choisissez votre mot de passe. Vous pouvez aussi ajouter une photo pour votre carte adhérent.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" enctype="multipart/form-data"><div class="field"><label>Mot de passe</label><input name="password" type="password" minlength="8" required></div><br>{% if user.role == 'adherent' %}<div class="field"><label>Profil adhérent</label><select name="member_profile"><option value="ouvrant_droit" {% if (user.member_profile or 'ouvrant_droit') == 'ouvrant_droit' %}selected{% endif %}>Ouvrant droit - personnel Thales, alternant, stagiaire, CDD</option><option value="ayant_droit" {% if user.member_profile == 'ayant_droit' %}selected{% endif %}>Ayant droit - proche d'un ouvrant droit</option><option value="exterieur" {% if user.member_profile == 'exterieur' %}selected{% endif %}>Extérieur - prestataire sur site Thales</option><option value="retraite" {% if user.member_profile == 'retraite' %}selected{% endif %}>Retraité</option></select></div><br><div class="field"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label><input name="rights_holder_name" value="{{ user.rights_holder_name or '' }}"></div><br>{% endif %}<div class="field"><label>Photo de profil (optionnel)</label><input name="profile_photo" type="file" accept="image/png,image/jpeg"></div><br><button class="btn" type="submit">Activer mon compte</button></form></div></div></body></html>
""".replace("__STYLE__", BASE_TEMPLATE_STYLE)
TEMPLATE_ACTIVATE = TEMPLATE_ACTIVATE.replace(
    """{% if user.role == 'adherent' %}<div class="field"><label>Profil adhérent</label>""",
    """{% if user.role == 'adherent' %}<div class="field"><label>Abonnement</label><input value="{{ user.subscription_type or '-' }} {{ user.subscription_year or '' }}" disabled></div><br><div class="field"><label>Profil adhérent</label>""",
    1,
)
TEMPLATE_ACTIVATE = TEMPLATE_ACTIVATE.replace(
    """<div class="field"><label>Abonnement</label><input value="{{ user.subscription_type or '-' }} {{ user.subscription_year or '' }}" disabled></div><br><div class="field"><label>Profil adhérent</label>""",
    """<div class="field"><label>Abonnement</label><input value="{{ user.subscription_type or '-' }} {{ user.subscription_year or '' }}" disabled></div><br><div class="field"><label>Prénom</label><input name="first_name" value="{{ user.first_name or split_name(user.full_name)[0] }}" required></div><br><div class="field"><label>Nom</label><input name="last_name" value="{{ user.last_name or split_name(user.full_name)[1] }}" required></div><br><div class="field"><label>Catégorie</label>""",
    1,
)
TEMPLATE_ACTIVATE = TEMPLATE_ACTIVATE.replace(
    """</select></div><br><div class="field"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label>""",
    """</select></div><br><div class="field"><label>Statut</label><select name="status"><option value="mensuel" {% if user.status == 'mensuel' %}selected{% endif %}>Mensuel</option><option value="cadre" {% if user.status == 'cadre' %}selected{% endif %}>Cadre</option><option value="autre" {% if user.status == 'autre' %}selected{% endif %}>Autre</option></select></div><br><div class="field"><label>Nom et prénom de l'ouvrant droit, si ayant droit</label>""",
    1,
)
TEMPLATE_ACTIVATE = TEMPLATE_ACTIVATE.replace(
    """<div class="field"><label>Photo de profil (optionnel)</label><input name="profile_photo" type="file" accept="image/png,image/jpeg"></div>""",
    """<div class="field"><label>Photo de profil</label><input name="profile_photo" type="file" accept="image/png,image/jpeg" {% if user.role == 'adherent' and not (user.profile_photo or user.profile_photo_data) %}required{% endif %}></div>""",
    1,
)

TEMPLATE_FORGOT_PASSWORD = """
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">__STYLE__<title>Mot de passe oublié</title></head><body><div class="login"><div class="card login-box"><h1>Mot de passe oublié</h1><p class="muted">Renseignez l'email de votre compte. Si le compte existe, un lien de réinitialisation sera envoyé.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post"><div class="field"><label>Email</label><input name="email" type="email" required></div><br><button class="btn" type="submit">Envoyer le lien</button> <a class="btn secondary" href="{{ url_for('login') }}">Retour</a></form></div></div></body></html>
""".replace("__STYLE__", BASE_TEMPLATE_STYLE)

TEMPLATE_RESET_PASSWORD = """
<!doctype html><html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">__STYLE__<title>Nouveau mot de passe</title></head><body><div class="login"><div class="card login-box"><h1>Nouveau mot de passe</h1><p class="muted">Choisissez un nouveau mot de passe pour votre compte.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post"><div class="field"><label>Mot de passe</label><input name="password" type="password" minlength="8" required></div><br><button class="btn" type="submit">Réinitialiser</button></form></div></div></body></html>
""".replace("__STYLE__", BASE_TEMPLATE_STYLE)

TEMPLATE_OFFICE = """
{% set content %}<div class="card"><h1>Bureau / Admins</h1><p class="muted">Nommer ou retirer des administrateurs.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><div class="form-grid"><div class="field"><label>Nom complet</label><input name="full_name"></div><div class="field"><label>Email</label><input name="email" type="email" required></div><div class="field"><label>Fonction</label><select name="admin_role"><option value="presidente">Présidente</option><option value="secretaire_general">Secrétaire général</option><option value="tresoriere">Trésorière</option><option value="membre_bureau">Membre du bureau</option></select></div></div><br><button class="btn" type="submit">Ajouter / nommer admin</button></form><br><table class="table"><tr><th>Nom</th><th>Email</th><th>Fonction</th><th>Statut compte</th><th>Action</th></tr>{% for u in admins %}<tr><td>{{ u.display_name() }}</td><td>{{ u.email }}</td><td>{{ u.admin_role or '-' }}</td><td>{{ u.account_status }}</td><td><a class="btn secondary" href="{{ url_for('admin_send_password_reset', user_id=u.id) }}">Réinitialiser MDP</a> {% if u.id != current_user.id %}<a class="btn danger" href="{{ url_for('admin_remove_admin', user_id=u.id) }}">Retirer droits admin</a>{% else %}<span class="muted">Compte actuel</span>{% endif %}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucun admin.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'office')|safe }}
"""

TEMPLATE_IMPORT_MEMBERS = """
{% set content %}<div class="card form-wrap"><h1>Import Excel adhérents</h1><p class="muted">Le fichier peut contenir uniquement une colonne d'emails. Les colonnes Nom, Prénom, Statut et Type d'abonnement restent prises en compte si elles existent.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" enctype="multipart/form-data"><div class="field"><label>Fichier Excel .xlsx</label><input type="file" name="excel_file" accept=".xlsx,.xlsm" required></div><br><div class="form-grid"><div class="field"><label>Année d'abonnement</label><input name="subscription_year" type="number" value="{{ current_year }}" required></div><div class="field"><label>Abonnement par défaut</label><select name="subscription_type"><option>Annuel</option><option>Semestre 1</option><option>Semestre 2</option><option>Trimestre 1</option><option>Trimestre 2</option><option>Trimestre 3</option><option>Trimestre 4</option></select></div><div class="field"><label>Statut par défaut</label><select name="status"><option value="autre">Autre</option><option value="mensuel">Mensuel</option><option value="cadre">Cadre</option></select></div></div><br><button class="btn" type="submit">Importer et envoyer les liens d'activation</button> <a class="btn secondary" href="{{ url_for('admin_members') }}">Retour adhérents</a></form></div>{% endset %}{{ shell(content, 'members')|safe }}
"""
TEMPLATE_IMPORT_MEMBERS = TEMPLATE_IMPORT_MEMBERS.replace(
    """<option>Trimestre 2</option><option>Trimestre 3</option><option>Trimestre 4</option>""",
    """<option>T2</option><option>T3</option><option>T4</option>""",
    1,
)

TEMPLATE_COACH_PLANNING = """
{% set content %}<div class="card"><div class="top"><div><h1>Planning coachs</h1><p class="muted">Agenda mensuel par coach. Les cours se modifient dans Paramètres.</p></div><form method="get"><input name="year" type="number" value="{{ year }}" style="width:90px;padding:10px;border-radius:10px;border:1px solid #ddd"> <input name="month" type="number" min="1" max="12" value="{{ month }}" style="width:70px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_coach_absences', year=year, month=month) }}">Exporter</a></form></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<div style="overflow:auto"><table class="table"><tr><th style="min-width:120px">Date</th>{% for coach in coach_names %}<th style="min-width:190px">{{ coach }}</th>{% endfor %}</tr>{% for day in month_days %}<tr><td><strong>{{ weekday_labels[day.weekday()] }}</strong><br>{{ day.strftime('%d/%m') }}</td>{% for coach in coach_names %}<td>{% set slots = coach_agenda.get((coach, day), []) %}{% for s in slots %}{% set a = abs_by_key.get((coach, day)) %}<div style="border:1px solid #e5e7eb;border-left:4px solid #34a853;border-radius:10px;padding:8px;margin:6px 0;background:#fff"><strong>{{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</strong><br>{{ s.course_name }}<br>{% if not s.is_reservable %}<span class="badge wait">Sans résa</span>{% endif %}{% if a %}<span class="badge {% if a.status in ['absent','conge'] %}full{% elif a.status == 'replaced' %}wait{% endif %}">{{ a.status }}</span>{% if a.replacement_name %}<br><small>Remplaçant : {{ a.replacement_name }}</small>{% endif %}{% endif %}</div>{% else %}<span class="muted">-</span>{% endfor %}</td>{% endfor %}</tr>{% endfor %}</table></div><br><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Déclarer une absence / remplacement</h3><div class="form-grid"><div class="field"><label>Coach</label><select name="coach_name">{% for c in coaches %}<option>{{ c }}</option>{% endfor %}</select></div><div class="field"><label>Date</label><input name="absence_date" type="date" required></div><div class="field"><label>Statut</label><select name="status"><option value="absent">Absent</option><option value="conge">Congé</option><option value="present">Présent</option><option value="replaced">Remplacé</option></select></div><div class="field"><label>Remplaçant</label><select name="replacement_name"><option value="">-</option>{% for c in replacement_coaches %}<option>{{ c }}</option>{% endfor %}</select></div><div class="field" style="grid-column:1/-1"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Enregistrer</button></form><br><h2>Suivi des absences / congés</h2><table class="table"><tr><th>Date</th><th>Coach</th><th>Type</th><th>Remplaçant</th><th>Notes coach</th><th>Suivi admin</th></tr>{% for a in absences %}<tr><td>{{ weekday_labels[a.absence_date.weekday()] }} {{ a.absence_date.strftime('%d/%m/%Y') }}</td><td>{{ a.coach_name }}</td><td><span class="badge {% if a.status in ['absent','conge'] %}full{% elif a.status == 'replaced' %}wait{% endif %}">{{ a.status }}</span></td><td>{{ a.replacement_name or '-' }}</td><td>{{ a.notes or '' }}</td><td><form method="post" action="{{ url_for('update_coach_absence_followup', absence_id=a.id) }}"><input type="hidden" name="year" value="{{ year }}"><input type="hidden" name="month" value="{{ month }}"><div class="field"><select name="followup_status"><option value="a_traiter" {% if a.followup_status == 'a_traiter' %}selected{% endif %}>À traiter</option><option value="en_cours" {% if a.followup_status == 'en_cours' %}selected{% endif %}>En cours</option><option value="remplacement_a_trouver" {% if a.followup_status == 'remplacement_a_trouver' %}selected{% endif %}>Remplacement à trouver</option><option value="remplacement_trouve" {% if a.followup_status == 'remplacement_trouve' %}selected{% endif %}>Remplacement trouvé</option><option value="valide" {% if a.followup_status == 'valide' %}selected{% endif %}>Validé</option><option value="refuse" {% if a.followup_status == 'refuse' %}selected{% endif %}>Refusé</option></select></div><div class="field" style="margin-top:8px"><input name="admin_notes" value="{{ a.admin_notes or '' }}" placeholder="Note admin"></div><button class="btn secondary" type="submit" style="margin-top:8px">Enregistrer suivi</button>{% if a.reviewed_at %}<br><small class="muted">MAJ {{ a.reviewed_at.strftime('%d/%m/%Y %H:%M') }} par {{ a.reviewed_by or '-' }}</small>{% endif %}</form></td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune absence déclarée ce mois.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'coach_planning')|safe }}
"""

TEMPLATE_MEMBER_COACH_PLANNING = """
{% set content %}<div class="card"><div class="top"><div><h1>Planning coachs</h1><p class="muted">Agenda visuel des cours. Par défaut : 30 jours glissants à partir d'aujourd'hui. Période affichée : {{ range_label }}.</p></div><form method="get" class="card" style="box-shadow:none;background:#f9fafb;min-width:360px"><div class="form-grid"><div class="field"><label>Affichage</label><select name="view_mode"><option value="rolling" {% if view_mode == 'rolling' %}selected{% endif %}>Glissant 30 jours</option><option value="month" {% if view_mode == 'month' %}selected{% endif %}>Mois entier</option><option value="range" {% if view_mode == 'range' %}selected{% endif %}>Date à date</option></select></div><div class="field"><label>Année</label><input name="year" type="number" value="{{ year }}" min="2024" max="2100"></div><div class="field"><label>Mois</label><input name="month" type="number" min="1" max="12" value="{{ month }}"></div><div class="field"><label>Début</label><input name="start_date" type="date" value="{{ start.isoformat() }}"></div><div class="field"><label>Fin</label><input name="end_date" type="date" value="{{ end.isoformat() }}"></div></div><br><button class="btn secondary" type="submit">Afficher</button></form></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<div style="overflow:auto"><table class="table"><tr><th style="min-width:120px">Date</th>{% for coach in coach_names %}<th style="min-width:210px">{{ coach }}</th>{% endfor %}</tr>{% for day in month_days %}<tr><td><strong>{{ weekday_labels[day.weekday()] }}</strong><br>{{ day.strftime('%d/%m') }}</td>{% for coach in coach_names %}<td>{% set slots = coach_agenda.get((coach, day), []) %}{% for s in slots %}{% set a = abs_by_key.get((coach, day)) %}{% set booking = active_booking_by_session.get(s.id) %}<div style="border:1px solid #e5e7eb;border-left:4px solid #34a853;border-radius:10px;padding:8px;margin:6px 0;background:#fff"><strong>{{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }}</strong><br>{{ s.course_name }}<br>{% if a %}<span class="badge {% if a.status in ['absent','conge'] %}full{% elif a.status == 'replaced' %}wait{% endif %}">{{ a.status }}</span>{% if a.replacement_name %}<br><small>Remplaçant : {{ a.replacement_name }}</small>{% endif %}<br>{% endif %}{% if not s.is_reservable %}<span class="badge wait">Sans réservation</span>{% elif a and a.status in ['absent','conge'] %}<span class="badge full">Indisponible</span>{% elif booking %}{% if booking.status == 'waiting_list' %}<span class="badge wait">Liste d’attente — rang {{ waitlist_rank(booking) }}</span>{% else %}<span class="badge">Réservé</span>{% endif %}<br><br><a class="btn danger" href="{{ url_for('cancel', booking_id=booking.id, next=request.full_path) }}">Annuler</a>{% else %}<span class="badge">{{ s.capacity - booked_count(s) if booked_count(s) < s.capacity else 0 }} places</span><br><br><a class="btn" href="{{ url_for('book', session_id=s.id, next=request.full_path) }}">Réserver</a>{% endif %}</div>{% else %}<span class="muted">-</span>{% endfor %}</td>{% endfor %}</tr>{% endfor %}</table></div></div>{% endset %}{{ shell(content, 'member_coach_planning')|safe }}
"""

_ABSENCE_BADGE_SNIPPET = """<span class="badge {% if a.status in ['absent','conge'] %}full{% elif a.status == 'replaced' %}wait{% endif %}">{{ a.status }}</span>"""
_ABSENCE_BADGE_RENDER = """<span class="badge {{ absence_badge_class(a) }}">{{ absence_display_label(a) }}</span>"""
TEMPLATE_COACH_PROFILE = TEMPLATE_COACH_PROFILE.replace(_ABSENCE_BADGE_SNIPPET, _ABSENCE_BADGE_RENDER)
TEMPLATE_COACH_PROFILE = TEMPLATE_COACH_PROFILE.replace(
    "<tr><th>Date</th><th>Type</th><th>Remplaçant</th><th>Notes</th><th>Action</th></tr>",
    "<tr><th>Date</th><th>Créneau</th><th>Type</th><th>Remplaçant</th><th>Notes</th><th>Action</th></tr>",
)
TEMPLATE_COACH_PROFILE = TEMPLATE_COACH_PROFILE.replace(
    "<tr><td>{{ a.absence_date.strftime('%d/%m/%Y') }}</td><td>",
    "<tr><td>{{ a.absence_date.strftime('%d/%m/%Y') }}</td><td>{{ absence_session_label(a) }}</td><td>",
)
TEMPLATE_COACH_PROFILE = TEMPLATE_COACH_PROFILE.replace(
    "<tr><td colspan=\"5\" class=\"muted\">Aucune absence enregistrée.</td></tr>",
    "<tr><td colspan=\"6\" class=\"muted\">Aucune absence enregistrée.</td></tr>",
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(_ABSENCE_BADGE_SNIPPET, _ABSENCE_BADGE_RENDER)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    "{% set a = abs_by_key.get((coach, day)) %}",
    "{% set a = absence_for_session(abs_by_key, s) %}",
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """<p class="muted">Agenda mensuel par coach. Les cours se modifient dans Paramètres.</p></div><form method="get"><input name="year" type="number" value="{{ year }}" style="width:90px;padding:10px;border-radius:10px;border:1px solid #ddd"> <input name="month" type="number" min="1" max="12" value="{{ month }}" style="width:70px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_coach_absences', year=year, month=month) }}">Exporter</a></form></div>""",
    """<p class="muted">Par défaut : planning glissant à partir d'aujourd'hui. Période affichée : {{ range_label }}.</p></div><form method="get" class="card" style="box-shadow:none;background:#f9fafb;min-width:360px"><div class="form-grid"><div class="field"><label>Affichage</label><select name="view_mode"><option value="rolling" {% if view_mode == 'rolling' %}selected{% endif %}>Glissant 30 jours</option><option value="month" {% if view_mode == 'month' %}selected{% endif %}>Mois entier</option><option value="range" {% if view_mode == 'range' %}selected{% endif %}>Date à date</option></select></div><div class="field"><label>Année</label><input name="year" type="number" value="{{ year }}" min="2024" max="2100"></div><div class="field"><label>Mois</label><input name="month" type="number" min="1" max="12" value="{{ month }}"></div><div class="field"><label>Début</label><input name="start_date" type="date" value="{{ start.isoformat() }}"></div><div class="field"><label>Fin</label><input name="end_date" type="date" value="{{ end.isoformat() }}"></div></div><br><button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_coach_absences', view_mode=view_mode, start_date=start.isoformat(), end_date=end.isoformat(), year=year, month=month) }}">Exporter cette période</a></form></div>""",
    1,
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    "<tr><th>Date</th><th>Coach</th><th>Type</th><th>Remplaçant</th><th>Notes coach</th><th>Suivi admin</th></tr>",
    "<tr><th>Date</th><th>Créneau</th><th>Coach</th><th>Type</th><th>Remplaçant</th><th>Notes coach</th><th>Suivi admin</th></tr>",
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    "<tr><td>{{ weekday_labels[a.absence_date.weekday()] }} {{ a.absence_date.strftime('%d/%m/%Y') }}</td><td>{{ a.coach_name }}</td>",
    "<tr><td>{{ weekday_labels[a.absence_date.weekday()] }} {{ a.absence_date.strftime('%d/%m/%Y') }}</td><td>{{ absence_session_label(a) }}</td><td>{{ a.coach_name }}</td>",
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    "<tr><th>Date</th><th>Créneau</th><th>Coach</th><th>Type</th><th>Remplaçant</th><th>Notes coach</th><th>Suivi admin</th></tr>",
    "<tr><th>Date</th><th>Créneau</th><th>Coach</th><th>Type</th><th>Remplaçant</th><th>Notes coach</th><th>Suivi admin</th><th>Action</th></tr>",
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    "</form></td></tr>{% else %}<tr><td colspan=\"6\" class=\"muted\">Aucune absence déclarée ce mois.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'coach_planning')|safe }}",
    "</form></td><td><a class=\"btn danger\" href=\"{{ url_for('delete_coach_absence', absence_id=a.id, source='admin_planning', view_mode=view_mode, start_date=start.isoformat(), end_date=end.isoformat(), year=year, month=month) }}\" onclick=\"return confirm('Supprimer cette demande d\\'absence/congé ?')\">Supprimer</a></td></tr>{% else %}<tr><td colspan=\"7\" class=\"muted\">Aucune absence déclarée ce mois.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'coach_planning')|safe }}",
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace("colspan=\"7\" class=\"muted\">Aucune absence déclarée ce mois.", "colspan=\"8\" class=\"muted\">Aucune absence déclarée ce mois.")
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """<div class="field"><label>Date</label><input name="absence_date" type="date" required></div>""",
    """<div class="field"><label>Début</label><input name="start_date" type="date" required></div><div class="field"><label>Fin</label><input name="end_date" type="date" required></div>""",
    1,
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """<div class="field" style="margin-top:8px"><input name="admin_notes" value="{{ a.admin_notes or '' }}" placeholder="Note admin"></div><button class="btn secondary" type="submit" style="margin-top:8px">Enregistrer suivi</button>""",
    """<div class="field" style="margin-top:8px"><label>Remplaçant</label><select name="replacement_name"><option value="">-</option>{% for c in replacement_coaches %}<option value="{{ c }}" {% if a.replacement_name == c %}selected{% endif %}>{{ c }}</option>{% endfor %}</select></div><div class="field" style="margin-top:8px"><input name="admin_notes" value="{{ a.admin_notes or '' }}" placeholder="Note admin"></div><button class="btn secondary" type="submit" style="margin-top:8px">Enregistrer suivi</button>""",
    1,
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """<option value="refuse" {% if a.followup_status == 'refuse' %}selected{% endif %}>Refusé</option>""",
    """<option value="refuse" {% if a.followup_status == 'refuse' %}selected{% endif %}>Refusé</option><option value="annule" {% if a.followup_status == 'annule' %}selected{% endif %}>Cours annulé</option>""",
    1,
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """<input type="hidden" name="year" value="{{ year }}"><input type="hidden" name="month" value="{{ month }}"><div class="field"><select name="followup_status">""",
    """<input type="hidden" name="year" value="{{ year }}"><input type="hidden" name="month" value="{{ month }}"><input type="hidden" name="view_mode" value="{{ view_mode }}"><input type="hidden" name="start_date" value="{{ start.isoformat() }}"><input type="hidden" name="end_date" value="{{ end.isoformat() }}"><div class="field"><label>Créneau</label><select name="session_id"><option value="">Toute la journée</option>{% for s in absence_session_options(a) %}<option value="{{ s.id }}" {% if a.session_id == s.id %}selected{% endif %}>{{ s.start_time.strftime('%H:%M') }} - {{ s.end_time.strftime('%H:%M') }} · {{ s.course_name }}</option>{% endfor %}</select></div><div class="field" style="margin-top:8px"><label>Suivi</label><select name="followup_status">""",
    1,
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<div style="overflow:auto">""",
    """{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<div class="card" style="box-shadow:none;background:#f9fafb"><h2>Récapitulatif mensuel facturation</h2><table class="table"><tr><th>Coach</th><th>Cours effectués</th><th>Remplacements</th><th>Absences</th><th>Cours annulés</th></tr>{% for row in invoice_rows %}<tr><td>{{ row.coach }}</td><td><strong>{{ row.cours }}</strong></td><td>{{ row.remplacements }}</td><td>{{ row.absences }}</td><td>{{ row.annules }}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune donnée sur cette période.</td></tr>{% endfor %}</table></div><br><div style="overflow:auto">""",
    1,
)
TEMPLATE_COACH_PLANNING = TEMPLATE_COACH_PLANNING.replace(
    """</table></div><br><div style="overflow:auto">""",
    """</table><details style="margin-top:16px"><summary style="cursor:pointer;font-weight:800">Détail facturation par coach</summary><p class="muted">Règle de facturation : tout créneau d'1h ou moins est compté 1h30.</p><br><table class="table"><tr><th>Coach</th><th>Date</th><th>Horaire</th><th>Durée réelle</th><th>Durée facturée</th><th>Cours</th><th>Statut</th><th>Coach initial</th><th>Remplaçant</th><th>Suivi admin</th><th>Notes admin</th><th>Notes coach</th></tr>{% for row in invoice_detail_rows %}<tr><td>{{ row.coach }}</td><td>{{ row.jour }} {{ row.date.strftime('%d/%m/%Y') }}</td><td>{{ row.horaire }}</td><td>{{ row.duration_label }}</td><td><strong>{{ row.billed_label }}</strong></td><td>{{ row.cours }}</td><td>{{ row.statut }}</td><td>{{ row.coach_initial }}</td><td>{{ row.remplacant or '-' }}</td><td>{{ row.suivi_admin or '-' }}</td><td>{{ row.notes_admin or '' }}</td><td>{{ row.notes_coach or '' }}</td></tr>{% else %}<tr><td colspan="12" class="muted">Aucun détail sur cette période.</td></tr>{% endfor %}</table></details></div><br><div style="overflow:auto">""",
    1,
)
TEMPLATE_MEMBER_COACH_PLANNING = TEMPLATE_MEMBER_COACH_PLANNING.replace(_ABSENCE_BADGE_SNIPPET, _ABSENCE_BADGE_RENDER)
TEMPLATE_MEMBER_COACH_PLANNING = TEMPLATE_MEMBER_COACH_PLANNING.replace(
    "{% set a = abs_by_key.get((coach, day)) %}",
    "{% set a = absence_for_session(abs_by_key, s) %}",
)
TEMPLATE_MEMBER_COACH_PLANNING = TEMPLATE_MEMBER_COACH_PLANNING.replace("{% elif a and a.status in ['absent','conge'] %}", "{% elif a and absence_blocks_booking(a) %}")
TEMPLATE_COACH_SCHEDULE = TEMPLATE_COACH_SCHEDULE.replace("<th>Réservation</th></tr>", "<th>Réservation</th><th>Suivi</th></tr>")
TEMPLATE_COACH_SCHEDULE = TEMPLATE_COACH_SCHEDULE.replace(
    """<td>{% if s.is_reservable %}<span class="badge">Réservable</span>{% else %}<span class="badge wait">Sans résa</span>{% endif %}</td></tr>""",
    """<td>{% if s.is_reservable %}<span class="badge">Réservable</span>{% else %}<span class="badge wait">Sans résa</span>{% endif %}</td><td>{% set a = absence_for_session(abs_by_key, s) %}{% if a %}<span class="badge {{ absence_badge_class(a) }}">{{ absence_display_label(a) }}</span>{% if a.replacement_name %}<br><small>Remplaçant : {{ a.replacement_name }}</small>{% endif %}{% else %}<span class="muted">-</span>{% endif %}</td></tr>""",
)
TEMPLATE_COACH_SCHEDULE = TEMPLATE_COACH_SCHEDULE.replace('<tr><td colspan="4" class="muted">Aucun cours rattaché sur ce mois.</td></tr>', '<tr><td colspan="5" class="muted">Aucun cours rattaché sur ce mois.</td></tr>')
TEMPLATE_COACH_SCHEDULE = TEMPLATE_COACH_SCHEDULE.replace("<td>{{ item.absence.followup_status or '-' }}</td>", """<td><span class="badge {{ absence_badge_class(item.absence) }}">{{ absence_display_label(item.absence) }}</span></td>""")
TEMPLATE_COACH_SCHEDULE = TEMPLATE_COACH_SCHEDULE.replace(
    """<div class="grid"><div class="card"><span class="muted">Cours rattachés</span><div class="stat">{{ sessions|length }}</div></div><div class="card"><span class="muted">Remplacements</span><div class="stat">{{ replacements|length }}</div></div></div><h2>Cours prévus</h2>""",
    """<div class="grid"><div class="card"><span class="muted">Cours rattachés</span><div class="stat">{{ sessions|length }}</div></div><div class="card"><span class="muted">Remplacements</span><div class="stat">{{ replacements|length }}</div></div></div><br><div class="card" style="box-shadow:none;background:#f9fafb"><h2>Récapitulatif mensuel facturation</h2><table class="table"><tr><th>Coach</th><th>Cours effectués</th><th>Remplacements</th><th>Absences</th><th>Cours annulés</th></tr>{% for row in invoice_rows %}<tr><td>{{ row.coach }}</td><td><strong>{{ row.cours }}</strong></td><td>{{ row.remplacements }}</td><td>{{ row.absences }}</td><td>{{ row.annules }}</td></tr>{% else %}<tr><td colspan="5" class="muted">Aucune donnée sur ce mois.</td></tr>{% endfor %}</table></div><br><h2>Cours prévus</h2>""",
    1,
)
TEMPLATE_COACH_SCHEDULE = TEMPLATE_COACH_SCHEDULE.replace(
    """</table></div><br><h2>Cours prévus</h2>""",
    """</table><details style="margin-top:16px"><summary style="cursor:pointer;font-weight:800">Détail facturation</summary><p class="muted">Règle de facturation : tout créneau d'1h ou moins est compté 1h30.</p><br><table class="table"><tr><th>Date</th><th>Horaire</th><th>Durée réelle</th><th>Durée facturée</th><th>Cours</th><th>Statut</th><th>Coach initial</th><th>Remplaçant</th><th>Suivi admin</th><th>Notes admin</th></tr>{% for row in invoice_detail_rows %}<tr><td>{{ row.jour }} {{ row.date.strftime('%d/%m/%Y') }}</td><td>{{ row.horaire }}</td><td>{{ row.duration_label }}</td><td><strong>{{ row.billed_label }}</strong></td><td>{{ row.cours }}</td><td>{{ row.statut }}</td><td>{{ row.coach_initial }}</td><td>{{ row.remplacant or '-' }}</td><td>{{ row.suivi_admin or '-' }}</td><td>{{ row.notes_admin or '' }}</td></tr>{% else %}<tr><td colspan="10" class="muted">Aucun détail sur ce mois.</td></tr>{% endfor %}</table></details></div><br><h2>Cours prévus</h2>""",
    1,
)

TEMPLATE_SETTINGS = """
{% set content %}<div class="card"><h1>Paramètres</h1>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><input type="hidden" name="settings_section" value="pricing"><h3>Tarifs des abonnements par statut</h3><p class="muted">Renseigner le montant attendu pour chaque combinaison abonnement / statut. Ces montants alimentent automatiquement l'onglet Budget.</p><div class="field" style="max-width:360px"><label>Cotisation annuelle première inscription (€)</label><input name="annual_membership_fee" value="{{ '%.2f'|format(annual_membership_fee) }}"></div><br><table class="table"><tr><th>Abonnement</th>{% for profile_key, profile_label in member_profile_labels.items() %}<th>{{ profile_label }}</th>{% endfor %}</tr>{% for name in subscription_prices %}<tr><td><strong>{{ name }}</strong></td>{% for profile_key, profile_label in member_profile_labels.items() %}<td><input name="{{ subscription_profile_price_key(name, profile_key) }}" value="{{ '%.2f'|format(subscription_price_matrix[name][profile_key]) }}" style="width:110px;padding:10px;border:1px solid #d1d5db;border-radius:10px"> €</td>{% endfor %}</tr>{% endfor %}</table><br><button class="btn" type="submit">Enregistrer les tarifs</button></form><br><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Créer un cours</h3><div class="form-grid"><div class="field"><label>Nom du coach</label><input name="coach_name" required></div><div class="field"><label>Intitulé cours</label><input name="course_name" required></div><div class="field"><label>Jour</label><select name="weekday">{% for label in weekday_labels %}<option value="{{ loop.index0 }}">{{ label }}</option>{% endfor %}</select></div><div class="field"><label>Semaine odd/even</label><select name="week_parity"><option value="all">Toutes</option><option value="even">Even / paire</option><option value="odd">Odd / impaire</option></select></div><div class="field"><label>Début</label><input name="start_time" type="time" required></div><div class="field"><label>Fin</label><input name="end_time" type="time" required></div><div class="field"><label>Jauge</label><input name="capacity" type="number" value="35" min="1" required></div><div class="field"><label>Réservation</label><label style="font-weight:600"><input name="is_reservable" type="checkbox" checked style="width:auto"> Créneau à réserver</label></div></div><br><button class="btn" type="submit">Créer le cours</button></form><br><table class="table"><tr><th>Jour</th><th>Semaine</th><th>Cours</th><th>Horaire</th><th>Jauge</th><th>Coach</th><th>Réservation</th><th>Statut</th><th>Actions</th></tr>{% for t in templates %}<tr><form method="post" action="{{ url_for('edit_template', template_id=t.id) }}"><td><select name="weekday">{% for label in weekday_labels %}<option value="{{ loop.index0 }}" {% if t.weekday == loop.index0 %}selected{% endif %}>{{ label }}</option>{% endfor %}</select></td><td><select name="week_parity"><option value="all" {% if t.week_parity == 'all' %}selected{% endif %}>Toutes</option><option value="even" {% if t.week_parity == 'even' %}selected{% endif %}>Even</option><option value="odd" {% if t.week_parity == 'odd' %}selected{% endif %}>Odd</option></select></td><td><input name="course_name" value="{{ t.course_name }}" required></td><td><input name="start_time" type="time" value="{{ t.start_time.strftime('%H:%M') }}" required><br><input name="end_time" type="time" value="{{ t.end_time.strftime('%H:%M') }}" required></td><td><input name="capacity" type="number" min="1" value="{{ t.capacity }}" required style="width:80px"></td><td><input name="coach_name" value="{{ t.coach_name or '' }}" required></td><td><label style="font-weight:600"><input name="is_reservable" type="checkbox" {% if t.is_reservable %}checked{% endif %} style="width:auto"> Oui</label></td><td>{% if t.active %}<span class="badge">Actif</span>{% else %}<span class="badge full">Inactif</span>{% endif %}</td><td><button class="btn secondary" type="submit">Modifier</button> <a class="btn secondary" href="{{ url_for('toggle_template', template_id=t.id) }}">Activer / désactiver</a> <a class="btn danger" href="{{ url_for('delete_template', template_id=t.id) }}" onclick="return confirm('Supprimer ce cours ? Les séances futures sans réservation seront supprimées.')">Supprimer</a></td></form></tr>{% else %}<tr><td colspan="9" class="muted">Aucun cours.</td></tr>{% endfor %}</table><br><h3>Profs</h3><table class="table"><tr><th>Prof</th><th>Action</th></tr>{% for coach in coaches %}<tr><td>{{ coach }}</td><td><a class="btn danger" href="{{ url_for('delete_settings_coach', coach_name=coach) }}" onclick="return confirm('Supprimer ce prof des cours paramétrés et des futurs cours sans réservation ?')">Supprimer ce prof des cours</a></td></tr>{% else %}<tr><td colspan="2" class="muted">Aucun prof.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'settings')|safe }}
"""
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    "</table></div>{% endset %}{{ shell(content, 'settings')|safe }}",
    """</table><br><h3>Coachs remplaçants</h3><form method="post" action="{{ url_for('add_replacement_coach') }}" class="card" style="box-shadow:none;background:#f9fafb"><div class="form-grid"><div class="field"><label>Nom du remplaçant</label><input name="replacement_name" required></div></div><br><button class="btn" type="submit">Ajouter un remplaçant</button></form><br><table class="table"><tr><th>Remplaçant</th><th>Action</th></tr>{% for replacement in replacement_coaches %}<tr><td>{{ replacement }}</td><td><a class="btn danger" href="{{ url_for('delete_replacement_coach', coach_name=replacement) }}" onclick="return confirm('Supprimer ce remplaçant de la liste ?')">Supprimer</a></td></tr>{% else %}<tr><td colspan="2" class="muted">Aucun coach remplaçant.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'settings')|safe }}""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """</form><br><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Créer un cours</h3>""",
    """</form><br><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><input type="hidden" name="settings_section" value="coach_planning_display"><h3>Affichage du planning coach</h3><p class="muted">Choisir les jours visibles dans l'agenda visuel des coachs.</p><div class="form-grid">{% for label in weekday_labels %}<label style="font-weight:600"><input type="checkbox" name="planning_weekdays" value="{{ loop.index0 }}" {% if loop.index0 in planning_weekdays %}checked{% endif %} style="width:auto"> {{ label }}</label>{% endfor %}</div><br><button class="btn" type="submit">Enregistrer l'affichage</button></form><br><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Créer un cours</h3>""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """<div class="field"><label>Jauge</label><input name="capacity" type="number" value="35" min="1" required></div><div class="field"><label>Réservation</label>""",
    """<div class="field"><label>Jauge</label><input name="capacity" type="number" value="35" min="1" required></div><div class="field"><label>Liste d'attente</label><input name="waitlist_capacity" type="number" value="5" min="0" required></div><div class="field"><label>Réservation</label>""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """<div class="field"><label>Semaine odd/even</label><select name="week_parity"><option value="all">Toutes</option><option value="even">Even / paire</option><option value="odd">Odd / impaire</option></select></div><div class="field"><label>Début</label>""",
    """<div class="field"><label>Récurrence</label><select name="week_parity"><option value="all">Toutes les semaines</option><option value="even">Even / paire</option><option value="odd">Odd / impaire</option><option value="single">Ponctuel - une seule date</option></select></div><div class="field"><label>Date si ponctuel</label><input name="session_date" type="date"></div><div class="field"><label>Début</label>""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """<tr><th>Jour</th><th>Semaine</th><th>Cours</th><th>Horaire</th><th>Jauge</th><th>Coach</th><th>Réservation</th><th>Statut</th><th>Actions</th></tr>""",
    """<tr><th>Jour</th><th>Semaine</th><th>Cours</th><th>Horaire</th><th>Jauge</th><th>Liste attente</th><th>Coach</th><th>Réservation</th><th>Statut</th><th>Actions</th></tr>""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """<td><input name="capacity" type="number" min="1" value="{{ t.capacity }}" required style="width:80px"></td><td><input name="coach_name" value="{{ t.coach_name or '' }}" required></td>""",
    """<td><input name="capacity" type="number" min="1" value="{{ t.capacity }}" required style="width:80px"></td><td><input name="waitlist_capacity" type="number" min="0" value="{{ t.waitlist_capacity }}" required style="width:80px"></td><td><input name="coach_name" value="{{ t.coach_name or '' }}" required></td>""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace("""<tr><td colspan="9" class="muted">Aucun cours.</td></tr>""", """<tr><td colspan="10" class="muted">Aucun cours.</td></tr>""", 1)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """<h3>Profs</h3><table class="table"><tr><th>Prof</th><th>Action</th></tr>{% for coach in coaches %}<tr><td>{{ coach }}</td><td><a class="btn danger" href="{{ url_for('delete_settings_coach', coach_name=coach) }}" onclick="return confirm('Supprimer ce prof des cours paramétrés et des futurs cours sans réservation ?')">Supprimer ce prof des cours</a></td></tr>{% else %}<tr><td colspan="2" class="muted">Aucun prof.</td></tr>{% endfor %}</table>""",
    """<h3>Ajouter un coach</h3><form method="post" action="{{ url_for('add_settings_coach') }}" class="card" style="box-shadow:none;background:#f9fafb"><div class="form-grid"><div class="field"><label>Nom coach</label><input name="full_name" placeholder="Ex. Coach Fitness" required></div><div class="field"><label>Email coach</label><input name="email" type="email" required></div><div class="field"><label>Rôle</label><select name="coach_type"><option value="titulaire">Titulaire</option><option value="remplacant">Remplaçant</option></select></div></div><br><button class="btn" type="submit">Ajouter et envoyer le lien d'activation</button></form><br><h3>Profs</h3><table class="table"><tr><th>Prof</th><th>Email</th><th>Rôle</th><th>Action</th></tr>{% for coach in coaches %}<tr><form method="post" action="{{ url_for('update_settings_coach', coach_name=coach.name) }}"><td><input name="coach_name" value="{{ coach.name }}" required></td><td><input name="coach_email" type="email" value="{{ coach.email or '' }}" placeholder="email de connexion"></td><td><select name="coach_type"><option value="titulaire" {% if coach.coach_type == 'titulaire' %}selected{% endif %}>Titulaire</option><option value="remplacant" {% if coach.coach_type == 'remplacant' %}selected{% endif %}>Remplaçant</option></select></td><td><button class="btn secondary" type="submit">Modifier</button> {% if coach.user_id %}<a class="btn secondary" href="{{ url_for('admin_send_password_reset', user_id=coach.user_id) }}">Réinitialiser MDP</a> <a class="btn secondary" href="{{ url_for('admin_send_coach_activation', user_id=coach.user_id) }}">Renvoyer activation</a>{% endif %} <a class="btn danger" href="{{ url_for('delete_settings_coach', coach_name=coach.name) }}" onclick="return confirm('Supprimer ce prof des cours paramétrés et des futurs cours sans réservation ?')">Supprimer</a></td></form></tr>{% else %}<tr><td colspan="4" class="muted">Aucun prof.</td></tr>{% endfor %}</table>""",
    1,
)
TEMPLATE_SETTINGS = TEMPLATE_SETTINGS.replace(
    """<br><h3>Coachs remplaçants</h3><form method="post" action="{{ url_for('add_replacement_coach') }}" class="card" style="box-shadow:none;background:#f9fafb"><div class="form-grid"><div class="field"><label>Nom du remplaçant</label><input name="replacement_name" required></div></div><br><button class="btn" type="submit">Ajouter un remplaçant</button></form><br><table class="table"><tr><th>Remplaçant</th><th>Action</th></tr>{% for replacement in replacement_coaches %}<tr><td>{{ replacement }}</td><td><a class="btn danger" href="{{ url_for('delete_replacement_coach', coach_name=replacement) }}" onclick="return confirm('Supprimer ce remplaçant de la liste ?')">Supprimer</a></td></tr>{% else %}<tr><td colspan="2" class="muted">Aucun coach remplaçant.</td></tr>{% endfor %}</table>""",
    "",
    1,
)

TEMPLATE_BUDGET = """
{% set content %}<div class="card"><h1>Budget</h1><div class="grid"><div class="card"><span class="muted">Recettes</span><div class="stat">{{ '%.2f'|format(income) }} €</div><small class="muted">Inclut {{ '%.2f'|format(expected_dues) }} € de cotisations attendues {{ dues_year }}</small></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><div class="card" style="box-shadow:none;background:#f9fafb"><div class="top"><div><h2>Cotisations attendues</h2><p class="muted">La cotisation annuelle de première inscription est ajoutée aux adhérents créés dans l'année sélectionnée. Le tarif d'abonnement est repris depuis les tarifs paramétrés par statut.</p></div><form method="get"><input name="dues_year" type="number" value="{{ dues_year }}" style="width:100px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_budget_dues', dues_year=dues_year) }}">Exporter</a></form></div><table class="table"><tr><th>Adhérent</th><th>Email</th><th>Profil</th><th>Abonnement</th><th>Tarif indicatif</th><th>Tarif statut</th><th>Cotisation annuelle</th><th>Total</th></tr>{% for row in dues_rows %}<tr><td>{{ row.user.display_name() }}</td><td>{{ row.user.email }}</td><td>{{ row.member_profile_label }}</td><td>{{ row.user.subscription_type or '-' }}</td><td>{{ '%.2f'|format(row.base_subscription_price) }} €</td><td>{{ '%.2f'|format(row.subscription_price) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee) }} €{% else %}<span class="muted">-</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total) }} €</strong></td></tr>{% else %}<tr><td colspan="8" class="muted">Aucune cotisation attendue pour cette année.</td></tr>{% endfor %}</table></div><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date.strftime('%d/%m/%Y') }}</td><td>{{ e.entry_type }}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td>{{ '%.2f'|format(e.amount) }} €</td><td>{{ e.notes or '' }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne budget.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""
TEMPLATE_BUDGET = TEMPLATE_BUDGET.replace(
    """<td>{{ row.user.subscription_type or '-' }}</td>""",
    """<td>{{ row.subscription_type or '-' }} {{ row.subscription_year or '' }}</td>""",
    1,
)
TEMPLATE_BUDGET = """
{% set content %}<div class="card"><h1>Budget</h1><div class="grid"><div class="card"><span class="muted">Recettes</span><div class="stat">{{ '%.2f'|format(income) }} €</div><small class="muted">Inclut {{ '%.2f'|format(expected_dues) }} € de cotisations / abonnements attendus {{ dues_year }}</small></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><div class="card" style="box-shadow:none;background:#f9fafb"><div class="top"><div><h2>Budget détaillé</h2><p class="muted">Les cotisations / abonnements attendus et les lignes ajoutées manuellement sont regroupés dans ce tableau.</p></div><form method="get"><input name="dues_year" type="number" value="{{ dues_year }}" style="width:100px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_budget_dues', dues_year=dues_year) }}">Exporter</a></form></div><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Personne</th><th>Profil</th><th>Abonnement</th><th>Tarif abonnement</th><th>Cotisation annuelle</th><th>Montant</th><th>Notes</th></tr>{% for row in dues_rows %}<tr><td>{{ dues_year }}</td><td>Recette attendue</td><td>Cotisation / abonnement</td><td>Adhésion {{ row.subscription_type or '-' }} {{ row.subscription_year or '' }}</td><td>{{ row.user.display_name() }}<br><small class="muted">{{ row.user.email }}</small></td><td>{{ row.member_profile_label }}</td><td>{{ row.subscription_type or '-' }} {{ row.subscription_year or '' }}</td><td>{{ '%.2f'|format(row.subscription_price) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee) }} €{% else %}<span class="muted">-</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total) }} €</strong></td><td>{% if row.annual_fee %}Première cotisation annuelle {{ dues_year }}{% else %}Renouvellement / cotisation annuelle déjà comptée{% endif %}</td></tr>{% endfor %}{% for e in entries %}<tr><td>{{ e.entry_date.strftime('%d/%m/%Y') }}</td><td>{% if e.entry_type == 'income' %}Recette{% else %}Dépense{% endif %}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td><span class="muted">-</span></td><td><span class="muted">-</span></td><td><span class="muted">-</span></td><td><span class="muted">-</span></td><td><span class="muted">-</span></td><td><strong>{{ '%.2f'|format(e.amount) }} €</strong></td><td>{{ e.notes or '' }}</td></tr>{% endfor %}{% if not dues_rows and not entries %}<tr><td colspan="11" class="muted">Aucune ligne budget.</td></tr>{% endif %}</table></div></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_BUDGET_SAFE = """
{% set content %}<div class="card"><h1>Budget</h1><div class="flash">Mode sécurisé : une ancienne donnée ou une colonne manquante empêche l'affichage complet. Les détails techniques sont affichés dans les logs Render.</div><div class="grid"><div class="card"><span class="muted">Recettes</span><div class="stat">{{ '%.2f'|format(income) }} €</div><small class="muted">Cotisations attendues calculées : {{ '%.2f'|format(expected_dues) }} €</small></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date.strftime('%d/%m/%Y') }}</td><td>{{ e.entry_type }}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td>{{ '%.2f'|format(e.amount) }} €</td><td>{{ e.notes or '' }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne budget lisible pour le moment.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_BUDGET_SIMPLE = """
{% set content %}<div class="card"><h1>Budget</h1><p class="muted">Recettes attendues d'après les adhésions enregistrées, plus les lignes manuelles. La cotisation annuelle est comptée une seule fois par adhérent et par année.</p><div class="grid"><div class="card"><span class="muted">Recettes</span><div class="stat">{{ '%.2f'|format(income) }} €</div><small class="muted">{{ '%.2f'|format(expected_dues) }} € cotisations / abonnements + {{ '%.2f'|format(manual_income) }} € lignes manuelles</small></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><div class="card" style="box-shadow:none;background:#f9fafb"><div class="top"><div><h2>Cotisations et abonnements attendus</h2><p class="muted">Basé sur l'historique des inscriptions / renouvellements. Les renouvellements de la même année affichent l'abonnement, sans nouvelle cotisation annuelle.</p></div><form method="get"><input name="dues_year" type="number" value="{{ dues_year }}" style="width:100px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_budget_dues', dues_year=dues_year) }}">Exporter</a></form></div><table class="table"><tr><th>Adhérent</th><th>Profil</th><th>Abonnement</th><th>Tarif abonnement</th><th>Cotisation annuelle</th><th>Total</th><th>Note</th></tr>{% for row in dues_rows %}<tr><td>{{ row.user.display_name() }}<br><small class="muted">{{ row.user.email }}</small></td><td>{{ row.member_profile_label }}</td><td>{{ row.subscription_type or '-' }} {{ row.subscription_year or '' }}</td><td>{{ '%.2f'|format(row.subscription_price or 0) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee) }} €{% else %}<span class="muted">Déjà comptée / renouvellement</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total or 0) }} €</strong></td><td>{% if row.annual_fee %}Première inscription annuelle{% else %}Pas de nouvelle cotisation{% endif %}</td></tr>{% else %}<tr><td colspan="7" class="muted">Aucune cotisation / abonnement attendu pour cette année.</td></tr>{% endfor %}</table></div><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date.strftime('%d/%m/%Y') }}</td><td>{% if e.entry_type == 'income' %}Recette{% else %}Dépense{% endif %}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td><strong>{{ '%.2f'|format(e.amount) }} €</strong></td><td>{{ e.notes or '' }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne manuelle.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_BUDGET_SIMPLE = """
{% set content %}<div class="card"><h1>Budget</h1><p class="muted">Recettes attendues d'après les adhésions enregistrées, plus les lignes manuelles. La cotisation annuelle est comptée une seule fois par adhérent et par année.</p>{% if budget_warning %}<div class="flash">{{ budget_warning }}</div>{% endif %}<div class="grid"><div class="card"><span class="muted">Recettes</span><div class="stat">{{ '%.2f'|format(income) }} €</div><small class="muted">{{ '%.2f'|format(expected_dues) }} € cotisations / abonnements + {{ '%.2f'|format(manual_income) }} € lignes manuelles</small></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><div class="card" style="box-shadow:none;background:#f9fafb"><div class="top"><div><h2>Cotisations et abonnements attendus</h2><p class="muted">Les tarifs affichés sont les montants figés à la date d'inscription / renouvellement. Un renouvellement sur la même année ne reprend pas la cotisation de 10 €.</p></div><form method="get"><input name="dues_year" type="number" value="{{ dues_year }}" style="width:100px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_budget_dues', dues_year=dues_year) }}">Exporter</a></form></div><table class="table"><tr><th>Adhérent</th><th>Profil</th><th>Abonnement</th><th>Tarif abonnement</th><th>Cotisation annuelle</th><th>Total</th><th>Note</th></tr>{% for row in dues_rows %}<tr><td>{{ row.name }}<br><small class="muted">{{ row.email }}</small></td><td>{{ row.member_profile_label }}</td><td>{{ row.subscription_type }} {{ row.subscription_year }}</td><td>{{ '%.2f'|format(row.subscription_price or 0) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee) }} €{% else %}<span class="muted">Non</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total or 0) }} €</strong></td><td>{{ row.note }}</td></tr>{% else %}<tr><td colspan="7" class="muted">Aucune cotisation / abonnement attendu pour cette année.</td></tr>{% endfor %}</table></div><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date.strftime('%d/%m/%Y') }}</td><td>{% if e.entry_type == 'income' %}Recette{% else %}Dépense{% endif %}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td><strong>{{ '%.2f'|format(e.amount) }} €</strong></td><td>{{ e.notes or '' }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne manuelle.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_BUDGET_MINIMAL = """
{% set content %}<div class="card"><h1>Budget</h1><div class="flash">Budget ouvert en mode secours : les cotisations n'ont pas pu être affichées, mais les lignes manuelles restent accessibles.</div><div class="grid"><div class="card"><span class="muted">Recettes manuelles</span><div class="stat">{{ '%.2f'|format(income) }} €</div></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde manuel</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date.strftime('%d/%m/%Y') }}</td><td>{{ e.entry_type }}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td>{{ '%.2f'|format(e.amount) }} €</td><td>{{ e.notes or '' }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne manuelle.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_BUDGET_SIMPLE = """
{% set content %}<div class="card"><h1>Budget</h1><p class="muted">Recettes attendues d'après les adhésions enregistrées, plus les lignes manuelles. La cotisation annuelle est comptée une seule fois par adhérent et par année.</p>{% if budget_warning %}<div class="flash">{{ budget_warning }}</div>{% endif %}<div class="grid"><div class="card"><span class="muted">Recettes</span><div class="stat">{{ '%.2f'|format(income) }} €</div><small class="muted">{{ '%.2f'|format(expected_dues) }} € cotisations / abonnements + {{ '%.2f'|format(manual_income) }} € lignes manuelles</small></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><div class="card" style="box-shadow:none;background:#f9fafb"><div class="top"><div><h2>Cotisations et abonnements attendus</h2><p class="muted">Les tarifs affichés sont les montants figés à la date d'inscription / renouvellement. Un renouvellement sur la même année ne reprend pas la cotisation de 10 €.</p></div><form method="get"><input name="dues_year" type="number" value="{{ dues_year }}" style="width:100px;padding:10px;border-radius:10px;border:1px solid #ddd"> <button class="btn secondary" type="submit">Afficher</button> <a class="btn" href="{{ url_for('export_budget_dues', dues_year=dues_year) }}">Exporter</a></form></div><table class="table"><tr><th>Adhérent</th><th>Profil</th><th>Abonnement</th><th>Tarif abonnement</th><th>Cotisation annuelle</th><th>Total</th><th>Note</th></tr>{% for row in dues_rows %}<tr><td>{{ row.name }}<br><small class="muted">{{ row.email }}</small></td><td>{{ row.member_profile_label }}</td><td>{{ row.subscription_type }} {{ row.subscription_year }}</td><td>{{ '%.2f'|format(row.subscription_price or 0) }} €</td><td>{% if row.annual_fee %}{{ '%.2f'|format(row.annual_fee) }} €{% else %}<span class="muted">Non</span>{% endif %}</td><td><strong>{{ '%.2f'|format(row.total or 0) }} €</strong></td><td>{{ row.note }}</td></tr>{% else %}<tr><td colspan="7" class="muted">Aucune cotisation / abonnement attendu pour cette année.</td></tr>{% endfor %}</table></div><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date }}</td><td>{{ e.entry_type_label }}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td><strong>{{ '%.2f'|format(e.amount or 0) }} €</strong></td><td>{{ e.notes }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne manuelle.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_BUDGET_MINIMAL = """
{% set content %}<div class="card"><h1>Budget</h1><div class="flash">Budget ouvert en mode secours : les cotisations n'ont pas pu être affichées, mais les lignes manuelles restent accessibles.</div><div class="grid"><div class="card"><span class="muted">Recettes manuelles</span><div class="stat">{{ '%.2f'|format(income) }} €</div></div><div class="card"><span class="muted">Dépenses</span><div class="stat">{{ '%.2f'|format(expenses) }} €</div></div><div class="card"><span class="muted">Solde manuel</span><div class="stat">{{ '%.2f'|format(balance) }} €</div></div></div><form method="post" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter une ligne</h3><div class="form-grid"><div class="field"><label>Date</label><input name="entry_date" type="date" value="{{ today.isoformat() }}" required></div><div class="field"><label>Type</label><select name="entry_type"><option value="income">Recette</option><option value="expense">Dépense</option></select></div><div class="field"><label>Catégorie</label><select name="category"><option>Abonnement</option><option>Cotisation adhérent</option><option>Facture coach</option><option>Achat matériel</option><option>Autre</option></select></div><div class="field"><label>Libellé</label><input name="label" required></div><div class="field"><label>Montant (€)</label><input name="amount" required></div><div class="field"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><table class="table"><tr><th>Date</th><th>Type</th><th>Catégorie</th><th>Libellé</th><th>Montant</th><th>Notes</th></tr>{% for e in entries %}<tr><td>{{ e.entry_date }}</td><td>{{ e.entry_type_label }}</td><td>{{ e.category }}</td><td>{{ e.label }}</td><td>{{ '%.2f'|format(e.amount or 0) }} €</td><td>{{ e.notes }}</td></tr>{% else %}<tr><td colspan="6" class="muted">Aucune ligne manuelle.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'budget')|safe }}
"""

TEMPLATE_INVENTORY = """
{% set content %}<div class="card"><h1>Inventaire</h1><p class="muted">Valeur estimée : <strong>{{ '%.2f'|format(inventory_value) }} €</strong></p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="post" enctype="multipart/form-data" class="card" style="box-shadow:none;background:#f9fafb"><h3>Ajouter un article</h3><div class="form-grid"><div class="field"><label>Nom</label><input name="name" required></div><div class="field"><label>Catégorie</label><input name="category"></div><div class="field"><label>Quantité</label><input name="quantity" type="number" value="1" required></div><div class="field"><label>Seuil d'alerte</label><input name="alert_threshold" type="number" value="1" required></div><div class="field"><label>Coût unitaire</label><input name="unit_cost"></div><div class="field"><label>Année d'acquisition</label><input name="acquisition_year" type="number" min="1900" max="2100" value="{{ current_year }}"></div><div class="field"><label>Facture</label><input name="invoice_file" type="file" accept=".pdf,.png,.jpg,.jpeg,.doc,.docx,.xls,.xlsx"></div><div class="field"><label>Demande achat CSE</label><input name="purchase_request_file" type="file" accept=".pdf,.png,.jpg,.jpeg,.doc,.docx,.xls,.xlsx"></div><div class="field" style="grid-column:1/-1"><label>Notes</label><input name="notes"></div></div><br><button class="btn" type="submit">Ajouter</button></form><br><table class="table"><tr><th>Article</th><th>Catégorie</th><th>Quantité</th><th>Seuil</th><th>Année</th><th>Valeur</th><th>Documents</th><th>Notes</th></tr>{% for i in items %}<tr><td>{{ i.name }}</td><td>{{ i.category or '-' }}</td><td>{% if i.quantity <= i.alert_threshold %}<span class="badge full">{{ i.quantity }}</span>{% else %}<span class="badge">{{ i.quantity }}</span>{% endif %}</td><td>{{ i.alert_threshold }}</td><td>{{ i.acquisition_year or '-' }}</td><td>{{ '%.2f'|format((i.quantity or 0) * (i.unit_cost or 0)) }} €</td><td>{% if i.invoice_file %}<a class="btn secondary" href="{{ url_for('static', filename=i.invoice_file) }}" target="_blank">Facture</a>{% endif %} {% if i.purchase_request_file %}<a class="btn secondary" href="{{ url_for('static', filename=i.purchase_request_file) }}" target="_blank">Demande CSE</a>{% endif %}{% if not i.invoice_file and not i.purchase_request_file %}<span class="muted">-</span>{% endif %}</td><td>{{ i.notes or '' }}</td></tr>{% else %}<tr><td colspan="8" class="muted">Aucun article.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'inventory')|safe }}
"""

TEMPLATE_ARCHIVED_MEMBERS = """
{% set content %}<div class="card"><h1>Archives adhérents</h1><p class="muted">Dossiers adhérents archivés par année civile. Si une personne revient plus tard, ouvrez sa fiche puis renouvelez son adhésion.</p>{% with messages = get_flashed_messages() %}{% if messages %}{% for msg in messages %}<div class="flash">{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}<form method="get" class="card" style="box-shadow:none;background:#f9fafb"><h3>Rechercher dans les archives</h3><div class="form-grid"><div class="field"><label>Recherche</label><input name="search" value="{{ filter_values.search }}" placeholder="Nom, email, ID"></div><div class="field"><label>Année d'archivage</label><input name="archived_year" type="number" value="{{ filter_values.archived_year }}" placeholder="2026"></div><div class="field"><label>Année d'abonnement</label><input name="subscription_year" type="number" value="{{ filter_values.subscription_year }}" placeholder="2026"></div></div><br><button class="btn secondary" type="submit">Filtrer</button> <a class="btn secondary" href="{{ url_for('archived_members') }}">Réinitialiser</a></form><br><table class="table"><tr><th>Nom</th><th>Email</th><th>Abonnement</th><th>Fin abonnement</th><th>Archivé le</th><th>Motif</th><th>Action</th></tr>{% for u in users %}<tr><td>{{ u.display_name() }}</td><td>{{ u.email }}</td><td>{{ u.subscription_type }} {{ u.subscription_year }}</td><td>{{ u.subscription_end_date or '-' }}</td><td>{{ u.archived_at or '-' }}</td><td>{{ u.archived_reason or '-' }}</td><td><a class="btn secondary" href="{{ url_for('admin_edit_member', user_id=u.id) }}">Ouvrir / renouveler</a></td></tr>{% else %}<tr><td colspan="7" class="muted">Aucun ancien adhérent archivé.</td></tr>{% endfor %}</table></div>{% endset %}{{ shell(content, 'archives')|safe }}
"""
# -------------------- Gestion avancée admin : bureau, import, budget, inventaire, coachs --------------------

WEEKDAY_LABELS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
DEFAULT_COURSE_TEMPLATES = [
    (0, "even", "Biking", time(11, 45), time(12, 45), 16, "Hayate"),
    (0, "even", "Pilates", time(12, 45), time(13, 45), 35, "Hayate"),
    (0, "odd", "Body Barres", time(11, 45), time(12, 45), 18, "Hayate"),
    (0, "odd", "Body Zen", time(12, 45), time(13, 45), 35, "Hayate"),
    (1, "all", "Pilates", time(18, 0), time(18, 45), 35, "Hayate"),
    (1, "all", "Biking", time(18, 45), time(19, 30), 16, "Hayate"),
]

REMOVED_DEFAULT_COURSE_TEMPLATES = [
    (1, "all", "Step", time(17, 30), time(18, 0), "Hayate"),
    (2, "all", "Body Sculpt", time(12, 45), time(13, 45), "Hayate"),
]


def seed_default_course_templates():
    cleanup_removed_default_course_templates()
    for weekday, parity, name, start, end, capacity, coach in DEFAULT_COURSE_TEMPLATES:
        existing = CourseTemplate.query.filter_by(weekday=weekday, week_parity=parity, course_name=name, start_time=start, end_time=end).first()
        if not existing:
            db.session.add(CourseTemplate(weekday=weekday, week_parity=parity, course_name=name, start_time=start, end_time=end, capacity=capacity, coach_name=coach, active=True, is_reservable=True))
    db.session.commit()


def cleanup_removed_default_course_templates():
    today = date.today()
    removed_evening_names = {("Cours du " + "soir " + str(index)) for index in range(1, 4)}
    for template in CourseTemplate.query.filter(CourseTemplate.course_name.in_(removed_evening_names)).all():
        db.session.delete(template)
    for session in CourseSession.query.filter(
        CourseSession.course_date >= today,
        CourseSession.course_name.in_(removed_evening_names),
    ).all():
        if not session.bookings:
            db.session.delete(session)

    for weekday, parity, name, start, end, coach in REMOVED_DEFAULT_COURSE_TEMPLATES:
        templates = CourseTemplate.query.filter_by(
            weekday=weekday,
            week_parity=parity,
            course_name=name,
            start_time=start,
            end_time=end,
            coach_name=coach,
        ).all()
        for template in templates:
            db.session.delete(template)

        sessions = CourseSession.query.filter(
            CourseSession.course_date >= today,
            CourseSession.course_name == name,
            CourseSession.start_time == start,
            CourseSession.end_time == end,
            CourseSession.coach_name == coach,
        ).all()
        for session in sessions:
            iso_week = session.course_date.isocalendar().week
            session_parity = "even" if iso_week % 2 == 0 else "odd"
            parity_matches = parity == "all" or parity == session_parity
            if session.course_date.weekday() == weekday and parity_matches and not session.bookings:
                db.session.delete(session)
    db.session.commit()


@app.route("/activate/<token>", methods=["GET", "POST"])
def activate_account(token):
    user = User.query.filter_by(activation_token=token).first()
    if not user or not user.activation_expires_at or user.activation_expires_at < datetime.utcnow():
        flash("Lien d'activation invalide ou expiré.")
        return redirect(url_for("login"))
    if request.method == "POST":
        password = request.form["password"]
        if len(password) < 8:
            flash("Merci de choisir un mot de passe d'au moins 8 caractères.")
            return render_template_string(TEMPLATE_ACTIVATE, user=user)
        if user.role == "adherent":
            first_name, last_name, full_name = form_full_name()
            if not first_name or not last_name:
                flash("Merci d'indiquer votre prénom et votre nom.")
                return render_template_string(TEMPLATE_ACTIVATE, user=user)
            member_profile = request.form.get("member_profile", user.member_profile or "ouvrant_droit")
            rights_holder_name = request.form.get("rights_holder_name", "").strip()
            status = normalize_member_status(member_profile, request.form.get("status", user.status or "autre"))
            if member_profile not in MEMBER_PROFILE_RATES:
                member_profile = "ouvrant_droit"
            if member_profile == "ayant_droit" and not rights_holder_name:
                flash("Merci d'indiquer le nom et prénom de l'ouvrant droit.")
                return render_template_string(TEMPLATE_ACTIVATE, user=user)
            photo = request.files.get("profile_photo")
            if not ((photo and photo.filename) or user.profile_photo or user.profile_photo_data):
                flash("Merci d'ajouter une photo de profil.")
                return render_template_string(TEMPLATE_ACTIVATE, user=user)
            user.first_name = first_name
            user.last_name = last_name
            user.full_name = full_name
            user.member_profile = member_profile
            user.status = status
            user.rights_holder_name = rights_holder_name if member_profile == "ayant_droit" else None
            if not user.subscription_end_date or user.subscription_end_date < date.today():
                user.subscription_year = date.today().year
                user.subscription_end_date = subscription_end(user.subscription_type or "Annuel", user.subscription_year)
            create_membership_period(user, user.subscription_type or "Annuel", user.subscription_year or date.today().year, annual_fee_applies=not MembershipPeriod.query.filter_by(user_id=user.id, subscription_year=user.subscription_year or date.today().year).first(), created_by=user.display_name(), notes="Activation compte")
        user.set_password(password)
        user.account_status = "active"
        user.activation_token = None
        user.activation_expires_at = None
        photo = request.files.get("profile_photo")
        if photo and photo.filename:
            try:
                persist_profile_photo(user, photo)
            except ValueError as exc:
                flash(str(exc))
                return render_template_string(TEMPLATE_ACTIVATE, user=user)
        if user.role == "adherent" and not user.member_card:
            generate_member_card(user)
        db.session.commit()
        flash("Compte activé. Vous pouvez vous connecter.")
        return redirect(url_for("login"))
    return render_template_string(TEMPLATE_ACTIVATE, user=user)


@app.route("/admin/office", methods=["GET", "POST"])
@login_required
def admin_office():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        full_name = request.form.get("full_name", "").strip() or email
        admin_role = request.form.get("admin_role", "membre_bureau")
        user = User.query.filter_by(email=email).first()
        if not user:
            user = User(email=email, role="admin", status="autre", full_name=full_name, admin_role=admin_role, account_status="pending")
            user.set_password(secrets.token_urlsafe(12))
            db.session.add(user)
            db.session.commit()
            activation_link = create_activation_link(user)
            sent = send_email(user.email, "Activation de votre compte Section Fitness", f"Bonjour {user.display_name()},\n\nVotre compte admin Section Fitness a été pré-enregistré. Merci de créer votre mot de passe avec ce lien :\n{activation_link}\n\nCe lien est valable 14 jours.\n\nSection Fitness")
            if sent:
                flash("Admin ajouté. Un lien d'activation a été envoyé.")
            else:
                flash(f"Admin ajouté. Email non envoyé : copiez ce lien d'activation et envoyez-le manuellement : {activation_link}")
        else:
            user.role = "admin"
            user.full_name = full_name
            user.admin_role = admin_role
            user.account_status = "active"
            db.session.commit()
            flash("Droits admin mis à jour.")
        return redirect(url_for("admin_office"))
    admins = User.query.filter_by(role="admin").order_by(User.admin_role, User.full_name, User.email).all()
    return render_template_string(TEMPLATE_OFFICE, admins=admins)


@app.route("/admin/office/remove/<int:user_id>")
@login_required
def admin_remove_admin(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("Vous ne pouvez pas retirer vos propres droits admin depuis cet écran.")
    elif user.role == "admin":
        user.role = "adherent"
        user.admin_role = None
        db.session.commit()
        flash("Droits admin retirés.")
    return redirect(url_for("admin_office"))


@app.route("/admin/members/import", methods=["GET", "POST"])
@login_required
def admin_import_members():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    if request.method == "POST":
        file = request.files.get("excel_file")
        year = int(request.form.get("subscription_year") or date.today().year)
        default_subscription_type = normalize_subscription_type(request.form.get("subscription_type", "Annuel"))
        if default_subscription_type not in SUBSCRIPTION_PRICES:
            default_subscription_type = "Annuel"
        default_status = request.form.get("status", "autre")
        if default_status not in ["cadre", "mensuel", "autre"]:
            default_status = "autre"
        if not file or not file.filename.endswith((".xlsx", ".xlsm")):
            flash("Merci d'importer un fichier Excel .xlsx.")
            return redirect(url_for("admin_import_members"))
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        def idx(*names):
            for n in names:
                if n in headers:
                    return headers.index(n)
            return None
        i_nom = idx("nom", "name")
        i_prenom = idx("prénom", "prenom", "first name")
        i_email = idx("adresse email", "email", "mail")
        i_statut = idx("statut", "status")
        i_abonnement = idx("type d'abonnement", "abonnement", "subscription")
        if i_email is None:
            for col_index, cell in enumerate(ws[1]):
                if valid_email(str(cell.value or "")):
                    i_email = col_index
                    break
        if i_email is None:
            flash("Le fichier doit contenir au minimum une colonne email/adresse email, ou une première ligne contenant une adresse email.")
            return redirect(url_for("admin_import_members"))
        created = updated = emailed = 0
        start_row = 2 if headers and any(h in headers for h in ["adresse email", "email", "mail"]) else 1
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            email = valid_email(str(row[i_email]) if len(row) > i_email and row[i_email] else "")
            if not email:
                continue
            nom = str(row[i_nom]).strip() if i_nom is not None and len(row) > i_nom and row[i_nom] else ""
            prenom = str(row[i_prenom]).strip() if i_prenom is not None and len(row) > i_prenom and row[i_prenom] else ""
            full_name = (prenom + " " + nom).strip() or email
            status = str(row[i_statut]).strip().lower() if i_statut is not None and len(row) > i_statut and row[i_statut] else default_status
            if status not in ["cadre", "mensuel", "autre"]:
                status = default_status
            subscription_type = normalize_subscription_type(str(row[i_abonnement]).strip()) if i_abonnement is not None and len(row) > i_abonnement and row[i_abonnement] else default_subscription_type
            if subscription_type not in SUBSCRIPTION_PRICES:
                subscription_type = default_subscription_type
            user = User.query.filter_by(email=email).first()
            if not user:
                user = User(email=email, role="adherent", status=status, full_name=full_name, subscription_type=subscription_type, subscription_year=year, subscription_end_date=subscription_end(subscription_type, year), account_status="pending", member_number=next_member_number(year))
                user.set_password(secrets.token_urlsafe(12))
                db.session.add(user)
                db.session.flush()
                created += 1
            else:
                user.role = "adherent"
                user.status = status
                user.full_name = full_name
                user.subscription_type = subscription_type
                user.subscription_year = year
                user.subscription_end_date = subscription_end(subscription_type, year)
                if user.account_status == "archived":
                    user.archived_at = None
                    user.archived_reason = None
                user.account_status = "pending"
                updated += 1
            db.session.commit()
            if send_activation_email(user):
                emailed += 1
        flash(f"Import terminé : {created} créé(s), {updated} mis à jour, {emailed} email(s) réellement envoyé(s).")
        return redirect(url_for("admin_members"))
    return render_template_string(TEMPLATE_IMPORT_MEMBERS, current_year=date.today().year)


@app.route("/admin/members/send-activation/<int:user_id>")
@login_required
def admin_send_activation(user_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    user = User.query.get_or_404(user_id)
    send_activation_email(user)
    flash("Lien d'activation envoyé ou affiché dans la console si SMTP non configuré.")
    return redirect(url_for("admin_members"))


@app.route("/admin/coach-planning", methods=["GET", "POST"])
@login_required
def admin_coach_planning():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    ensure_coach_absence_schema()
    view_mode, start, end, year, month = coach_planning_period(request.values)
    if request.method == "POST":
        coach_name = request.form["coach_name"]
        start_date = datetime.strptime(request.form.get("start_date") or request.form["absence_date"], "%Y-%m-%d").date()
        end_date = datetime.strptime(request.form.get("end_date") or request.form.get("start_date") or request.form["absence_date"], "%Y-%m-%d").date()
        if end_date < start_date:
            flash("La date de fin doit être postérieure ou égale à la date de début.")
            return redirect(url_for("admin_coach_planning", view_mode=view_mode, start_date=start.isoformat(), end_date=end.isoformat(), year=year, month=month))
        status = request.form.get("status", "absent")
        replacement = request.form.get("replacement_name", "").strip()
        notes = request.form.get("notes", "").strip()
        current_day = start_date
        saved = 0
        while current_day <= end_date:
            target_sessions = absence_target_sessions(coach_name, current_day)
            if target_sessions:
                for session in target_sessions:
                    upsert_coach_absence(coach_name, current_day, status, replacement, notes, session=session)
                    saved += 1
            else:
                pass
            current_day += timedelta(days=1)
        db.session.commit()
        if saved == 0:
            flash("Aucune absence créée : aucun cours n'existe pour cette coach sur la période sélectionnée.")
            return redirect(url_for("admin_coach_planning", view_mode=view_mode, start_date=start.isoformat(), end_date=end.isoformat(), year=year, month=month))
        member_sent = notify_members_of_coach_absence(coach_name, start_date, end_date, status, replacement, notes)
        flash(f"Planning coach mis à jour sur {saved} jour(s). Email envoyé à {member_sent} adhérent(s) inscrit(s)." if member_sent else f"Planning coach mis à jour sur {saved} jour(s).")
        return redirect(url_for("admin_coach_planning", view_mode=view_mode, start_date=start.isoformat(), end_date=end.isoformat(), year=year, month=month))
    sessions = CourseSession.query.filter(CourseSession.course_date >= start, CourseSession.course_date <= end).order_by(CourseSession.course_date, CourseSession.start_time).all()
    absences = CoachAbsence.query.filter(CoachAbsence.absence_date >= start, CoachAbsence.absence_date <= end).order_by(CoachAbsence.absence_date, CoachAbsence.coach_name, CoachAbsence.session_id).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in absences}
    coaches = titular_coach_names()
    effective_names = {effective_coach_for_session(s, abs_by_key) for s in sessions if effective_coach_for_session(s, abs_by_key) != "-"}
    coach_names = sorted(set(coaches) | {name for name in effective_names if coach_type_for_name(name) == "titulaire"} | {name for name in effective_names if name in get_replacement_coaches()})
    planning_weekdays = set(get_coach_planning_weekdays())
    month_days = [start + timedelta(days=i) for i in range((end - start).days + 1) if (start + timedelta(days=i)).weekday() in planning_weekdays]
    coach_agenda = {}
    for session in sessions:
        coach_agenda.setdefault((effective_coach_for_session(session, abs_by_key), session.course_date), []).append(session)
    range_label = f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"
    invoice_rows = coach_monthly_invoice_rows(start, end)
    invoice_detail_rows = coach_invoice_detail_rows(start, end)
    invoice_summary_rows = coach_invoice_summary_rows(start, end)
    return render_template_string(TEMPLATE_COACH_PLANNING, sessions=sessions, absences=absences, abs_by_key=abs_by_key, coaches=coaches, replacement_coaches=coach_replacement_options(), coach_names=coach_names, month_days=month_days, coach_agenda=coach_agenda, invoice_rows=invoice_rows, invoice_detail_rows=invoice_detail_rows, invoice_summary_rows=invoice_summary_rows, year=year, month=month, view_mode=view_mode, start=start, end=end, range_label=range_label, weekday_labels=WEEKDAY_LABELS)


@app.route("/admin/coach-absence/<int:absence_id>/followup", methods=["POST"])
@login_required
def update_coach_absence_followup(absence_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    ensure_coach_absence_schema()
    absence = CoachAbsence.query.get_or_404(absence_id)
    allowed = {"a_traiter", "en_cours", "remplacement_a_trouver", "remplacement_trouve", "valide", "refuse", "annule"}
    status = request.form.get("followup_status", "a_traiter")
    absence.followup_status = status if status in allowed else "a_traiter"
    if "session_id" in request.form:
        raw_session_id = request.form.get("session_id", "").strip()
        if raw_session_id:
            session = CourseSession.query.get(int(raw_session_id))
            if session and session.coach_name == absence.coach_name and session.course_date == absence.absence_date:
                duplicate = CoachAbsence.query.filter_by(
                    coach_name=absence.coach_name,
                    absence_date=absence.absence_date,
                    session_id=session.id,
                ).filter(CoachAbsence.id != absence.id).first()
                if duplicate:
                    duplicate.status = absence.status
                    duplicate.replacement_name = absence.replacement_name
                    duplicate.notes = absence.notes
                    duplicate.followup_status = absence.followup_status
                    duplicate.admin_notes = absence.admin_notes
                    duplicate.reviewed_at = absence.reviewed_at
                    duplicate.reviewed_by = absence.reviewed_by
                    db.session.delete(absence)
                    absence = duplicate
                else:
                    absence.session_id = session.id
        else:
            absence.session_id = None
    replacement_name = request.form.get("replacement_name")
    if replacement_name is not None:
        absence.replacement_name = replacement_name.strip()
        if absence.replacement_name and absence.followup_status in ["a_traiter", "remplacement_a_trouver", "valide"]:
            absence.followup_status = "remplacement_trouve"
    notify_replacement = False
    notify_absent = False
    notify_search = False
    if absence.followup_status == "valide":
        notify_absent = True
        if absence.replacement_name:
            absence.followup_status = "remplacement_trouve"
            notify_replacement = True
        else:
            absence.followup_status = "remplacement_a_trouver"
            notify_search = True
    elif absence.followup_status == "remplacement_trouve" and absence.replacement_name:
        notify_absent = True
        notify_replacement = True
    elif absence.followup_status == "remplacement_a_trouver":
        notify_absent = True
        notify_search = True
    absence.admin_notes = request.form.get("admin_notes", "").strip()
    absence.reviewed_at = datetime.utcnow()
    absence.reviewed_by = current_user.display_name()
    db.session.commit()
    sent_parts = []
    if notify_absent and notify_coach_absence_validated(absence):
        sent_parts.append("coach absent notifié")
    if notify_replacement and notify_replacement_assigned(absence):
        sent_parts.append("coach remplaçant notifié")
    if notify_search:
        sent = notify_coaches_replacement_needed(absence)
        if sent:
            sent_parts.append(f"{sent} coach(s) contacté(s) pour remplacement")
    flash("Suivi de la demande mis à jour." + (f" Emails : {', '.join(sent_parts)}." if sent_parts else ""))
    return redirect(url_for("admin_coach_planning", view_mode=request.form.get("view_mode", "rolling"), start_date=request.form.get("start_date", ""), end_date=request.form.get("end_date", ""), year=request.form.get("year", absence.absence_date.year), month=request.form.get("month", absence.absence_date.month)))


@app.route("/admin/coach-planning/export")
@login_required
def export_coach_absences():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    ensure_coach_absence_schema()
    view_mode, start, end, year, month = coach_planning_period(request.args)
    absences = CoachAbsence.query.filter(CoachAbsence.absence_date >= start, CoachAbsence.absence_date <= end).order_by(CoachAbsence.absence_date, CoachAbsence.coach_name, CoachAbsence.session_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Absences coachs"
    ws.append(["Date", "Créneau", "Coach", "Type", "Remplaçant", "Notes coach", "Suivi admin", "Notes admin", "Mis à jour le", "Mis à jour par", "Créé le"])
    for a in absences:
        ws.append([
            a.absence_date.strftime("%d/%m/%Y"),
            absence_session_label(a),
            a.coach_name,
            a.status,
            a.replacement_name or "",
            a.notes or "",
            a.followup_status or "a_traiter",
            a.admin_notes or "",
            a.reviewed_at.strftime("%d/%m/%Y %H:%M") if a.reviewed_at else "",
            a.reviewed_by or "",
            a.created_at.strftime("%d/%m/%Y %H:%M") if a.created_at else "",
        ])
    sessions = CourseSession.query.filter(CourseSession.course_date >= start, CourseSession.course_date <= end).order_by(CourseSession.course_date, CourseSession.coach_name, CourseSession.start_time).all()
    abs_by_key = {(a.coach_name, a.absence_date, a.session_id): a for a in absences}
    ws2 = wb.create_sheet("Planning cours")
    ws2.append(["Date", "Jour", "Coach", "Cours", "Horaire", "Réservation", "Statut planning", "Remplaçant"])
    for session in sessions:
        absence = absence_for_session(abs_by_key, session)
        ws2.append([
            session.course_date.strftime("%d/%m/%Y"),
            WEEKDAY_LABELS[session.course_date.weekday()],
            session.coach_name or "",
            session.course_name,
            f"{session.start_time.strftime('%H:%M')} - {session.end_time.strftime('%H:%M')}",
            "Oui" if session.is_reservable else "Non",
            absence_display_label(absence) if absence else "",
            absence.replacement_name if absence and absence.replacement_name else "",
        ])
    ws3 = wb.create_sheet("Récap facturation")
    ws3.append(["Coach", "Cours effectués", "Remplacements", "Absences", "Cours annulés"])
    for row in coach_monthly_invoice_rows(start, end):
        ws3.append([row["coach"], row["cours"], row["remplacements"], row["absences"], row["annules"]])
    ws4 = wb.create_sheet("Détail facturation")
    ws4.append(["Règle", "Tout créneau d'1h ou moins est compté 1h30 pour la facturation."])
    ws4.append([])
    ws4.append(["Coach", "Date", "Jour", "Horaire", "Durée réelle", "Durée facturée", "Cours", "Statut", "Coach initial", "Remplaçant", "Suivi admin", "Notes admin", "Notes coach"])
    for row in coach_invoice_detail_rows(start, end):
        ws4.append([
            row["coach"],
            row["date"].strftime("%d/%m/%Y"),
            row["jour"],
            row["horaire"],
            row["duration_label"],
            row["billed_label"],
            row["cours"],
            row["statut"],
            row["coach_initial"],
            row["remplacant"],
            row["suivi_admin"],
            row["notes_admin"],
            row["notes_coach"],
        ])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name=f"suivi_absences_coachs_{start.isoformat()}_{end.isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/settings", methods=["GET", "POST"])
@login_required
def admin_settings():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    if request.method == "POST":
        if request.form.get("settings_section") == "coach_planning_display":
            selected_days = request.form.getlist("planning_weekdays")
            if not selected_days:
                flash("Sélectionnez au moins un jour à afficher dans le planning coach.")
                return redirect(url_for("admin_settings"))
            set_coach_planning_weekdays(selected_days)
            db.session.commit()
            flash("Affichage du planning coach mis à jour.")
            return redirect(url_for("admin_settings"))
        if request.form.get("settings_section") == "pricing":
            set_setting_value("annual_membership_fee", parse_amount(request.form.get("annual_membership_fee"), DEFAULT_ANNUAL_MEMBERSHIP_FEE))
            for name in SUBSCRIPTION_PRICES:
                if request.form.get(subscription_price_key(name)) is not None:
                    set_setting_value(subscription_price_key(name), parse_amount(request.form.get(subscription_price_key(name)), SUBSCRIPTION_PRICES[name]))
                for member_profile in MEMBER_PROFILE_LABELS:
                    key = subscription_profile_price_key(name, member_profile)
                    default_price = SUBSCRIPTION_PRICES[name] * member_profile_rate(member_profile)
                    set_setting_value(key, parse_amount(request.form.get(key), default_price))
            db.session.commit()
            flash("Tarifs des abonnements mis à jour.")
            return redirect(url_for("admin_settings"))
        parity = request.form.get("week_parity", "all")
        if parity == "single":
            if not request.form.get("session_date"):
                flash("Merci d'indiquer une date pour une session ponctuelle.")
                return redirect(url_for("admin_settings"))
            single_date = datetime.strptime(request.form["session_date"], "%Y-%m-%d").date()
            course_name = request.form["course_name"].strip()
            start = datetime.strptime(request.form["start_time"], "%H:%M").time()
            end = datetime.strptime(request.form["end_time"], "%H:%M").time()
            if end <= start:
                flash("L'heure de fin doit être après l'heure de début.")
                return redirect(url_for("admin_settings"))
            capacity = int(request.form.get("capacity", 35))
            waitlist_capacity_value = int(request.form.get("waitlist_capacity", 5))
            coach_name = request.form.get("coach_name", "").strip()
            is_reservable = request.form.get("is_reservable") == "on"
            created = create_session_if_missing(single_date, course_name, start, end, capacity, date.today(), coach_name, is_reservable, waitlist_capacity_value)
            db.session.commit()
            flash("Session ponctuelle créée." if created else "Cette session ponctuelle existe déjà.")
            return redirect(url_for("admin_settings"))
        weekday = int(request.form["weekday"])
        course_name = request.form["course_name"].strip()
        start = datetime.strptime(request.form["start_time"], "%H:%M").time()
        end = datetime.strptime(request.form["end_time"], "%H:%M").time()
        if end <= start:
            flash("L'heure de fin doit être après l'heure de début.")
            return redirect(url_for("admin_settings"))
        capacity = int(request.form.get("capacity", 35))
        waitlist_capacity_value = int(request.form.get("waitlist_capacity", 5))
        coach_name = request.form.get("coach_name", "").strip()
        is_reservable = request.form.get("is_reservable") == "on"
        db.session.add(CourseTemplate(weekday=weekday, week_parity=parity, course_name=course_name, start_time=start, end_time=end, capacity=capacity, waitlist_capacity=waitlist_capacity_value, coach_name=coach_name, active=True, is_reservable=is_reservable))
        db.session.commit()
        generate_rolling_sessions(days_ahead=28)
        flash("Cours créé. Il apparaît dans le planning coach et sera généré automatiquement sur le planning glissant.")
        return redirect(url_for("admin_settings"))
    templates = CourseTemplate.query.order_by(CourseTemplate.weekday, CourseTemplate.start_time).all()
    return render_template_string(TEMPLATE_SETTINGS, templates=templates, coaches=configured_coach_rows(), replacement_coaches=get_replacement_coaches(), planning_weekdays=get_coach_planning_weekdays(), weekday_labels=WEEKDAY_LABELS, subscription_prices=get_subscription_prices(), subscription_price_matrix=get_subscription_price_matrix(), member_profile_labels=MEMBER_PROFILE_LABELS, annual_membership_fee=get_annual_membership_fee(), subscription_price_key=subscription_price_key, subscription_profile_price_key=subscription_profile_price_key)


@app.route("/admin/settings/template/<int:template_id>/edit", methods=["POST"])
@login_required
def edit_template(template_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    tpl = CourseTemplate.query.get_or_404(template_id)
    old = (tpl.weekday, tpl.week_parity, tpl.course_name, tpl.start_time, tpl.end_time, tpl.coach_name)
    start = datetime.strptime(request.form["start_time"], "%H:%M").time()
    end = datetime.strptime(request.form["end_time"], "%H:%M").time()
    if end <= start:
        flash("L'heure de fin doit être après l'heure de début.")
        return redirect(url_for("admin_settings"))
    tpl.weekday = int(request.form["weekday"])
    tpl.week_parity = request.form.get("week_parity", "all")
    tpl.course_name = request.form["course_name"].strip()
    tpl.start_time = start
    tpl.end_time = end
    tpl.capacity = int(request.form.get("capacity", 35))
    tpl.waitlist_capacity = int(request.form.get("waitlist_capacity", 5))
    tpl.coach_name = request.form.get("coach_name", "").strip()
    tpl.is_reservable = request.form.get("is_reservable") == "on"

    today = date.today()
    future_sessions = CourseSession.query.filter(
        CourseSession.course_date >= today,
        CourseSession.course_name == old[2],
        CourseSession.start_time == old[3],
        CourseSession.end_time == old[4],
        CourseSession.coach_name == old[5],
    ).all()
    for session in future_sessions:
        iso_week = session.course_date.isocalendar().week
        old_parity = "even" if iso_week % 2 == 0 else "odd"
        if session.course_date.weekday() == old[0] and (old[1] == "all" or old[1] == old_parity) and not session.bookings:
            session.course_name = tpl.course_name
            session.start_time = tpl.start_time
            session.end_time = tpl.end_time
            session.capacity = tpl.capacity
            session.waitlist_capacity = tpl.waitlist_capacity
            session.coach_name = tpl.coach_name
            session.is_reservable = tpl.is_reservable
    db.session.commit()
    generate_rolling_sessions(days_ahead=28)
    flash("Cours modifié.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/template/<int:template_id>/toggle")
@login_required
def toggle_template(template_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    tpl = CourseTemplate.query.get_or_404(template_id)
    tpl.active = not tpl.active
    db.session.commit()
    flash("Créneau type mis à jour.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/template/<int:template_id>/delete")
@login_required
def delete_template(template_id):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    tpl = CourseTemplate.query.get_or_404(template_id)
    today = date.today()
    future_sessions = CourseSession.query.filter(
        CourseSession.course_date >= today,
        CourseSession.course_name == tpl.course_name,
        CourseSession.start_time == tpl.start_time,
        CourseSession.end_time == tpl.end_time,
        CourseSession.coach_name == tpl.coach_name,
    ).all()
    for session in future_sessions:
        iso_week = session.course_date.isocalendar().week
        parity = "even" if iso_week % 2 == 0 else "odd"
        if session.course_date.weekday() == tpl.weekday and (tpl.week_parity == "all" or tpl.week_parity == parity) and not session.bookings:
            db.session.delete(session)
    db.session.delete(tpl)
    db.session.commit()
    flash("Cours supprimé. Les séances futures sans réservation ont été retirées.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/coach/<path:coach_name>/delete")
@login_required
def delete_settings_coach(coach_name):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    today = date.today()
    for tpl in CourseTemplate.query.filter_by(coach_name=coach_name).all():
        db.session.delete(tpl)
    for session in CourseSession.query.filter(CourseSession.course_date >= today, CourseSession.coach_name == coach_name).all():
        if not session.bookings:
            db.session.delete(session)
    coach_user = User.query.filter_by(role="coach", full_name=coach_name).first()
    if coach_user:
        db.session.delete(coach_user)
    save_replacement_coaches([name for name in get_replacement_coaches() if name != coach_name])
    db.session.commit()
    flash("Prof supprimé des cours paramétrés.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/coach/add", methods=["POST"])
@login_required
def add_settings_coach():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    email = request.form.get("email", "").strip().lower()
    full_name = request.form.get("full_name", "").strip() or email
    coach_type = request.form.get("coach_type", "titulaire")
    if coach_type not in {"titulaire", "remplacant"}:
        coach_type = "titulaire"
    if not email:
        flash("Merci d'indiquer l'email du coach.")
        return redirect(url_for("admin_settings"))
    existing = User.query.filter_by(email=email).first()
    if existing:
        existing.role = "coach"
        existing.full_name = full_name
        existing.coach_type = coach_type
        existing.account_status = "pending"
        existing.set_password(secrets.token_urlsafe(12))
        user = existing
    else:
        user = User(email=email, role="coach", status="autre", full_name=full_name, account_status="pending", coach_type=coach_type)
        user.set_password(secrets.token_urlsafe(12))
        db.session.add(user)
    replacements = [name for name in get_replacement_coaches() if name != full_name]
    if coach_type == "remplacant":
        replacements.append(full_name)
    save_replacement_coaches(replacements)
    db.session.commit()
    send_activation_email(user)
    flash("Coach ajouté. Un lien unique de création de mot de passe a été envoyé si le SMTP est configuré.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/coach/<path:coach_name>/update", methods=["POST"])
@login_required
def update_settings_coach(coach_name):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    new_name = request.form.get("coach_name", coach_name).strip()
    new_email = valid_email(request.form.get("coach_email", ""))
    coach_type = request.form.get("coach_type", "titulaire")
    if coach_type not in {"titulaire", "remplacant"}:
        coach_type = "titulaire"
    if not new_name:
        flash("Merci d'indiquer un nom de coach.")
        return redirect(url_for("admin_settings"))

    user = User.query.filter(User.role == "coach").filter((User.full_name == coach_name) | (User.email == coach_name)).first()
    send_link = False
    if new_email:
        email_owner = User.query.filter_by(email=new_email).first()
        if email_owner and (not user or email_owner.id != user.id):
            flash("Cet email est déjà utilisé par un autre compte.")
            return redirect(url_for("admin_settings"))
        if not user:
            user = User(email=new_email, role="coach", full_name=new_name, account_status="pending", coach_type=coach_type)
            user.set_password(secrets.token_urlsafe(12))
            db.session.add(user)
            send_link = True
        elif user.email != new_email:
            user.account_status = "pending"
            user.set_password(secrets.token_urlsafe(12))
            send_link = True
    if user:
        user.full_name = new_name
        if new_email:
            user.email = new_email
        user.coach_type = coach_type

    for tpl in CourseTemplate.query.filter_by(coach_name=coach_name).all():
        tpl.coach_name = new_name
    for session in CourseSession.query.filter(CourseSession.course_date >= date.today(), CourseSession.coach_name == coach_name).all():
        if not session.bookings:
            session.coach_name = new_name

    replacements = [name for name in get_replacement_coaches() if name not in {coach_name, new_name}]
    if coach_type == "remplacant":
        replacements.append(new_name)
    save_replacement_coaches(replacements)
    db.session.commit()
    if send_link and user:
        send_activation_email(user)
        flash("Coach mis à jour. Un lien unique de création de mot de passe a été envoyé si le SMTP est configuré.")
    else:
        flash("Coach mis à jour.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/replacement-coach/add", methods=["POST"])
@login_required
def add_replacement_coach():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    name = request.form.get("replacement_name", "").strip()
    if not name:
        flash("Merci d'indiquer le nom du remplaçant.")
        return redirect(url_for("admin_settings"))
    names = get_replacement_coaches()
    names.append(name)
    user = User.query.filter(User.role == "coach").filter((User.full_name == name) | (User.email == name)).first()
    if user:
        user.coach_type = "remplacant"
    save_replacement_coaches(names)
    db.session.commit()
    flash("Coach remplaçant ajouté.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/replacement-coach/<path:coach_name>/delete")
@login_required
def delete_replacement_coach(coach_name):
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    save_replacement_coaches([name for name in get_replacement_coaches() if name != coach_name])
    user = User.query.filter(User.role == "coach").filter((User.full_name == coach_name) | (User.email == coach_name)).first()
    if user:
        user.coach_type = "titulaire"
    db.session.commit()
    flash("Coach remplaçant supprimé.")
    return redirect(url_for("admin_settings"))


@app.route("/admin/budget", methods=["GET", "POST"])
@login_required
def admin_budget():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    dues_year = int(request.args.get("dues_year", date.today().year))
    try:
        db.create_all()
        if request.method == "POST":
            entry_date = datetime.strptime(request.form["entry_date"], "%Y-%m-%d").date()
            amount = float(request.form["amount"].replace(",", "."))
            entry = BudgetEntry(
                entry_date=entry_date,
                entry_type=request.form.get("entry_type", "expense"),
                category=request.form.get("category", "Autre"),
                label=request.form.get("label", "").strip() or "Ligne budget",
                amount=amount,
                notes=request.form.get("notes", "").strip(),
            )
            db.session.add(entry)
            db.session.commit()
            flash("Ligne budget ajoutée.")
            return redirect(url_for("admin_budget"))
        entry_objects = BudgetEntry.query.order_by(BudgetEntry.entry_date.desc(), BudgetEntry.id.desc()).all()
        entries = budget_entry_rows_plain(entry_objects)
    except Exception:
        db.session.rollback()
        print("\n--- ERREUR BUDGET SIMPLE ---")
        traceback.print_exc()
        print("----------------------------\n")
        entries = []
        flash("Budget ouvert en mode simplifié : anciennes données temporairement ignorées.")
    budget_warning = None
    try:
        dues_rows = budget_due_rows_plain(dues_year)
    except Exception:
        db.session.rollback()
        print("\n--- ERREUR COTISATIONS BUDGET ---")
        traceback.print_exc()
        print("---------------------------------\n")
        dues_rows = []
        budget_warning = "Les lignes cotisations / abonnements n'ont pas pu être calculées. Les lignes manuelles restent disponibles."
        flash(budget_warning)
    expected_dues = sum((row.get("total") or 0) for row in dues_rows)
    manual_income = sum((e.get("amount") or 0) for e in entries if e.get("entry_type") == "income")
    income = manual_income + expected_dues
    expenses = sum((e.get("amount") or 0) for e in entries if e.get("entry_type") == "expense")
    balance = income - expenses
    try:
        return render_template_string(TEMPLATE_BUDGET_SIMPLE, entries=entries, dues_rows=dues_rows, dues_year=dues_year, expected_dues=expected_dues, manual_income=manual_income, income=income, expenses=expenses, balance=balance, today=date.today(), budget_warning=budget_warning)
    except Exception:
        print("\n--- ERREUR RENDU BUDGET ---")
        traceback.print_exc()
        print("---------------------------\n")
        return render_template_string(TEMPLATE_BUDGET_MINIMAL, entries=entries, income=manual_income, expenses=expenses, balance=manual_income - expenses, today=date.today())


@app.route("/admin/budget/dues/export")
@login_required
def export_budget_dues():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    dues_year = int(request.args.get("dues_year", date.today().year))
    dues_rows = budget_due_rows_plain(dues_year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotisations attendues"
    ws.append(["Année", "Nom", "Email", "Profil", "Abonnement", "Tarif abonnement", "Cotisation annuelle", "Total attendu", "Première inscription"])
    for row in dues_rows:
        ws.append([
            dues_year,
            row["name"],
            row["email"],
            row["member_profile_label"],
            f"{row['subscription_type'] or ''} {row['subscription_year'] or ''}".strip(),
            row["subscription_price"],
            row["annual_fee"],
            row["total"],
            "Oui" if row["annual_fee"] else "Non",
        ])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return send_file(file, as_attachment=True, download_name=f"cotisations_attendues_{dues_year}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/inventory", methods=["GET", "POST"])
@login_required
def admin_inventory():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    ensure_inventory_schema()
    if request.method == "POST":
        try:
            invoice_file = save_inventory_document(request.files.get("invoice_file"), "facture")
            purchase_request_file = save_inventory_document(request.files.get("purchase_request_file"), "demande_cse")
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("admin_inventory"))
        item = InventoryItem(name=request.form["name"].strip(), category=request.form.get("category", "").strip(), quantity=int(request.form.get("quantity", 0)), alert_threshold=int(request.form.get("alert_threshold", 1)), unit_cost=float((request.form.get("unit_cost") or "0").replace(",", ".")), acquisition_year=int(request.form.get("acquisition_year") or date.today().year), invoice_file=invoice_file, purchase_request_file=purchase_request_file, notes=request.form.get("notes", "").strip(), updated_at=datetime.utcnow())
        db.session.add(item)
        db.session.commit()
        flash("Article ajouté à l'inventaire.")
        return redirect(url_for("admin_inventory"))
    items = InventoryItem.query.order_by(InventoryItem.category, InventoryItem.name).all()
    inventory_value = sum((i.quantity or 0) * (i.unit_cost or 0) for i in items)
    return render_template_string(TEMPLATE_INVENTORY, items=items, inventory_value=inventory_value, current_year=date.today().year)


@app.route("/admin/archives/members")
@login_required
def archived_members():
    if not is_admin():
        flash("Accès réservé à l’admin.")
        return redirect(url_for("index"))
    query = User.query.filter(User.role == "adherent", User.account_status == "archived")
    search = request.args.get("search", "").strip()
    archived_year = request.args.get("archived_year", "").strip()
    subscription_year = request.args.get("subscription_year", "").strip()
    if search:
        like = f"%{search.lower()}%"
        query = query.filter(db.or_(db.func.lower(User.full_name).like(like), db.func.lower(User.first_name).like(like), db.func.lower(User.last_name).like(like), db.func.lower(User.email).like(like), db.func.lower(User.member_number).like(like)))
    if archived_year.isdigit():
        archive_start = date(int(archived_year), 1, 1)
        archive_end = date(int(archived_year), 12, 31)
        query = query.filter(User.archived_at >= archive_start, User.archived_at <= archive_end)
    if subscription_year.isdigit():
        query = query.filter(User.subscription_year == int(subscription_year))
    users = query.order_by(User.archived_at.desc(), User.full_name).all()
    filter_values = {"search": search, "archived_year": archived_year, "subscription_year": subscription_year}
    return render_template_string(TEMPLATE_ARCHIVED_MEMBERS, users=users, filter_values=filter_values)


def ensure_inventory_schema():
    db.create_all()
    if db.engine.dialect.name == "postgresql":
        inventory_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'inventory_item'
        """)).fetchall()}
        inventory_additions = {
            "acquisition_year": "ALTER TABLE inventory_item ADD COLUMN acquisition_year INTEGER",
            "invoice_file": "ALTER TABLE inventory_item ADD COLUMN invoice_file VARCHAR(255)",
            "purchase_request_file": "ALTER TABLE inventory_item ADD COLUMN purchase_request_file VARCHAR(255)",
        }
        for col, sql in inventory_additions.items():
            if col not in inventory_columns:
                db.session.execute(db.text(sql))
        db.session.commit()
        return
    if db.engine.dialect.name != "sqlite":
        return
    inventory_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(inventory_item)")).fetchall()}
    inventory_additions = {
        "acquisition_year": "ALTER TABLE inventory_item ADD COLUMN acquisition_year INTEGER",
        "invoice_file": "ALTER TABLE inventory_item ADD COLUMN invoice_file VARCHAR(255)",
        "purchase_request_file": "ALTER TABLE inventory_item ADD COLUMN purchase_request_file VARCHAR(255)",
    }
    for col, sql in inventory_additions.items():
        if col not in inventory_columns:
            db.session.execute(db.text(sql))
    db.session.commit()


def ensure_coach_absence_schema():
    db.create_all()
    if db.engine.dialect.name == "postgresql":
        absence_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'coach_absence'
        """)).fetchall()}
        absence_additions = {
            "session_id": "ALTER TABLE coach_absence ADD COLUMN session_id INTEGER",
            "followup_status": "ALTER TABLE coach_absence ADD COLUMN followup_status VARCHAR(30) DEFAULT 'a_traiter' NOT NULL",
            "admin_notes": "ALTER TABLE coach_absence ADD COLUMN admin_notes VARCHAR(500)",
            "reviewed_at": "ALTER TABLE coach_absence ADD COLUMN reviewed_at TIMESTAMP",
            "reviewed_by": "ALTER TABLE coach_absence ADD COLUMN reviewed_by VARCHAR(150)",
        }
        for col, sql in absence_additions.items():
            if col not in absence_columns:
                db.session.execute(db.text(sql))
        booking_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'booking'
        """)).fetchall()}
        if "attendance_status" not in booking_columns:
            db.session.execute(db.text("ALTER TABLE booking ADD COLUMN attendance_status VARCHAR(30)"))
        db.session.commit()
        return
    if db.engine.dialect.name != "sqlite":
        return
    absence_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(coach_absence)")).fetchall()}
    absence_additions = {
        "session_id": "ALTER TABLE coach_absence ADD COLUMN session_id INTEGER",
        "followup_status": "ALTER TABLE coach_absence ADD COLUMN followup_status VARCHAR(30) DEFAULT 'a_traiter' NOT NULL",
        "admin_notes": "ALTER TABLE coach_absence ADD COLUMN admin_notes VARCHAR(500)",
        "reviewed_at": "ALTER TABLE coach_absence ADD COLUMN reviewed_at DATETIME",
        "reviewed_by": "ALTER TABLE coach_absence ADD COLUMN reviewed_by VARCHAR(150)",
    }
    for col, sql in absence_additions.items():
        if col not in absence_columns:
            db.session.execute(db.text(sql))
    db.session.commit()


def ensure_useful_documents_schema():
    db.create_all()
    if db.engine.dialect.name == "postgresql":
        document_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'useful_document'
        """)).fetchall()}
        document_additions = {
            "title": "ALTER TABLE useful_document ADD COLUMN title VARCHAR(150)",
            "category": "ALTER TABLE useful_document ADD COLUMN category VARCHAR(80)",
            "file_path": "ALTER TABLE useful_document ADD COLUMN file_path VARCHAR(255)",
            "notes": "ALTER TABLE useful_document ADD COLUMN notes VARCHAR(500)",
            "uploaded_at": "ALTER TABLE useful_document ADD COLUMN uploaded_at TIMESTAMP",
            "uploaded_by": "ALTER TABLE useful_document ADD COLUMN uploaded_by VARCHAR(150)",
        }
        for col, sql in document_additions.items():
            if col not in document_columns:
                db.session.execute(db.text(sql))
        db.session.commit()
        return
    if db.engine.dialect.name != "sqlite":
        return
    document_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(useful_document)")).fetchall()}
    document_additions = {
        "title": "ALTER TABLE useful_document ADD COLUMN title VARCHAR(150)",
        "category": "ALTER TABLE useful_document ADD COLUMN category VARCHAR(80)",
        "file_path": "ALTER TABLE useful_document ADD COLUMN file_path VARCHAR(255)",
        "notes": "ALTER TABLE useful_document ADD COLUMN notes VARCHAR(500)",
        "uploaded_at": "ALTER TABLE useful_document ADD COLUMN uploaded_at DATETIME",
        "uploaded_by": "ALTER TABLE useful_document ADD COLUMN uploaded_by VARCHAR(150)",
    }
    for col, sql in document_additions.items():
        if col not in document_columns:
            db.session.execute(db.text(sql))
    db.session.commit()


def backfill_membership_tariff_snapshots():
    periods = MembershipPeriod.query.join(User).filter(
        db.or_(
            MembershipPeriod.subscription_price_snapshot.is_(None),
            MembershipPeriod.annual_fee_snapshot.is_(None),
            MembershipPeriod.total_snapshot.is_(None),
        )
    ).all()
    changed = False
    for period in periods:
        subscription_price, annual_fee, total = membership_tariff_snapshot(period.user, period.subscription_type, period.annual_fee_applies)
        if period.subscription_price_snapshot is None:
            period.subscription_price_snapshot = subscription_price
        if period.annual_fee_snapshot is None:
            period.annual_fee_snapshot = annual_fee
        if period.total_snapshot is None:
            period.total_snapshot = (period.subscription_price_snapshot or 0.0) + (period.annual_fee_snapshot or 0.0)
        if period.tariff_snapshot_at is None:
            period.tariff_snapshot_at = period.created_at or datetime.utcnow()
        changed = True
    if changed:
        db.session.commit()


def ensure_membership_period_schema():
    db.create_all()
    if db.engine.dialect.name == "postgresql":
        columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'membership_period'
        """)).fetchall()}
        additions = {
            "subscription_type": "ALTER TABLE membership_period ADD COLUMN subscription_type VARCHAR(50) DEFAULT 'Annuel' NOT NULL",
            "subscription_year": "ALTER TABLE membership_period ADD COLUMN subscription_year INTEGER DEFAULT 2026 NOT NULL",
            "start_date": "ALTER TABLE membership_period ADD COLUMN start_date DATE DEFAULT CURRENT_DATE NOT NULL",
            "end_date": "ALTER TABLE membership_period ADD COLUMN end_date DATE DEFAULT CURRENT_DATE NOT NULL",
            "annual_fee_applies": "ALTER TABLE membership_period ADD COLUMN annual_fee_applies BOOLEAN DEFAULT FALSE NOT NULL",
            "subscription_price_snapshot": "ALTER TABLE membership_period ADD COLUMN subscription_price_snapshot DOUBLE PRECISION",
            "annual_fee_snapshot": "ALTER TABLE membership_period ADD COLUMN annual_fee_snapshot DOUBLE PRECISION",
            "total_snapshot": "ALTER TABLE membership_period ADD COLUMN total_snapshot DOUBLE PRECISION",
            "tariff_snapshot_at": "ALTER TABLE membership_period ADD COLUMN tariff_snapshot_at TIMESTAMP",
            "created_at": "ALTER TABLE membership_period ADD COLUMN created_at TIMESTAMP",
            "created_by": "ALTER TABLE membership_period ADD COLUMN created_by VARCHAR(150)",
            "notes": "ALTER TABLE membership_period ADD COLUMN notes VARCHAR(500)",
        }
        for col, sql in additions.items():
            if col not in columns:
                db.session.execute(db.text(sql))
        db.session.commit()
        backfill_membership_tariff_snapshots()
        return
    if db.engine.dialect.name != "sqlite":
        return
    columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(membership_period)")).fetchall()}
    additions = {
        "subscription_price_snapshot": "ALTER TABLE membership_period ADD COLUMN subscription_price_snapshot FLOAT",
        "annual_fee_snapshot": "ALTER TABLE membership_period ADD COLUMN annual_fee_snapshot FLOAT",
        "total_snapshot": "ALTER TABLE membership_period ADD COLUMN total_snapshot FLOAT",
        "tariff_snapshot_at": "ALTER TABLE membership_period ADD COLUMN tariff_snapshot_at DATETIME",
    }
    for col, sql in additions.items():
        if col not in columns:
            db.session.execute(db.text(sql))
    db.session.commit()
    backfill_membership_tariff_snapshots()


def ensure_schema():
    db.create_all()
    if db.engine.dialect.name == "postgresql":
        user_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'user'
        """)).fetchall()}
        postgres_user_additions = {
            "full_name": "ALTER TABLE \"user\" ADD COLUMN full_name VARCHAR(150)",
            "first_name": "ALTER TABLE \"user\" ADD COLUMN first_name VARCHAR(80)",
            "last_name": "ALTER TABLE \"user\" ADD COLUMN last_name VARCHAR(80)",
            "profile_photo": "ALTER TABLE \"user\" ADD COLUMN profile_photo VARCHAR(255)",
            "profile_photo_data": "ALTER TABLE \"user\" ADD COLUMN profile_photo_data TEXT",
            "profile_photo_mime": "ALTER TABLE \"user\" ADD COLUMN profile_photo_mime VARCHAR(80)",
            "subscription_type": "ALTER TABLE \"user\" ADD COLUMN subscription_type VARCHAR(50)",
            "subscription_year": "ALTER TABLE \"user\" ADD COLUMN subscription_year INTEGER",
            "member_profile": "ALTER TABLE \"user\" ADD COLUMN member_profile VARCHAR(30)",
            "rights_holder_name": "ALTER TABLE \"user\" ADD COLUMN rights_holder_name VARCHAR(150)",
            "member_number": "ALTER TABLE \"user\" ADD COLUMN member_number VARCHAR(30)",
            "member_card": "ALTER TABLE \"user\" ADD COLUMN member_card VARCHAR(255)",
            "blocked_at": "ALTER TABLE \"user\" ADD COLUMN blocked_at DATE",
            "blocked_reason": "ALTER TABLE \"user\" ADD COLUMN blocked_reason VARCHAR(255)",
            "preferred_course": "ALTER TABLE \"user\" ADD COLUMN preferred_course VARCHAR(100)",
            "preferred_coach": "ALTER TABLE \"user\" ADD COLUMN preferred_coach VARCHAR(150)",
            "preferred_slot": "ALTER TABLE \"user\" ADD COLUMN preferred_slot VARCHAR(80)",
            "admin_role": "ALTER TABLE \"user\" ADD COLUMN admin_role VARCHAR(50)",
            "account_status": "ALTER TABLE \"user\" ADD COLUMN account_status VARCHAR(30) DEFAULT 'active' NOT NULL",
            "activation_token": "ALTER TABLE \"user\" ADD COLUMN activation_token VARCHAR(255)",
            "activation_expires_at": "ALTER TABLE \"user\" ADD COLUMN activation_expires_at TIMESTAMP",
            "subscription_end_date": "ALTER TABLE \"user\" ADD COLUMN subscription_end_date DATE",
            "archived_at": "ALTER TABLE \"user\" ADD COLUMN archived_at DATE",
            "archived_reason": "ALTER TABLE \"user\" ADD COLUMN archived_reason VARCHAR(255)",
            "created_at": "ALTER TABLE \"user\" ADD COLUMN created_at TIMESTAMP",
            "coach_type": "ALTER TABLE \"user\" ADD COLUMN coach_type VARCHAR(30) DEFAULT 'titulaire' NOT NULL",
        }
        for col, sql in postgres_user_additions.items():
            if col not in user_columns:
                db.session.execute(db.text(sql))
        course_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'course_session'
        """)).fetchall()}
        course_additions = {
            "coach_name": "ALTER TABLE course_session ADD COLUMN coach_name VARCHAR(150)",
            "is_reservable": "ALTER TABLE course_session ADD COLUMN is_reservable BOOLEAN DEFAULT TRUE NOT NULL",
            "waitlist_capacity": "ALTER TABLE course_session ADD COLUMN waitlist_capacity INTEGER DEFAULT 5 NOT NULL",
        }
        for col, sql in course_additions.items():
            if col not in course_columns:
                db.session.execute(db.text(sql))
        template_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'course_template'
        """)).fetchall()}
        template_additions = {
            "is_reservable": "ALTER TABLE course_template ADD COLUMN is_reservable BOOLEAN DEFAULT TRUE NOT NULL",
            "waitlist_capacity": "ALTER TABLE course_template ADD COLUMN waitlist_capacity INTEGER DEFAULT 5 NOT NULL",
        }
        for col, sql in template_additions.items():
            if col not in template_columns:
                db.session.execute(db.text(sql))
        booking_columns = {row[0] for row in db.session.execute(db.text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'booking'
        """)).fetchall()}
        booking_additions = {
            "archived": "ALTER TABLE booking ADD COLUMN archived BOOLEAN DEFAULT FALSE NOT NULL",
            "attendance_status": "ALTER TABLE booking ADD COLUMN attendance_status VARCHAR(30)",
        }
        for col, sql in booking_additions.items():
            if col not in booking_columns:
                db.session.execute(db.text(sql))
        db.session.commit()
        seed_default_2026_tariffs_once()
        ensure_membership_period_schema()
        ensure_coach_absence_schema()
        ensure_inventory_schema()
        ensure_useful_documents_schema()
        return
    if db.engine.dialect.name != "sqlite":
        db.session.commit()
        return
    # Migration simple pour une base SQLite déjà existante.
    columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(user)")).fetchall()}
    additions = {
        "full_name": "ALTER TABLE user ADD COLUMN full_name VARCHAR(150)",
        "first_name": "ALTER TABLE user ADD COLUMN first_name VARCHAR(80)",
        "last_name": "ALTER TABLE user ADD COLUMN last_name VARCHAR(80)",
        "profile_photo": "ALTER TABLE user ADD COLUMN profile_photo VARCHAR(255)",
        "profile_photo_data": "ALTER TABLE user ADD COLUMN profile_photo_data TEXT",
        "profile_photo_mime": "ALTER TABLE user ADD COLUMN profile_photo_mime VARCHAR(80)",
        "subscription_type": "ALTER TABLE user ADD COLUMN subscription_type VARCHAR(50)",
        "subscription_year": "ALTER TABLE user ADD COLUMN subscription_year INTEGER",
        "member_profile": "ALTER TABLE user ADD COLUMN member_profile VARCHAR(30)",
        "rights_holder_name": "ALTER TABLE user ADD COLUMN rights_holder_name VARCHAR(150)",
        "member_number": "ALTER TABLE user ADD COLUMN member_number VARCHAR(30)",
        "member_card": "ALTER TABLE user ADD COLUMN member_card VARCHAR(255)",
        "blocked_at": "ALTER TABLE user ADD COLUMN blocked_at DATE",
        "blocked_reason": "ALTER TABLE user ADD COLUMN blocked_reason VARCHAR(255)",
        "preferred_course": "ALTER TABLE user ADD COLUMN preferred_course VARCHAR(100)",
        "preferred_coach": "ALTER TABLE user ADD COLUMN preferred_coach VARCHAR(150)",
        "preferred_slot": "ALTER TABLE user ADD COLUMN preferred_slot VARCHAR(80)",
        "admin_role": "ALTER TABLE user ADD COLUMN admin_role VARCHAR(50)",
        "account_status": "ALTER TABLE user ADD COLUMN account_status VARCHAR(30) DEFAULT 'active' NOT NULL",
        "activation_token": "ALTER TABLE user ADD COLUMN activation_token VARCHAR(255)",
        "activation_expires_at": "ALTER TABLE user ADD COLUMN activation_expires_at DATETIME",
        "subscription_end_date": "ALTER TABLE user ADD COLUMN subscription_end_date DATE",
        "archived_at": "ALTER TABLE user ADD COLUMN archived_at DATE",
        "archived_reason": "ALTER TABLE user ADD COLUMN archived_reason VARCHAR(255)",
        "created_at": "ALTER TABLE user ADD COLUMN created_at DATETIME",
        "coach_type": "ALTER TABLE user ADD COLUMN coach_type VARCHAR(30) DEFAULT 'titulaire' NOT NULL",
    }
    for col, sql in additions.items():
        if col not in columns:
            db.session.execute(db.text(sql))

    course_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(course_session)")).fetchall()}
    if "coach_name" not in course_columns:
        db.session.execute(db.text("ALTER TABLE course_session ADD COLUMN coach_name VARCHAR(150)"))
    if "is_reservable" not in course_columns:
        db.session.execute(db.text("ALTER TABLE course_session ADD COLUMN is_reservable BOOLEAN NOT NULL DEFAULT 1"))
    if "waitlist_capacity" not in course_columns:
        db.session.execute(db.text("ALTER TABLE course_session ADD COLUMN waitlist_capacity INTEGER NOT NULL DEFAULT 5"))

    template_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(course_template)")).fetchall()}
    if "is_reservable" not in template_columns:
        db.session.execute(db.text("ALTER TABLE course_template ADD COLUMN is_reservable BOOLEAN NOT NULL DEFAULT 1"))
    if "waitlist_capacity" not in template_columns:
        db.session.execute(db.text("ALTER TABLE course_template ADD COLUMN waitlist_capacity INTEGER NOT NULL DEFAULT 5"))

    booking_columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(booking)")).fetchall()}
    if "archived" not in booking_columns:
        db.session.execute(db.text("ALTER TABLE booking ADD COLUMN archived BOOLEAN NOT NULL DEFAULT 0"))
    if "attendance_status" not in booking_columns:
        db.session.execute(db.text("ALTER TABLE booking ADD COLUMN attendance_status VARCHAR(30)"))

    ensure_coach_absence_schema()
    ensure_inventory_schema()
    ensure_useful_documents_schema()
    seed_default_2026_tariffs_once()
    ensure_membership_period_schema()
    db.session.commit()

    # Renseigner les fins d'abonnement manquantes dans les anciennes bases.
    for user in User.query.filter(User.role == "adherent", User.subscription_end_date.is_(None)).all():
        if user.subscription_type and user.subscription_year:
            user.subscription_end_date = subscription_end(user.subscription_type, user.subscription_year)
    db.session.commit()


def create_default_admin():
    admin_email = os.getenv("ADMIN_EMAIL", "admin@fitness.local").strip().lower()
    if admin_email == "oceane.allard@gmail.com":
        admin_email = "admin@fitness.local"
    admin_password = os.getenv("ADMIN_PASSWORD", "admin123")
    admin_name = os.getenv("ADMIN_NAME", "Admin Fitness")
    admin = User.query.filter_by(email=admin_email).first()
    if not admin:
        admin = User(email=admin_email, role="admin", status="autre", full_name=admin_name, admin_role="presidente", account_status="active")
        admin.set_password(admin_password)
        db.session.add(admin)
        db.session.commit()
        print(f"Admin créé : {admin_email}")
    elif not admin.admin_role:
        admin.admin_role = "presidente"
        db.session.commit()

    coach = User.query.filter_by(email=DEMO_COACH_EMAIL).first()
    if not coach:
        coach = User(email=DEMO_COACH_EMAIL, role="coach", status="autre", full_name="Coach Fitness", account_status="active")
        db.session.add(coach)
    else:
        coach.role = "coach"
        coach.status = coach.status or "autre"
        coach.full_name = coach.full_name or "Coach Fitness"
        coach.account_status = "active"
    coach.set_password("coach123")
    db.session.commit()
    print("Coach démo disponible : coach@fitness.local / coach123")

    demo = User.query.filter_by(email=DEMO_ADHERENT_EMAIL).first()
    if not demo:
        demo = User(
            email=DEMO_ADHERENT_EMAIL,
            role="adherent",
            status="mensuel",
            full_name="Adhérent Démo",
            first_name="Adhérent",
            last_name="Démo",
            member_profile="ouvrant_droit",
            subscription_type="Annuel",
            subscription_year=2026,
            subscription_end_date=date(2026, 12, 31),
            account_status="active",
            member_number="DEMO-2026",
            preferred_course="Pilates",
            preferred_coach="Hayate",
            preferred_slot="Lundi midi",
        )
        db.session.add(demo)
    else:
        demo.role = "adherent"
        demo.status = "mensuel"
        demo.full_name = demo.full_name or "Adhérent Démo"
        demo.first_name = demo.first_name or "Adhérent"
        demo.last_name = demo.last_name or "Démo"
        demo.member_profile = demo.member_profile or "ouvrant_droit"
        demo.subscription_type = demo.subscription_type or "Annuel"
        demo.subscription_year = demo.subscription_year or 2026
        demo.subscription_end_date = demo.subscription_end_date or date(2026, 12, 31)
        demo.account_status = "active"
        demo.member_number = demo.member_number or "DEMO-2026"
    demo.set_password("adherent123")
    db.session.commit()
    print("Adhérent démo disponible : adherent@fitness.local / adherent123")


def start_scheduler():
    """Lance les tâches automatiques tous les jours, tant que l'application est ouverte."""
    if BackgroundScheduler is None:
        print("APScheduler non installé : lance `pip3 install apscheduler` pour activer la génération quotidienne en arrière-plan.")
        return None

    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(
        lambda: run_daily_automation(force=True),
        trigger="interval",
        days=1,
        id="fitness_daily_maintenance",
        replace_existing=True,
    )
    scheduler.start()
    print("Automatisation active : génération/archivage vérifiés chaque jour.")
    return scheduler


@app.context_processor
def inject_helpers():
    return dict(shell=shell)


if __name__ == "__main__":
    with app.app_context():
        ensure_schema()
        create_default_admin()
        seed_default_course_templates()
        run_daily_automation(force=True)
    start_scheduler()
    app.run(debug=False, use_reloader=False)
